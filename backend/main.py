import os
import uuid
import json
import shutil
import threading
import time
import logging
import signal
import sys
from typing import List, Optional
from datetime import datetime, timedelta, timezone
from contextlib import contextmanager
from functools import lru_cache
import concurrent.futures

# Suppress noisy Windows asyncio connection reset errors (harmless, client closed early)
try:
    import asyncio
    from asyncio.proactor_events import _ProactorBasePipeTransport

    def _silence_connection_reset_error(*args, **kwargs):
        try:
            return _ProactorBasePipeTransport._call_connection_lost_original(*args, **kwargs)
        except (ConnectionResetError, ConnectionAbortedError, OSError) as _err:
            # Only swallow the specific Windows errors 10053/10054
            if hasattr(_err, 'winerror') and _err.winerror in (10053, 10054):
                pass
            elif isinstance(_err, (ConnectionResetError, ConnectionAbortedError)):
                pass
            else:
                raise

    if not hasattr(_ProactorBasePipeTransport, '_call_connection_lost_original'):
        _ProactorBasePipeTransport._call_connection_lost_original = _ProactorBasePipeTransport._call_connection_lost
        _ProactorBasePipeTransport._call_connection_lost = _silence_connection_reset_error
except Exception:
    pass

import duckdb
import pandas as pd
from fastapi import FastAPI, UploadFile, File, Form, Request, HTTPException, BackgroundTasks, Query, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.exceptions import RequestValidationError
from pydantic import ValidationError
import traceback

# ---- Smart Error Helper ----
MASTER_ERROR_MAP = {
    "no_files": {
        "reason": "No source files found in selected folder",
        "suggestion": "Upload files to this folder first, then try again"
    },
    "no_files_merged": {
        "reason": "No files could be merged",
        "suggestion": "Check file formats and column names, then try again"
    },
    "missing_columns": {
        "reason_template": "Column '{col}' not found in source files",
        "suggestion": "Check available columns in file details or type 'All' to use all columns"
    },
    "db_connection": {
        "reason": "Database connection failed",
        "suggestion": "Restart the application or check server logs"
    },
    "network_timeout": {
        "reason": "Request timed out after 30s",
        "suggestion": "Try with fewer files or check your network connection"
    },
    "no_folder_selected": {
        "reason": "No folder selected",
        "suggestion": "Select a folder from the left sidebar first"
    },
    "master_exists": {
        "reason": "Master file already exists for this folder",
        "suggestion": "Delete the existing master file first, or view it in Master View"
    },
    "invalid_folder_id": {
        "reason": "Invalid folder ID provided",
        "suggestion": "Refresh the page and select a folder again"
    },
    "master_not_found": {
        "reason": "Master file not found",
        "suggestion": "Create a master file first using the 'Create Master File' button"
    },
    "multiple_sheets": {
        "reason_template": "File '{file}' has multiple sheets",
        "suggestion": "Ensure all files have only one sheet, or split multi-sheet files"
    },
    "unknown": {
        "reason": "An unexpected error occurred",
        "suggestion": "Please try again or contact support if the issue persists"
    }
}

def get_error_response(error_key="unknown", extra_context=None, status_code=500):
    """Build a structured error response with reason and suggestion."""
    mapping = MASTER_ERROR_MAP.get(error_key, MASTER_ERROR_MAP["unknown"])
    reason = mapping.get("reason_template", mapping["reason"])
    if extra_context and "{col}" in reason:
        reason = reason.replace("{col}", str(extra_context.get("col", "")))
    if extra_context and "{file}" in reason:
        reason = reason.replace("{file}", str(extra_context.get("file", "")))
    return JSONResponse(
        status_code=status_code,
        content={"detail": reason, "suggestion": mapping["suggestion"]}
    )
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

def excel_column_to_number(col_letter: str) -> int:
    """Convert Excel column letter to number (A=1, Z=26, AA=27, ZZ=702)"""
    result = 0
    for ch in col_letter.upper():
        result = result * 26 + (ord(ch) - 64)
    return result

def sort_excel_columns(columns_dict: dict) -> list:
    """
    Sort output columns by Excel column letter order (A, B, C... Z, AA, AB... ZZ).
    Returns list of (column_letter, column_name) tuples in proper sequence.
    """
    items = list(columns_dict.items())
    items.sort(key=lambda x: excel_column_to_number(x[0]))
    return items

# ---- Matplotlib setup (conditional import for robustness) ----
# Force non-interactive backend BEFORE any pyplot import.
# This prevents "Failed to generate chart" errors on Windows with no display.
_MATPLOTLIB_AVAILABLE = False
try:
    import matplotlib
    matplotlib.use('Agg', force=True)
    _MATPLOTLIB_AVAILABLE = True
except ImportError:
    logging.warning("matplotlib not installed - chart generation will be disabled")

# Detect PyInstaller/frozen runtime for correct paths FIRST
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Setup logging for production
LOG_DIR = os.path.join(BASE_DIR, '..', 'data', 'logs')
os.makedirs(LOG_DIR, exist_ok=True)
log_file = os.path.join(LOG_DIR, f"server_{datetime.now().strftime('%Y%m%d')}.log")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger("reconciliation_tool")

sys.path.insert(0, BASE_DIR)

from database import (
    init_db,
    create_folder, get_folders, delete_folder,
    save_file_metadata, get_files_by_folder, delete_file, move_file,
    save_master_file, get_master_file, delete_master_file,
    save_rule, get_rules_by_phase, get_all_rules, delete_rule,
    get_db_connection, get_master_formulas, update_master_formulas,
    move_to_recycle_bin, get_recycle_bin_items, get_recycle_bin_item,
    restore_from_recycle_bin, permanent_delete_from_recycle_bin,
    get_physical_storage_path,
    get_company_storage_path, get_user_folder_path,
    format_storage_error, is_valid_folder_name,
    # Activity Window (master_activities) helpers
    list_master_activities, get_master_activity, create_master_activity,
    update_master_activity, delete_master_activity, reorder_master_activities,
    migrate_legacy_master_formulas,
    # Dedup / duplicate-detection (concat) helpers
    get_dedup_config, save_dedup_config,
    save_rejected_artefact, list_rejected_artefacts,
    get_latest_rejected_artefact_for_file, get_rejected_artefact_by_id,
)

# Concat-based duplicate detection engine
from backend.dedup_engine import (
    is_dedup_active, resolve_dedup_columns, build_concat,
    normalize_concat, load_existing_concat_set, detect_duplicate_rows,
    ensure_dedup_table, populate_dedup_table, write_rejected_artefact,
    fuzzy_match_column, resolve_columns_fuzzy,
    DEDUP_COL_NAME,
)

from primary_data import generate_primary_data, preview_primary_data, list_primary_files, get_primary_file_path, read_primary_file, get_primary_field_columns
from filename_parser import parse_filename, generate_processed_filename, get_storage_path
from formula_engine import parse_formula, validate_formula, FormulaError
from database import create_activity_from_action as _create_activity_from_action

# Dummy dependencies to bypass authentication
async def get_current_active_user(): return {"id": 1, "user_id": 1, "company_id": 1, "role": "admin"}
def require_role(role: str):
    async def role_checker(): return True
    return role_checker
def require_page_permission(page: str):
    async def permission_checker(): return True
    return permission_checker
async def is_first_login_required(): return False
async def get_company_id(): return 1
async def get_module_id(): return 1
async def get_optional_user(request: Request):
    module_id = request.headers.get("X-Module-ID", 1)
    try:
        module_id = int(module_id)
    except:
        module_id = 1
    return {"id": 1, "user_id": 1, "company_id": 1, "module_id": module_id, "role": "admin"}

app = FastAPI(title="Reconciliation Tool", version="3.0.0")

# Local Modules APIs
@app.get("/api/local-modules")
async def get_local_modules():
    conn = get_db_connection()
    try:
        # All modules are available locally
        modules = conn.execute("SELECT id, name, description FROM modules").fetchall()
        return {"success": True, "modules": [dict(m) for m in modules]}
    finally:
        conn.close()

from pydantic import BaseModel
class ModuleCreate(BaseModel):
    name: str
    description: str = ""

@app.post("/api/local-modules")
async def create_local_module(module: ModuleCreate):
    conn = get_db_connection()
    try:
        # code is a required column, so we just use the name as the code
        conn.execute("INSERT INTO modules (name, code, description) VALUES (?, ?, ?)", (module.name, module.name, module.description))
        conn.commit()
        module_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        return {"success": True, "module_id": module_id, "name": module.name}
    finally:
        conn.close()

@app.delete("/api/local-modules/{module_id}")
async def delete_local_module(module_id: int):
    # Prevent deleting the default module (ID 1)
    if module_id == 1:
        return JSONResponse(status_code=400, content={"success": False, "error": "Cannot delete the Website module."})
        
    conn = get_db_connection()
    try:
        # Delete the module
        conn.execute("DELETE FROM modules WHERE id = ?", (module_id,))
        conn.commit()
        return {"success": True}
    finally:
        conn.close()

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize DB on startup (ensure tables and migrations run)
@app.on_event("startup")
async def startup_event():
    try:
        import os, shutil
        db_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data', 'metadata.db')
        template_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'template.db')
        if not os.path.exists(db_path) and os.path.exists(template_path):
            os.makedirs(os.path.dirname(db_path), exist_ok=True)
            shutil.copy2(template_path, db_path)
            logger.info("Successfully initialized metadata.db from template.db!")
            
            # --- Auto-Deployment Path & Data Restoration ---
            import sqlite3
            try:
                base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                target_uploads_dir = os.path.join(base_dir, 'data', 'uploads')
                template_data_dir = os.path.join(base_dir, 'data', 'template_master_dbs')
                
                # 1. Copy Physical Master DBs
                if os.path.exists(template_data_dir):
                    shutil.copytree(template_data_dir, target_uploads_dir, dirs_exist_ok=True)
                    logger.info("Restored 10-line template physical master databases.")
                
                # 2. Re-path Absolute Database URLs
                conn = sqlite3.connect(db_path)
                cursor = conn.cursor()
                
                # Update master_files
                cursor.execute("SELECT id, db_path FROM master_files WHERE db_path IS NOT NULL")
                for m_id, path in cursor.fetchall():
                    # Support both slash types
                    norm_path = path.replace('\\', '/')
                    if '/data/uploads/' in norm_path:
                        rel_path = norm_path.split('/data/uploads/')[1]
                        new_abs_path = os.path.join(target_uploads_dir, rel_path.replace('/', os.sep))
                        conn.execute("UPDATE master_files SET db_path = ? WHERE id = ?", (new_abs_path, m_id))
                        
                # Update files (if any templates exist)
                cursor.execute("SELECT id, file_path FROM files WHERE file_path IS NOT NULL")
                for f_id, path in cursor.fetchall():
                    norm_path = path.replace('\\', '/')
                    if '/data/uploads/' in norm_path:
                        rel_path = norm_path.split('/data/uploads/')[1]
                        new_abs_path = os.path.join(target_uploads_dir, rel_path.replace('/', os.sep))
                        conn.execute("UPDATE files SET file_path = ? WHERE id = ?", (new_abs_path, f_id))
                        
                conn.commit()
                conn.close()
                logger.info("Successfully re-mapped all absolute file paths for the current environment.")
            except Exception as e:
                logger.error(f"Error during template path restoration: {e}")
            # ---------------------------------------------
            
        from database import init_db, cleanup_old_notifications
        init_db()
        logger.info("Database initialized successfully.")
        
        # Start a background thread to cleanup old notifications periodically (e.g. daily)
        def cleanup_loop():
            import time
            while True:
                try:
                    cleanup_old_notifications(30)
                except Exception as e:
                    logger.error(f"Error in cleanup loop: {e}")
                time.sleep(86400) # Sleep for 24 hours
                
        import threading
        t = threading.Thread(target=cleanup_loop, daemon=True)
        t.start()
        
    except Exception as e:
        logger.error(f"Failed to initialize database: {e}")

# Directories (use BASE_DIR already set above for frozen runtime compatibility)
UPLOAD_DIR = os.path.join(BASE_DIR, '..', 'data', 'uploads')
MASTER_DIR = os.path.join(BASE_DIR, '..', 'data', 'master_files')
PROCESSED_DIR = os.path.join(BASE_DIR, '..', 'data', 'processed')
LOG_DIR = os.path.join(BASE_DIR, '..', 'data', 'logs')
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(MASTER_DIR, exist_ok=True)
os.makedirs(PROCESSED_DIR, exist_ok=True)
os.makedirs(LOG_DIR, exist_ok=True)

# Serve frontend
FRONTEND_DIR = os.path.join(BASE_DIR, '..', 'frontend')
app.mount("/static", StaticFiles(directory=FRONTEND_DIR), name="static")

# ============ VALIDATION ERROR HANDLER (Detailed 422 Diagnostics) ============
@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request, exc):
    """
    Custom handler for FastAPI validation errors (422 Unprocessable Content).
    Returns detailed information about which fields failed and why.
    """
    errors = []
    for error in exc.errors():
        field = ".".join(str(loc) for loc in error.get("loc", []))
        msg = error.get("msg", "Unknown error")
        errors.append({
            "field": field,
            "message": msg,
            "type": error.get("type", "unknown"),
            "input": str(error.get("input", "N/A"))[:100]
        })
    
    # Log the error details for server-side debugging
    logger.warning(f"422 Validation Error: {errors} | Path: {request.url.path} | Method: {request.method}")
    
    return JSONResponse(
        status_code=422,
        content={
            "success": False,
            "status": "error",
            "message": f"Validation failed: {len(errors)} field(s) invalid",
            "detail": errors,
            "path": request.url.path,
            "method": request.method,
            "hint": "Check that all required form fields are provided with correct types"
        }
    )

@app.exception_handler(ValidationError)
async def pydantic_validation_exception_handler(request, exc):
    """Handler for Pydantic validation errors"""
    logger.warning(f"Pydantic Validation Error: {str(exc)} | Path: {request.url.path}")
    return JSONResponse(
        status_code=422,
        content={
            "success": False,
            "status": "error",
            "message": f"Pydantic validation failed: {str(exc)}",
            "path": request.url.path,
            "hint": "Check data types and required fields"
        }
    )

# ============ STORAGE ERROR HANDLER ============
@app.exception_handler(OSError)
async def storage_exception_handler(request, exc):
    """
    Global handler for filesystem errors (disk full, permission denied, etc.).
    Returns structured error with actionable suggestion.
    """
    path = getattr(exc, 'filename', None) or str(exc)
    logger.error(f"OSError on [{request.method}] {request.url.path}: {exc} (errno={exc.errno})")
    
    if exc.errno == 28:  # No space left on device
        err = format_storage_error("disk_full", {"path": path})
        return JSONResponse(status_code=507, content=err)
    elif exc.errno == 13:  # Permission denied
        err = format_storage_error("permission_denied", {"path": path, "os_error": str(exc)})
        return JSONResponse(status_code=403, content=err)
    elif exc.errno == 36:  # File name too long
        err = format_storage_error("path_too_long", {"path": path})
        return JSONResponse(status_code=400, content=err)
    else:
        err = format_storage_error("dir_create_failed", {"path": path, "os_error": str(exc)})
        return JSONResponse(status_code=500, content=err)

# Global processing status tracker
processing_status = {
    "is_processing": False,
    "result": None,
    "error": None,
    "start_time": None,
    "progress": "idle"
}
processing_lock = threading.Lock()

# ============ FILE METADATA CACHE ============
_file_metadata_cache = {}
_file_cache_lock = threading.Lock()

def get_cached_file_info(file_id):
    """Get cached file metadata to avoid repeated DB queries"""
    with _file_cache_lock:
        return _file_metadata_cache.get(file_id)

def set_cached_file_info(file_id, info):
    with _file_cache_lock:
        _file_metadata_cache[file_id] = info

def clear_file_cache():
    with _file_cache_lock:
        _file_metadata_cache.clear()

# ============ NaN CLEANUP UTILITY ============
import math

def clean_nan_values(obj):
    """
    Recursively replace NaN, Inf, -Inf float values with None.
    Required because Python's json module cannot serialize NaN values.
    """
    if isinstance(obj, dict):
        return {k: clean_nan_values(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [clean_nan_values(item) for item in obj]
    elif isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj
    return obj

# ============ FAST EXCEL INFO ============
def get_excel_info_fast(file_path):
    """
    Optimized Excel/CSV info extraction.
    Returns: (sheet_count, row_count, col_count, sheet_names_json)
    """
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == '.csv':
        try:
            # Fast CSV line count without loading full file
            total_rows = 0
            total_cols = 0
            sample = []
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                for i, line in enumerate(f):
                    stripped = line.strip()
                    if i < 5:
                        sample.append(line)
                    if stripped:  # Skip blank lines
                        total_rows += 1
            
            if sample:
                import csv
                first_row = next(csv.reader([sample[0]]))
                total_cols = len(first_row) if first_row else 0
            
            # Subtract 1 for header row
            data_rows = max(0, total_rows - 1)
            
            return 1, data_rows, total_cols, json.dumps(["Sheet1"])
        except Exception as e:
            logger.warning(f"Fast CSV parsing error: {e}. Falling back to pandas...")
            try:
                import pandas as pd
                # Fallback to pandas which handles weird encodings better
                df = pd.read_csv(file_path, on_bad_lines='skip')
                return 1, len(df), len(df.columns), json.dumps(["Sheet1"])
            except Exception as e2:
                logger.error(f"CSV info fallback error: {e2}")
                return 0, 0, 0, "[]"
    
    # Handle Excel files - optimized with read_only
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        sheet_count = len(sheet_names)
        
        total_rows = 0
        total_cols = 0
        
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            # read_only mode gives approximate max_row quickly
            rows = ws.max_row or 0
            cols = ws.max_column or 0
            total_rows += rows
            if cols > total_cols:
                total_cols = cols
        
        wb.close()
        return sheet_count, total_rows, total_cols, json.dumps(sheet_names)
    except Exception as e:
        logger.warning(f"openpyxl failed to read {file_path} ({e}). Falling back to pandas...")
        try:
            import pandas as pd
            # Use pandas ExcelFile to get sheet names and dimensions
            xl = pd.ExcelFile(file_path)
            sheet_names = xl.sheet_names
            sheet_count = len(sheet_names)
            
            total_rows = 0
            total_cols = 0
            
            for sheet in sheet_names:
                df = pd.read_excel(xl, sheet_name=sheet, nrows=0) # Just read headers
                cols = len(df.columns)
                if cols > total_cols:
                    total_cols = cols
                
                # To get row count without loading whole df into memory, we might just load it
                df_full = pd.read_excel(xl, sheet_name=sheet)
                total_rows += len(df_full)
                
            return sheet_count, total_rows, total_cols, json.dumps(sheet_names)
        except Exception as e2:
            logger.error(f"Excel info fallback error: {e2}")
            return 0, 0, 0, "[]"


# ============ OPTIMIZED AUTO-FIT (SAMPLED) ============
def _autofit_columns_fast(ws, sample_rows=100):
    """Fast column auto-fit using sampling instead of scanning all cells"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        # Sample first N cells + last cell
        cells_to_check = list(column[:sample_rows])
        if len(column) > sample_rows:
            cells_to_check.append(column[-1])
        
        for cell in cells_to_check:
            try:
                if cell.value:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


# ============ VECTORIZED EXCEL WRITING ============
def write_pivot_to_worksheet_fast(ws, pivot_df, start_row=1):
    """
    Vectorized pivot table writing to Excel worksheet.
    Professional formatting with Grand Total highlighting and number formatting.
    Ensures Grand Total row is always at the bottom with special formatting.
    """
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # Reset index to convert row fields from index to regular columns
    if not pivot_df.index.names == [None]:
        pivot_df = pivot_df.reset_index()
    
    # Identify row field columns (they come before value columns after reset_index)
    # Grand Total detection: check all columns since with MultiIndex, GT can be in any row-field column
    row_field_count = 0
    for col in pivot_df.columns:
        # Heuristic: if column name doesn't contain aggregation keywords, it's a row field
        col_str = str(col).lower()
        if not any(agg in col_str for agg in ['sum', 'count', 'avg', 'mean', 'min', 'max']):
            row_field_count += 1
        else:
            break
    
    # Write headers
    for col_idx, col_name in enumerate(pivot_df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            bottom=Side(style='medium', color='1F4E79')
        )
    
    # Track Grand Total rows for professional formatting
    grand_total_rows = set()
    
    # Pre-scan to identify Grand Total rows (last row is always Grand Total when margins=True)
    data_rows = list(dataframe_to_rows(pivot_df, index=False, header=False))
    
    # Mark last row as Grand Total if it contains 'Grand Total' anywhere
    if data_rows:
        last_row = data_rows[-1]
        last_row_str = [str(v) for v in last_row]
        if any('Grand Total' in v for v in last_row_str):
            grand_total_rows.add(start_row + len(data_rows))
    
    # Use dataframe_to_rows for bulk writing
    for row_idx, row in enumerate(data_rows, start_row + 1):
        is_grand_total = row_idx in grand_total_rows
        
        # Also check any row that has 'Grand Total' in any row-field column
        if not is_grand_total:
            for col_idx in range(min(row_field_count, len(row))):
                if 'Grand Total' in str(row[col_idx]):
                    is_grand_total = True
                    grand_total_rows.add(row_idx)
                    break
        
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            is_numeric = isinstance(value, (int, float))
            col_name = str(pivot_df.columns[col_idx - 1]) if col_idx <= len(pivot_df.columns) else ''
            
            if is_numeric:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                # Use Indian number format (lakhs/crores) for large numbers
                if abs(value) >= 100000:
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '#,##0.00'
                
                if is_grand_total:
                    cell.font = Font(bold=True, color='1F4E79', size=10)
                    cell.fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
                else:
                    if value > 0:
                        cell.font = Font(color='27AE60', size=9)
                    elif value < 0:
                        cell.font = Font(color='E74C3C', size=9)
                    else:
                        cell.font = Font(size=9)
            else:
                if is_grand_total:
                    cell.font = Font(bold=True, color='1F4E79', size=10)
                    cell.fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.font = Font(size=9, color='374151')
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Grand Total row gets top and bottom borders
            if is_grand_total:
                cell.border = Border(
                    top=Side(style='medium', color='1F4E79'),
                    bottom=Side(style='medium', color='1F4E79')
                )
            else:
                # Light border for normal rows
                cell.border = Border(
                    bottom=Side(style='thin', color='E5E7EB')
                )
    
    # Apply column widths based on content
    for col_idx in range(1, len(pivot_df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 12  # minimum width
        
        # Check header length
        header_val = str(pivot_df.columns[col_idx - 1])
        if len(header_val) > max_length:
            max_length = len(header_val)
        
        ws.column_dimensions[col_letter].width = min(max_length + 3, 30)


def _get_context(current_user):
    """Extract company_id and module_id from optional auth context."""
    if current_user:
        return current_user.get("company_id"), current_user.get("module_id")
    return None, None


# ============ HEALTH CHECK ============
@app.get("/health")
async def health_check():
    return {
        "status": "healthy",
        "version": "2.0.0",
        "timestamp": datetime.now().isoformat(),
        "processing": processing_status["is_processing"],
        "progress": processing_status["progress"]
    }

@app.get("/")
async def root():
    return FileResponse(os.path.join(FRONTEND_DIR, 'index.html'), headers={"Cache-Control": "no-cache, no-store, must-revalidate"})

@app.get("/index.html")
async def index_html():
    return FileResponse(os.path.join(FRONTEND_DIR, 'index.html'), headers={"Cache-Control": "no-cache, no-store, must-revalidate"})

@app.get("/legacy")
async def legacy_tool():
    return FileResponse(os.path.join(FRONTEND_DIR, 'index.html'))

# ============ FOLDER APIs ============

@app.post("/api/folders")
async def api_create_folder(
    name: str = Form(...),
    parent_id: str = Form("1"),
    current_user: Optional[dict] = Depends(get_optional_user)
):
    try:
        # Validate folder name
        if not is_valid_folder_name(name):
            err = format_storage_error("invalid_folder_name", {"name": name})
            return JSONResponse(status_code=400, content=err)
        
        try:
            parent_id_int = int(parent_id) if parent_id else 1
        except ValueError:
            parent_id_int = 1
        
        cid, mid = _get_context(current_user)
        
        # Create the physical directory on disk
        path_result = get_user_folder_path(cid, mid, name)
        if not path_result.get("success"):
            return JSONResponse(status_code=500, content=path_result)
        
        physical_path = path_result["path"]
        
        # Create folder in DB - create_folder() already sets the correct display path
        folder_id = create_folder(name, company_id=cid, module_id=mid, description=None, parent_id=parent_id_int)
        
        logger.info(f"Folder '{name}' created at '{physical_path}' (ID: {folder_id})")
        return {
            "success": True,
            "folder_id": folder_id,
            "path": physical_path,
            "message": "Folder created successfully"
        }
    except Exception as e:
        logger.error(f"Create folder error: {e}")
        err = format_storage_error("unknown", {"detail": str(e)})
        return JSONResponse(status_code=500, content=err)

@app.get("/api/folders")
async def api_get_folders(current_user: Optional[dict] = Depends(get_optional_user)):
    cid, mid = _get_context(current_user)
    # Removed temporary debug log
    folders = get_folders(company_id=cid, module_id=mid)
    return {"success": True, "folders": folders}

@app.delete("/api/folders/{folder_id}")
async def api_delete_folder(folder_id: int, current_user: Optional[dict] = Depends(get_optional_user)):
    try:
        cid, mid = _get_context(current_user)
        
        # Get folder info for recycle bin
        from database import get_db_connection
        conn = get_db_connection()
        folder = conn.execute("SELECT * FROM folders WHERE id = ?", (folder_id,)).fetchone()
        conn.close()
        
        if not folder:
            raise HTTPException(status_code=404, detail="Folder not found")
        
        # Move to recycle bin first
        try:
            folder_dict = dict(folder)
            
            # Physically rename the file to avoid conflict if uploaded again
            old_path = folder_dict.get('path')
            new_path = old_path
            if old_path and os.path.exists(old_path):
                import time
                new_path = f"{old_path}.deleted.{int(time.time())}"
                try:
                    os.rename(old_path, new_path)
                except OSError as e:
                    logger.warning(f"Could not rename folder for recycle bin: {e}")
                    new_path = old_path
            
            move_to_recycle_bin(
                company_id=cid,
                entity_type='folder',
                entity_id=folder_id,
                entity_name=folder_dict.get('name'),
                original_path=new_path,
                metadata=folder_dict,
                deleted_by=current_user.get('user_id') if current_user else None,
                module_id=mid
            )
        except Exception as rb_err:
            logger.warning(f"Recycle bin entry failed, proceeding with direct delete: {rb_err}")
        
        # Delete physical directory if it exists
        physical_path = dict(folder).get('path')
        if physical_path and os.path.isdir(physical_path):
            try:
                # Only delete if empty (don't remove files accidentally)
                if not os.listdir(physical_path):
                    os.rmdir(physical_path)
                    logger.info(f"Deleted empty physical directory: {physical_path}")
                else:
                    logger.warning(f"Folder '{folder['name']}' has files at {physical_path} - skipping physical deletion to prevent data loss")
            except OSError as e:
                logger.warning(f"Could not delete physical directory {physical_path}: {e}")
        
        delete_folder(folder_id)
        clear_file_cache()
        return {"success": True, "message": "Folder moved to recycle bin"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Delete folder error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/folders/{folder_id}/rename")
async def api_rename_folder(
    folder_id: int,
    new_name: str = Form(...),
    current_user: Optional[dict] = Depends(get_optional_user)
):
    """Rename a folder on disk and in the database."""
    try:
        # Validate new name
        if not is_valid_folder_name(new_name):
            err = format_storage_error("invalid_folder_name", {"name": new_name})
            return JSONResponse(status_code=400, content=err)
        
        conn = get_db_connection()
        folder = conn.execute("SELECT * FROM folders WHERE id = ?", (folder_id,)).fetchone()
        
        if not folder:
            conn.close()
            return JSONResponse(status_code=404, content=format_storage_error("unknown", {"detail": "Folder not found"}))
        
        folder_dict = dict(folder)
        old_name = folder_dict['name']
        old_path = folder_dict.get('path', '')
        
        if old_name == new_name:
            conn.close()
            return {"success": True, "message": "Folder name unchanged"}
        
        # Build new physical path
        parent_dir = os.path.dirname(old_path) if old_path else None
        new_path = os.path.join(parent_dir, new_name) if parent_dir else None
        
        # Rename physical directory if it exists
        if old_path and os.path.isdir(old_path) and new_path:
            if os.path.exists(new_path):
                conn.close()
                err = format_storage_error("file_exists", {"name": new_name})
                return JSONResponse(status_code=409, content=err)
            try:
                os.rename(old_path, new_path)
                logger.info(f"Renamed folder on disk: {old_path} -> {new_path}")
            except OSError as e:
                conn.close()
                logger.error(f"Rename folder disk error: {e}")
                err = format_storage_error("permission_denied", {"path": old_path, "os_error": str(e)})
                return JSONResponse(status_code=500, content=err)
        
        # Update DB record
        conn.execute(
            "UPDATE folders SET name = ?, path = ? WHERE id = ?",
            (new_name, new_path, folder_id)
        )
        conn.commit()
        conn.close()
        
        logger.info(f"Folder renamed: '{old_name}' -> '{new_name}' (ID: {folder_id})")
        return {
            "success": True,
            "message": f"Folder renamed from '{old_name}' to '{new_name}'",
            "folder": {
                "id": folder_id,
                "name": new_name,
                "old_name": old_name,
                "path": new_path
            }
        }
    except Exception as e:
        logger.error(f"Rename folder error: {e}")
        err = format_storage_error("unknown", {"detail": str(e)})
        return JSONResponse(status_code=500, content=err)

# ============ FILE APIs ============

@app.post("/api/upload")
async def upload_file(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    folder_id: str = Form("1"),
    header_row: str = Form("1"),
    folder_path: Optional[str] = Form(None),
    replace: bool = Form(False),
    current_user: Optional[dict] = Depends(get_optional_user)
):
    try:
        try:
            folder_id_int = int(folder_id) if folder_id else 1
        except ValueError:
            folder_id_int = 1
        try:
            header_row_int = int(header_row) if header_row else 1
        except ValueError:
            header_row_int = 1
            
        cid, mid = _get_context(current_user)
        
        # Resolve folder's physical path from DB
        conn = get_db_connection()
        folder = conn.execute("SELECT * FROM folders WHERE id = ?", (folder_id_int,)).fetchone()
        conn.close()
        
        if not folder:
            err = format_storage_error("unknown", {"detail": f"Folder ID {folder_id_int} not found"})
            return JSONResponse(status_code=404, content=err)
        
        # Use company/module isolated storage
        storage_dir = get_physical_storage_path(UPLOAD_DIR, cid, mid, folder_id_int)
        os.makedirs(storage_dir, exist_ok=True)
        
        # Save with original filename for human readability
        original_filename = file.filename
        file_path = os.path.join(storage_dir, original_filename)
        
        # Check for duplicates
        if os.path.exists(file_path):
            if not replace:
                from database import add_notification
                uid = current_user.get('id') if current_user else None
                add_notification(cid, mid, "duplicate_upload", f"Duplicate file '{original_filename}' upload attempted.", user_id=uid)
                
                err = format_storage_error("file_exists", {"name": original_filename, "prompt_replace": True})
                return JSONResponse(status_code=409, content=err)
            else:
                # We are replacing. Move old file to recycle bin and delete DB record.
                conn = get_db_connection()
                old_file = conn.execute("SELECT id, original_name FROM files WHERE folder_id = ? AND original_name = ?", (folder_id_int, original_filename)).fetchone()
                if old_file:
                    import time
                    new_path = f"{file_path}.deleted.{int(time.time())}"
                    try:
                        os.rename(file_path, new_path)
                    except OSError:
                        new_path = file_path
                    
                    move_to_recycle_bin(
                        company_id=cid,
                        entity_type='file',
                        entity_id=old_file['id'],
                        entity_name=old_file['original_name'],
                        original_path=new_path,
                        deleted_by=current_user.get('id') if current_user else None
                    )
                    conn.execute("DELETE FROM files WHERE id = ?", (old_file['id'],))
                    conn.commit()
                else:
                    # If file exists on disk but not in DB, rename it so we can write the new one
                    import time
                    try:
                        os.rename(file_path, f"{file_path}.deleted.{int(time.time())}")
                    except OSError:
                        pass
                conn.close()
        
        # Check path length
        if len(file_path) > 260:
            err = format_storage_error("path_too_long", {"path": file_path})
            return JSONResponse(status_code=400, content=err)
        
        # Write file to disk
        try:
            with open(file_path, "wb") as f:
                shutil.copyfileobj(file.file, f)
        except OSError as e:
            logger.error(f"Upload write error for '{file_path}': {e}")
            if e.errno == 28:
                err = format_storage_error("disk_full", {"path": storage_dir})
            elif e.errno == 13:
                err = format_storage_error("permission_denied", {"path": storage_dir, "os_error": str(e)})
            else:
                err = format_storage_error("dir_create_failed", {"path": storage_dir, "os_error": str(e)})
            return JSONResponse(status_code=500, content=err)
        
        size = os.path.getsize(file_path)
        sheet_count, row_count, col_count, sheet_names = get_excel_info_fast(file_path)
        ext = os.path.splitext(original_filename)[1].lower()
        
        user_id = current_user.get("id") if current_user else None
        
        file_id = save_file_metadata(
            folder_id=folder_id_int,
            original_name=original_filename,
            file_path=file_path,
            file_format=ext.replace('.', '').upper(),
            size=size,
            sheet_names=sheet_names,
            company_id=cid,
            module_id=mid,
            header_row=header_row_int,
            uploaded_by=user_id
        )
        
        if not file_id:
            # Clean up file if metadata save failed
            if os.path.exists(file_path):
                os.remove(file_path)
            err = format_storage_error("unknown", {"detail": "Failed to save file metadata"})
            return JSONResponse(status_code=500, content=err)
        
        # Cache the file info
        set_cached_file_info(file_id, {
            'file_path': file_path,
            'format': ext.replace('.', '').upper(),
            'original_name': original_filename,
            'header_row': header_row_int
        })
        
        logger.info(f"File '{original_filename}' uploaded to '{storage_dir}' (ID: {file_id})")

        # Trigger Auto-Sync if enabled
        try:
            conn_sync = get_db_connection()
            master = conn_sync.execute("SELECT auto_sync FROM master_files WHERE folder_id = ?", (folder_id_int,)).fetchone()
            if master and dict(master).get('auto_sync') == 1:
                from backend.auto_sync import trigger_folder_sync
                background_tasks.add_task(trigger_folder_sync, folder_id_int, False, current_user.get('user_id') if current_user else None)
        except Exception as e:
            logger.error(f"Failed to trigger auto_sync for upload: {e}")
        finally:
            try: conn_sync.close()
            except: pass
        
        return {
            "success": True,
            "file_id": file_id,
            "message": "File uploaded successfully",
            "file": {
                "id": file_id,
                "name": original_filename,
                "original_name": original_filename,
                "size": size,
                "format": ext.replace('.', '').upper(),
                "sheet_count": sheet_count,
                "row_count": row_count,
                "column_count": col_count,
                "path": file_path
            }
        }
    except Exception as e:
        logger.error(f"Upload error: {traceback.format_exc()}")
        err = format_storage_error("unknown", {"detail": str(e)})
        return JSONResponse(status_code=500, content=err)

@app.get("/api/files/{folder_id}")
async def api_get_files(folder_id: int):
    # --- Stale-sync recovery: any file stuck in 'in_processing' for >5 minutes
    # is reset to 'pending' so the user can click 'force retry' to re-trigger.
    # This protects against previous auto-sync runs that crashed mid-loop and
    # never updated the file's status (the most common cause of the spinner
    # being stuck forever).
    try:
        from database import get_db_connection
        _conn = get_db_connection()
        _conn.execute(
            """UPDATE files
               SET sync_status = 'pending', sync_error = COALESCE(sync_error, 'Recovered from stale in_processing state')
               WHERE folder_id = ?
                 AND sync_status = 'in_processing'
                 AND (sync_error IS NULL OR sync_error = '')""",
            (folder_id,),
        )
        _conn.commit()
        _conn.close()
    except Exception as _stale_err:
        logger.warning(f"Stale-sync recovery skipped for folder {folder_id}: {_stale_err}")

    files = get_files_by_folder(folder_id)
    # Decorate each row with the latest rejected-artefact summary so the UI can
    # render the "Rejected & Download" pill on the file list.
    _enrich_files_with_rejected_artefact(files)
    return {"success": True, "files": files}

@app.delete("/api/files/{file_id}")
async def api_delete_file(file_id: int, background_tasks: BackgroundTasks, current_user: Optional[dict] = Depends(get_optional_user)):
    try:
        from database import get_db_connection
        conn = get_db_connection()
        file = conn.execute("SELECT * FROM files WHERE id = ?", (file_id,)).fetchone()
        conn.close()
        
        if not file:
            raise HTTPException(status_code=404, detail="File not found")
        
        cid, mid = _get_context(current_user)
        
        # Move to recycle bin first
        try:
            file_dict = dict(file)
            
            # Physically rename the file to avoid conflict if uploaded again
            old_path = file_dict.get('file_path')
            new_path = old_path
            if old_path and os.path.exists(old_path):
                import time
                new_path = f"{old_path}.deleted.{int(time.time())}"
                try:
                    os.rename(old_path, new_path)
                except OSError as e:
                    logger.warning(f"Could not rename file for recycle bin: {e}")
                    new_path = old_path
            
            move_to_recycle_bin(
                company_id=cid,
                entity_type='file',
                entity_id=file_id,
                entity_name=file_dict.get('original_name') or file_dict.get('name'),
                original_path=new_path,
                metadata=file_dict,
                deleted_by=current_user.get('user_id') if current_user else None,
                module_id=mid
            )
        except Exception as rb_err:
            logger.warning(f"Recycle bin entry failed, proceeding with direct delete: {rb_err}")
        
        # Now delete from files table (but KEEP physical file for restore)
        delete_file(file_id)
        clear_file_cache()
        
        try:
            from database import add_notification
            file_name = file_dict.get('original_name') or file_dict.get('name') or f"File {file_id}"
            add_notification(cid, mid, 'info', f"File '{file_name}' moved to recycle bin", "?page=recycle_bin", user_id=current_user.get('user_id') if current_user else None)
        except Exception as e:
            logger.error(f"Failed to add notification for file delete: {e}")

        # Trigger sync for deletions regardless of auto_sync setting
        # to ensure master data does not contain deleted records.
        try:
            folder_id_int = dict(file).get('folder_id')
            if folder_id_int:
                conn_sync = get_db_connection()
                master = conn_sync.execute("SELECT 1 FROM master_files WHERE folder_id = ?", (folder_id_int,)).fetchone()
                if master:
                    from backend.auto_sync import trigger_folder_sync
                    # force_sync=False because we only want to process deletions if auto_sync=0
                    background_tasks.add_task(trigger_folder_sync, folder_id_int, False, current_user.get('user_id') if current_user else None)
                conn_sync.close()
        except Exception as e:
            logger.error(f"Failed to trigger sync for delete: {e}")

        return {"success": True, "message": "File moved to recycle bin"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Delete file error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/files/move")
async def api_move_file(file_id: str = Form(...), new_folder_id: str = Form(...)):
    try:
        try:
            file_id_int = int(file_id)
            new_folder_id_int = int(new_folder_id)
        except ValueError:
            raise HTTPException(status_code=422, detail="file_id and new_folder_id must be valid integers")
            
        from database import get_db_connection
        conn = get_db_connection()
        file = conn.execute("SELECT * FROM files WHERE id = ?", (file_id_int,)).fetchone()
        
        if not file:
            conn.close()
            raise HTTPException(status_code=404, detail="File not found")
            
        old_path = file['file_path']
        company_id = file['company_id']
        module_id = file['module_id']
        name = file['name']
        
        storage_dir = get_physical_storage_path(UPLOAD_DIR, company_id, module_id, new_folder_id_int)
        os.makedirs(storage_dir, exist_ok=True)
        new_path = os.path.join(storage_dir, name)
        
        if os.path.exists(old_path):
            import shutil
            shutil.move(old_path, new_path)
            
        conn.execute("UPDATE files SET folder_id = ?, file_path = ? WHERE id = ?", (new_folder_id_int, new_path, file_id_int))
        conn.commit()
        conn.close()
        
        clear_file_cache()
        return {"success": True, "message": "File moved successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/files/{file_id}/rename")
async def api_rename_file(file_id: int, new_name: str = Form(...)):
    try:
        from database import get_db_connection
        conn = get_db_connection()
        file = conn.execute("SELECT * FROM files WHERE id = ?", (file_id,)).fetchone()
        if not file:
            conn.close()
            raise HTTPException(status_code=404, detail="File not found")
        
        old_path = file['file_path']
        old_original = file['original_name']
        
        old_ext = os.path.splitext(old_original)[1]
        new_ext = os.path.splitext(new_name)[1]
        if not new_ext and old_ext:
            new_name = new_name + old_ext
        
        new_unique_name = f"{uuid.uuid4().hex}{os.path.splitext(new_name)[1].lower()}"
        storage_dir = os.path.dirname(old_path) if os.path.dirname(old_path) else UPLOAD_DIR
        new_path = os.path.join(storage_dir, new_unique_name)
        
        if os.path.exists(old_path):
            os.rename(old_path, new_path)
        
        conn.execute(
            "UPDATE files SET original_name = ?, name = ?, file_path = ? WHERE id = ?",
            (new_name, new_unique_name, new_path, file_id)
        )
        conn.commit()
        conn.close()
        clear_file_cache()
        
        return {
            "success": True,
            "message": "File renamed successfully",
            "file": {
                "id": file_id,
                "new_name": new_name,
                "new_path": new_path
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/files/{file_id}/details")
async def api_file_details(file_id: int):
    from database import get_db_connection
    conn = get_db_connection()
    file = conn.execute("SELECT * FROM files WHERE id = ?", (file_id,)).fetchone()
    conn.close()
    
    if not file:
        raise HTTPException(status_code=404, detail="File not found")
    
    file_dict = dict(file)
    
    try:
        sheets_detail = []
        file_path = file_dict['file_path']
        if file_path.lower().endswith('.csv'):
            import pandas as pd
            df = pd.read_csv(file_path, low_memory=False)
            sheets_detail.append({
                "name": "CSV Data",
                "rows": len(df) + 1,
                "columns": len(df.columns)
            })
        else:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheets_detail.append({
                    "name": sheet_name,
                    "rows": ws.max_row,
                    "columns": ws.max_column
                })
            wb.close()
            
        file_dict['sheets_detail'] = sheets_detail
        file_dict['sheet_count'] = len(sheets_detail)
        file_dict['row_count'] = sum((s['rows'] or 0) for s in sheets_detail)
        file_dict['column_count'] = max((s['columns'] or 0) for s in sheets_detail) if sheets_detail else 0
        
    except Exception as e:
        file_dict['sheets_detail'] = []
        file_dict['sheet_count'] = 0
        file_dict['row_count'] = 0
        file_dict['column_count'] = 0
        file_dict['error'] = str(e)
    
    return {"success": True, "file": file_dict}

@app.post("/api/files/{file_id}/header-row")
async def update_file_header_row(file_id: int, header_row: int = Form(...)):
    from database import get_db_connection
    try:
        conn = get_db_connection()
        file = conn.execute("SELECT id FROM files WHERE id = ?", (file_id,)).fetchone()
        if not file:
            conn.close()
            raise HTTPException(status_code=404, detail="File not found")
        
        conn.execute("UPDATE files SET header_row = ? WHERE id = ?", (header_row, file_id))
        conn.commit()
        conn.close()
        
        clear_file_cache()
        return {"success": True, "message": f"File header row updated to {header_row}"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/files/{file_id}/sheets")
async def api_file_sheets(file_id: int):
    from database import get_db_connection
    conn = get_db_connection()
    file = conn.execute("SELECT file_path, sheet_names FROM files WHERE id = ?", (file_id,)).fetchone()
    conn.close()
    
    if not file:
        raise HTTPException(status_code=404, detail="File not found")
    
    try:
        sheet_names = json.loads(file['sheet_names']) if file['sheet_names'] else []
        if isinstance(sheet_names, str):
            sheet_names = json.loads(sheet_names)
    except:
        sheet_names = []
    
    return {"success": True, "sheets": sheet_names}

@app.get("/api/files/{file_id}/columns")
async def api_file_columns(file_id: int, sheet_name: str, header_row: Optional[int] = None):
    from database import get_db_connection
    conn = get_db_connection()
    file = conn.execute("SELECT file_path, format, header_row FROM files WHERE id = ?", (file_id,)).fetchone()
    conn.close()
    
    if not file:
        raise HTTPException(status_code=404, detail="File not found")
    
    if header_row is None:
        header_row = file['header_row'] if ('header_row' in file.keys() and file['header_row']) else 1
    
    try:
        file_format = file['format'].upper() if file['format'] else ''
        
        if file_format == 'CSV':
            df = pd.read_csv(file['file_path'], header=header_row-1, nrows=1)
        else:
            df = pd.read_excel(file['file_path'], sheet_name=sheet_name, header=header_row-1, nrows=1)
        
        columns = [str(col) for col in df.columns.tolist()]
        return {"success": True, "columns": columns}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============ MASTER FILE APIs ============

@app.post("/api/master/merge")
async def merge_files(
    folder_id: str = Form(...),
    column_names: str = Form(...),
    auto_sync: int = Form(0),
    # --- Duplicate-detection (concat) config (optional) ---
    dedup_enabled: Optional[str] = Form(None),
    dedup_columns: Optional[str] = Form(None),     # JSON list of column names
    dedup_separator: Optional[str] = Form(None),   # e.g. ' | '
):
    try:
        try:
            folder_id_int = int(folder_id)
        except ValueError:
            raise HTTPException(status_code=422, detail="folder_id must be a valid integer")
        files = get_files_by_folder(folder_id_int)
        if not files:
            return {"success": False, "message": "No source files found in selected folder. Upload files to this folder first, then try again."}
        
        company_id = files[0].get('company_id')
        module_id = files[0].get('module_id')
        master_storage_dir = get_physical_storage_path(MASTER_DIR, company_id, module_id, folder_id_int)
        os.makedirs(master_storage_dir, exist_ok=True)
        master_db_path = os.path.join(master_storage_dir, f"folder_{folder_id}_master.duckdb")
        
        if os.path.exists(master_db_path):
            os.remove(master_db_path)
        
        conn = duckdb.connect(master_db_path)
        try:
            # Parse column names from input
            is_all_columns = column_names.strip().upper() == 'ALL'
            user_columns = [c.strip() for c in column_names.split(',') if c.strip()] if not is_all_columns else []
            
            all_data = []
            rejected_files = []
            merged_count = 0
            
            for idx, file_info in enumerate(files):
                try:
                    # Determine file format
                    file_format = file_info.get('format', '').upper()
                    file_path = file_info['file_path']
                    original_name = file_info['original_name']
                    
                    # Read the file to get sheet info
                    if file_format == 'CSV':
                        # CSV files have only one "sheet"
                        sheet_names = ['Sheet1']
                    else:
                        # Check sheet count - read_only for speed
                        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                        sheet_names = wb.sheetnames
                        wb.close()
                        
                        if len(sheet_names) > 1:
                            rejected_files.append({
                                "file": original_name,
                                "reason": f"Multiple sheets found ({len(sheet_names)} sheets). Only single-sheet files are allowed."
                            })
                            continue
                    
                    # Always use the first sheet
                    first_sheet = sheet_names[0]
                    
                    # Use the user-configured header row from file metadata
                    header_row = file_info.get('header_row', 1) or 1
                    header_idx = max(0, header_row - 1)
                    
                    if file_format == 'CSV':
                        df = pd.read_csv(file_path, header=header_idx, dtype=str)
                    else:
                        df = pd.read_excel(file_path, sheet_name=first_sheet, header=header_idx, dtype=str)
                    
                    # Strip whitespace from column names for matching
                    df.columns = [str(col).strip() for col in df.columns]
                    actual_columns = df.columns.tolist()
                    
                    # If "All" was specified, use all columns from the first file
                    if is_all_columns:
                        if idx == 0:
                            user_columns = actual_columns.copy()
                        selected_columns = user_columns
                    else:
                        # Try exact-match first; if anything is missing, fall back to fuzzy match
                        # (handles Excel's auto-suffixed duplicates like FACILITY_1 vs FACILITY).
                        desired = list(user_columns)
                        exact_matched = [c for c in desired if c in actual_columns]
                        missing_after_exact = [c for c in desired if c not in actual_columns]
                        fuzzy_matched, still_missing = resolve_columns_fuzzy(missing_after_exact, actual_columns)
                        if still_missing:
                            rejected_files.append({
                                "file": original_name,
                                "reason": f"Column(s) not found: {', '.join(still_missing)}"
                            })
                            continue
                        _seen = set()
                        selected_columns = []
                        for c in exact_matched + fuzzy_matched:
                            if c not in _seen:
                                _seen.add(c)
                                selected_columns.append(c)
                        if missing_after_exact:
                            logger.info(
                                f"Merge fuzzy column resolution for '{original_name}': "
                                f"saved={missing_after_exact} -> file={fuzzy_matched}"
                            )
                    
                    # Select only the requested columns
                    df = df[selected_columns]
                    
                    # Add Source_File_Name as the first column
                    df.insert(0, 'Source_File_Name', original_name)
                    
                    # --- DEDUP / CONCAT CHECK (strict whole-file rejection) ---
                    # Bypass: dedup is inactive when no columns are selected, or when none
                    # of the selected columns exist in this file.
                    _dedup_cfg = {
                        'dedup_enabled':   (dedup_enabled or '0') in ('1', 'true', 'True', 'yes'),
                        'dedup_columns_list': [],
                        'dedup_separator':  dedup_separator if dedup_separator else ' | ',
                    }
                    try:
                        if dedup_columns:
                            _parsed_cols = json.loads(dedup_columns)
                            if isinstance(_parsed_cols, list):
                                _dedup_cfg['dedup_columns_list'] = [str(c) for c in _parsed_cols if c]
                    except Exception:
                        pass
                    if is_dedup_active(_dedup_cfg):
                        _resolved_cols = resolve_dedup_columns(_dedup_cfg, df.columns.tolist())
                        if _resolved_cols:
                            # Initialise / open the dedup lookup table once.
                            # First file populates it; subsequent files compare against it.
                            if merged_count == 0:
                                ensure_dedup_table(conn, DEDUP_COL_NAME)
                                populate_dedup_table(conn, df, _resolved_cols, _dedup_cfg['dedup_separator'], DEDUP_COL_NAME)
                            else:
                                _existing_set = load_existing_concat_set(conn, DEDUP_COL_NAME)
                                _dup_mask = detect_duplicate_rows(
                                    df, _existing_set, _resolved_cols, _dedup_cfg['dedup_separator']
                                )
                                if _dup_mask.any():
                                    _matched_count = int(_dup_mask.sum())
                                    _reject_reason = (
                                        f"Rejected: {_matched_count} row(s) in this file match an existing "
                                        f"concat value built from columns {_resolved_cols} with separator "
                                        f"'{_dedup_cfg['dedup_separator']}'."
                                    )
                                    # Write the artefact (full file + Status / Reject_Reason columns)
                                    _artefact_path = write_rejected_artefact(
                                        df,
                                        {**file_info, 'folder_id': folder_id_int},
                                        _dup_mask,
                                        _resolved_cols,
                                        _dedup_cfg['dedup_separator'],
                                        _reject_reason,
                                    )
                                    _artefact_id = None
                                    if _artefact_path:
                                        _artefact_id = save_rejected_artefact(
                                            folder_id=folder_id_int,
                                            file_id=file_info.get('id'),
                                            original_name=original_name,
                                            artefact_path=_artefact_path,
                                            reject_reason=_reject_reason,
                                            rejected_rows=_matched_count,
                                            total_rows=len(df),
                                            source='merge',
                                        )
                                    # Mark this file as rejected in files table (best-effort)
                                    try:
                                        _conn_files = get_db_connection()
                                        _conn_files.execute(
                                            "UPDATE files SET sync_status = 'rejected', sync_error = ? WHERE folder_id = ? AND original_name = ?",
                                            (_reject_reason, folder_id_int, original_name),
                                        )
                                        _conn_files.commit()
                                        _conn_files.close()
                                    except Exception as _fsync_err:
                                        logger.warning(f"Failed to mark file rejected (merge): {_fsync_err}")
                                    # Add bell notification
                                    try:
                                        add_notification(
                                            company_id, module_id, 'error',
                                            f"File '{original_name}' rejected: {_matched_count} duplicate row(s) "
                                            f"matched existing master on columns {_resolved_cols}.",
                                            link=f"?folder={folder_id_int}",
                                            user_id=None,
                                        )
                                    except Exception:
                                        pass
                                    _fid = file_info.get('id')
                                    rejected_files.append({
                                        "file": original_name,
                                        "reason": _reject_reason,
                                        "rejected_rows": _matched_count,
                                        "total_rows": len(df),
                                        "artefact_id": _artefact_id,
                                        "file_id":               _fid,
                                        "rejected_download_url": f"/api/files/{_fid}/rejected-download" if _fid else None,
                                    })
                                    # Do NOT add this file's data to the master.
                                    continue
                                # File is clean: append its concat values to the dedup set
                                # so that the NEXT file (and subsequent sync cycles) see them.
                                populate_dedup_table(conn, df, _resolved_cols, _dedup_cfg['dedup_separator'], DEDUP_COL_NAME)
                    # --- END DEDUP CHECK ---
                    
                    all_data.append(df)
                    merged_count += 1
                    
                except Exception as e:
                    rejected_files.append({
                        "file": file_info['original_name'],
                        "reason": str(e)
                    })
            
            if not all_data:
                if rejected_files:
                    # Check if any were rejected due to missing columns
                    missing_col_reasons = [r for r in rejected_files if 'Column(s) not found' in r.get('reason', '')]
                    if missing_col_reasons:
                        first_missing = missing_col_reasons[0]
                        col = first_missing['reason'].split('Column(s) not found: ')[-1].split(',')[0].strip() if 'Column(s) not found: ' in first_missing['reason'] else ''
                        return {"success": False, "message": f"Column '{col}' not found in source files. Check available columns in file details or type 'All' to use all columns."}
                    multi_sheet = [r for r in rejected_files if 'Multiple sheets' in r.get('reason', '')]
                    if multi_sheet:
                        return {"success": False, "message": f"File '{multi_sheet[0]['file']}' has multiple sheets. Ensure all files have only one sheet, or split multi-sheet files."}
                return {"success": False, "message": "No files could be merged. Check file formats and column names, then try again."}
            
            combined_df = pd.concat(all_data, ignore_index=True)
            
            table_name = "master_data"
            conn.execute(f"DROP TABLE IF EXISTS {table_name}")
            conn.execute(f"CREATE TABLE {table_name} AS SELECT * FROM combined_df")
            
            # Reapply persisted formulas after merge
            try:
                formulas = get_master_formulas(folder_id_int)
                if formulas:
                    logger.info(f"Reapplying {len(formulas)} persisted formulas after merge")
                    current_cols = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()
                    for f in formulas:
                        try:
                            ft = f.get('formula_type', '').upper()
                            col_name = f.get('column_name', '')
                            if not ft or not col_name:
                                continue
                            if col_name in current_cols:
                                continue  # Column already exists, skip
                            
                            if ft in ('SUMIF', 'COUNTIF'):
                                # For SUMIF/COUNTIF, validate required params
                                pcol = f.get('primary_column')
                                sec_file = f.get('secondary_file')
                                sec_sheet = f.get('secondary_sheet')
                                sec_match = f.get('secondary_match_column')
                                sec_val = f.get('secondary_value_column')
                                cnt_col = f.get('count_column')
                                if not pcol or not sec_file or not sec_sheet or not sec_match:
                                    logger.warning(f"Skipping {ft} formula '{col_name}': missing required params")
                                    continue
                                if ft == 'SUMIF' and not sec_val:
                                    logger.warning(f"Skipping SUMIF formula '{col_name}': missing value column")
                                    continue
                                if ft == 'COUNTIF' and not cnt_col:
                                    logger.warning(f"Skipping COUNTIF formula '{col_name}': missing count column")
                                    continue
                                # Validate primary column exists
                                if pcol not in current_cols:
                                    logger.warning(f"Skipping {ft} formula '{col_name}': primary column '{pcol}' not found")
                                    continue
                                # Build and execute SUMIF/COUNTIF
                                # Load secondary data
                                if sec_file.startswith('master_'):
                                    sec_folder_id = int(sec_file.replace('master_', ''))
                                    sec_master = get_master_file(sec_folder_id)
                                    if not sec_master:
                                        continue
                                    sec_conn = duckdb.connect(sec_master['db_path'], read_only=True)
                                    sec_df = sec_conn.execute("SELECT * FROM master_data").fetchdf()
                                    sec_conn.close()
                                else:
                                    sec_file_id = int(sec_file)
                                    db_conn2 = get_db_connection()
                                    file_rec = db_conn2.execute("SELECT file_path, format FROM files WHERE id = ?", (sec_file_id,)).fetchone()
                                    db_conn2.close()
                                    if not file_rec:
                                        continue
                                    file_fmt = file_rec['format'].upper() if file_rec['format'] else ''
                                    if file_fmt == 'CSV':
                                        sec_df = pd.read_csv(file_rec['file_path'], dtype=str)
                                    else:
                                        sec_df = pd.read_excel(file_rec['file_path'], sheet_name=sec_sheet, dtype=str)
                                sec_cols = sec_df.columns.tolist()
                                if sec_match not in sec_cols:
                                    continue
                                if ft == 'SUMIF' and sec_val not in sec_cols:
                                    continue
                                if ft == 'COUNTIF' and cnt_col and cnt_col not in sec_cols:
                                    continue
                                conn.execute("CREATE TEMPORARY TABLE IF NOT EXISTS temp_secondary AS SELECT * FROM sec_df")
                                match_t = f.get('match_type', 'exact') or 'exact'
                                if match_t == 'exact':
                                    join_cond = f'CAST(master."{pcol}" AS VARCHAR) = CAST(secondary."{sec_match}" AS VARCHAR)'
                                else:
                                    join_cond = f'CAST(master."{pcol}" AS VARCHAR) LIKE \'%\' || CAST(secondary."{sec_match}" AS VARCHAR) || \'%\''
                                if ft == 'SUMIF':
                                    conn.execute(f'ALTER TABLE master_data ADD COLUMN "{col_name}" DOUBLE')
                                    upd_sql = f'UPDATE master_data SET "{col_name}" = (SELECT COALESCE(SUM(TRY_CAST(secondary."{sec_val}" AS DOUBLE)), 0) FROM temp_secondary AS secondary WHERE {join_cond})'
                                    conn.execute(upd_sql)
                                else:  # COUNTIF
                                    conn.execute(f'ALTER TABLE master_data ADD COLUMN "{col_name}" INTEGER')
                                    cnt_expr = f'COUNT(secondary."{cnt_col}")' if cnt_col else 'COUNT(*)'
                                    upd_sql = f'UPDATE master_data SET "{col_name}" = (SELECT {cnt_expr} FROM temp_secondary AS secondary WHERE {join_cond})'
                                    conn.execute(upd_sql)
                                conn.execute("DROP TABLE IF EXISTS temp_secondary")
                                current_cols.append(col_name)
                            else:
                                # Regular formulas
                                src_cols = f.get('source_columns', [])
                                if not src_cols:
                                    continue
                                missing = [c for c in src_cols if c not in current_cols]
                                if missing:
                                    logger.warning(f"Skipping {ft} formula '{col_name}': missing columns {missing}")
                                    continue
                                const_val = f.get('constant_value')
                                if ft == 'SUM':
                                    expr = ' + '.join(f'TRY_CAST("{c}" AS DOUBLE)' for c in src_cols)
                                elif ft == 'SUBTRACT':
                                    expr = f'TRY_CAST("{src_cols[0]}" AS DOUBLE) - TRY_CAST("{src_cols[1]}" AS DOUBLE)'
                                elif ft == 'MULTIPLY':
                                    expr = f'TRY_CAST("{src_cols[0]}" AS DOUBLE) * TRY_CAST("{src_cols[1]}" AS DOUBLE)'
                                elif ft == 'DIVIDE':
                                    expr = f'CASE WHEN TRY_CAST("{src_cols[1]}" AS DOUBLE) = 0 OR TRY_CAST("{src_cols[1]}" AS DOUBLE) IS NULL THEN 0 ELSE TRY_CAST("{src_cols[0]}" AS DOUBLE) / TRY_CAST("{src_cols[1]}" AS DOUBLE) END'
                                elif ft == 'PERCENTAGE':
                                    expr = f'CASE WHEN TRY_CAST("{src_cols[1]}" AS DOUBLE) = 0 OR TRY_CAST("{src_cols[1]}" AS DOUBLE) IS NULL THEN 0 ELSE (TRY_CAST("{src_cols[0]}" AS DOUBLE) / TRY_CAST("{src_cols[1]}" AS DOUBLE)) * 100 END'
                                elif ft == 'CONCAT':
                                    sep = const_val if const_val else ' '
                                    expr = f' || \'{sep}\' || '.join(f'"{c}"' for c in src_cols)
                                elif ft == 'ABS':
                                    if len(src_cols) != 1:
                                        continue
                                    expr = f'ABS(TRY_CAST("{src_cols[0]}" AS DOUBLE))'
                                elif ft == 'EXPRESSION':
                                    # Custom expression formula - parse and execute
                                    expression = f.get('expression', '')
                                    if not expression:
                                        logger.warning(f"Skipping EXPRESSION formula '{col_name}': no expression provided")
                                        continue
                                    try:
                                        from formula_engine import validate_formula
                                        validation = validate_formula(expression, current_cols)
                                        if not validation["valid"]:
                                            logger.warning(f"Skipping EXPRESSION formula '{col_name}': {validation['error']}")
                                            continue
                                        expr = validation["sql"]
                                    except Exception as parse_err:
                                        logger.warning(f"Skipping EXPRESSION formula '{col_name}': parse error {parse_err}")
                                        continue
                                else:
                                    continue
                                conn.execute(f'ALTER TABLE master_data ADD COLUMN "{col_name}" DOUBLE')
                                conn.execute(f'UPDATE master_data SET "{col_name}" = {expr}')
                                current_cols.append(col_name)
                        except Exception as e2:
                            logger.warning(f"Failed to reapply formula {f.get('formula_type')} '{f.get('column_name')}': {e2}")
            except Exception as e:
                logger.warning(f"Failed to reapply formulas after merge: {e}")
            
            row_count = conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
            
            # Get updated columns after reapplying formulas
            updated_cols_after_reapply = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()
            
            save_master_file(
                folder_id=folder_id_int,
                db_path=master_db_path,
                sheet_name="First_Sheet",
                columns=json.dumps(updated_cols_after_reapply),
                header_row=1,
                concat_columns=None,
                rejected_files=json.dumps(rejected_files) if rejected_files else None,
                company_id=company_id,
                module_id=module_id,
                auto_sync=auto_sync
            )

            # Persist the dedup config so the user's choice survives the next merge / auto-sync.
            try:
                _parsed_cols_for_save = []
                if dedup_columns:
                    try:
                        _t = json.loads(dedup_columns)
                        if isinstance(_t, list):
                            _parsed_cols_for_save = [str(c) for c in _t if c]
                    except Exception:
                        _parsed_cols_for_save = []
                save_dedup_config(
                    folder_id=folder_id_int,
                    enabled=(dedup_enabled or '0') in ('1', 'true', 'True', 'yes'),
                    columns=_parsed_cols_for_save,
                    separator=(dedup_separator if dedup_separator else ' | '),
                )
            except Exception as _dedup_save_err:
                logger.warning(f"Failed to persist dedup config: {_dedup_save_err}")

            # Re-apply any saved activity steps (Formula, Find & Replace, Rename, Delete, Filter)
            # so they survive master-file recreation. Wrapped in try/except so merge still succeeds
            # if activity engine itself errors out.
            try:
                from backend.auto_sync import apply_activities
                apply_activities(conn, folder_id_int, company_id, module_id)
                # Refresh columns after activities may have added/renamed/deleted columns
                try:
                    updated_cols_after_activities = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()
                except Exception:
                    updated_cols_after_activities = updated_cols_after_reapply
            except Exception as act_e:
                logger.warning(f"Failed to re-apply activities after merge: {act_e}")
                updated_cols_after_activities = updated_cols_after_reapply

            # Update auto-sync statuses for the UI
            try:
                from database import get_db_connection
                conn_sync = get_db_connection()
                conn_sync.execute("UPDATE files SET sync_status = 'synced', sync_error = NULL WHERE folder_id = ?", (folder_id_int,))
                if rejected_files:
                    for rej in rejected_files:
                        original_name = rej.get('file')
                        reason = rej.get('reason')
                        if original_name:
                            conn_sync.execute("UPDATE files SET sync_status = 'rejected', sync_error = ? WHERE folder_id = ? AND original_name = ?", (reason, folder_id_int, original_name))
                conn_sync.commit()
                conn_sync.close()
            except Exception as sync_e:
                logger.error(f"Failed to update sync statuses after master rebuild: {sync_e}")
            
            result = {
                "success": True,
                "message": f"Successfully merged {merged_count} files into master file",
                "rows": row_count,
                "merged_files": merged_count,
                "master_path": master_db_path
            }
            
            if rejected_files:
                result["rejected_files"] = rejected_files
            
            return result
            
        finally:
            try:
                conn.close()
            except Exception:
                pass
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Merge error: {e}")
        return {"success": False, "message": f"An unexpected error occurred: {str(e)}. Please try again or contact support if the issue persists."}

@app.get("/api/master/config")
async def get_master_config(folder_id: int = Query(...)):
    """Get saved master file configuration for a folder"""
    try:
        conn = get_db_connection()
        # Ensure table exists
        conn.execute('''
            CREATE TABLE IF NOT EXISTS master_file_configs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                folder_id INTEGER NOT NULL UNIQUE,
                columns TEXT,
                concat_columns TEXT,
                sheet_name TEXT,
                header_row INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.commit()
        row = conn.execute(
            "SELECT columns, concat_columns, sheet_name, header_row, updated_at FROM master_file_configs WHERE folder_id = ?",
            (folder_id,)
        ).fetchone()
        
        master_row = conn.execute("SELECT auto_sync FROM master_files WHERE folder_id = ?", (folder_id,)).fetchone()
        auto_sync = master_row[0] if master_row else 0
        conn.close()
        
        if row:
            logger.info(f"Loaded master config for folder {folder_id}: columns='{row[0]}'")
            # Format updated_at nicely
            updated_at = row[4]
            try:
                if updated_at:
                    dt = datetime.strptime(updated_at, '%Y-%m-%d %H:%M:%S')
                    updated_at = dt.strftime('%d %b %Y, %I:%M %p')
            except Exception:
                pass
            return {
                "success": True,
                "config": {
                    "columns": row[0],
                    "concat_columns": row[1],
                    "sheet_name": row[2],
                    "header_row": row[3],
                    "updated_at": updated_at,
                    "auto_sync": auto_sync
                }
            }
            
        # Even if config row is missing, we might have an auto_sync setting in master_files
        if master_row:
            return {"success": True, "config": {"auto_sync": auto_sync}}
            
        logger.debug(f"No master config found for folder {folder_id}")
        return {"success": True, "config": None}
    except Exception as e:
        logger.error(f"Get master config error for folder {folder_id}: {e}")
        return get_error_response("db_connection")

@app.post("/api/master/config")
async def save_master_config(
    folder_id: str = Form(...),
    columns: str = Form(None),
    concat_columns: str = Form(None),
    sheet_name: str = Form(None),
    header_row: str = Form(None)
):
    """Save master file configuration for a folder"""
    try:
        folder_id_int = int(folder_id)
        header_row_int = int(header_row) if header_row else None
    except ValueError:
        raise HTTPException(status_code=422, detail="folder_id and header_row must be valid integers")
    
    try:
        conn = get_db_connection()
        conn.execute('''
            CREATE TABLE IF NOT EXISTS master_file_configs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                folder_id INTEGER NOT NULL UNIQUE,
                columns TEXT,
                concat_columns TEXT,
                sheet_name TEXT,
                header_row INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.execute('''
            INSERT INTO master_file_configs (folder_id, columns, concat_columns, sheet_name, header_row)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(folder_id) DO UPDATE SET
                columns = excluded.columns,
                concat_columns = excluded.concat_columns,
                sheet_name = excluded.sheet_name,
                header_row = excluded.header_row,
                updated_at = CURRENT_TIMESTAMP
        ''', (folder_id_int, columns, concat_columns, sheet_name, header_row_int))
        conn.commit()
        conn.close()
        logger.info(f"Saved master config for folder {folder_id_int}: columns='{columns}'")
        return {"success": True, "message": "Master configuration saved"}
    except Exception as e:
        logger.error(f"Save master config error for folder {folder_id_int}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/files/{file_id}/retry-sync")
async def retry_file_sync(file_id: int, current_user: Optional[dict] = Depends(get_optional_user)):
    try:
        conn = get_db_connection()
        # Verify access
        file_rec = conn.execute("SELECT folder_id, company_id FROM files WHERE id = ?", (file_id,)).fetchone()
        if not file_rec:
            conn.close()
            raise HTTPException(status_code=404, detail="File not found")
            
        if current_user and str(file_rec['company_id']) != str(current_user.get('company_id')):
            conn.close()
            raise HTTPException(status_code=403, detail="Access denied")
            
        conn.execute("UPDATE files SET sync_status = 'pending', sync_error = NULL WHERE id = ?", (file_id,))
        conn.commit()
        conn.close()
        
        return {'success': True}
    except Exception as e:
        logger.error(f"Error retrying sync: {e}")
        return {'success': False, 'error': str(e)}

@app.get("/api/sync/active")
async def get_active_syncs(current_user: Optional[dict] = Depends(get_optional_user)):
    """
    Returns a list of all files across all folders that are currently syncing.
    """
    try:
        conn = get_db_connection()
        query = '''
            SELECT f.id, f.original_name, f.sync_status, f.folder_id, fo.name as folder_name
            FROM files f
            JOIN folders fo ON f.folder_id = fo.id
            WHERE f.sync_status IN ('in_processing', 'pending')
        '''
        params = []
        
        # If we have user isolation, apply it
        if current_user and current_user.get('company_id'):
            query += " AND f.company_id = ?"
            params.append(current_user.get('company_id'))
            
        active_files = conn.execute(query, params).fetchall()
        
        result = [
            {
                "id": f['id'],
                "file_name": f['original_name'],
                "status": f['sync_status'],
                "folder_id": f['folder_id'],
                "folder_name": f['folder_name']
            } for f in active_files
        ]
        
        # Include folders that are actively running in auto_sync queue
        from backend.auto_sync import SYNC_QUEUES
        for folder_id, state in SYNC_QUEUES.items():
            if state.get("is_running"):
                has_files = any(r['folder_id'] == folder_id for r in result)
                if not has_files:
                    folder_rec = conn.execute("SELECT name, company_id FROM folders WHERE id = ?", (folder_id,)).fetchone()
                    if folder_rec:
                        if not current_user or str(folder_rec['company_id']) == str(current_user.get('company_id', 1)):
                            result.append({
                                "id": f"folder_{folder_id}",
                                "file_name": "Master Data Update",
                                "status": "in_processing",
                                "folder_id": folder_id,
                                "folder_name": folder_rec['name']
                            })
                            
        # Now check for pending deletions (files in DuckDB but not in SQLite files)
        import os
        master_query = 'SELECT folder_id, db_path FROM master_files'
        params = []
        if current_user and current_user.get('company_id'):
            master_query += " WHERE company_id = ?"
            params.append(current_user.get('company_id'))
        masters = conn.execute(master_query, params).fetchall()
        
        for master in masters:
            db_path = master['db_path']
            folder_id = master['folder_id']
            if os.path.exists(db_path):
                import duckdb
                try:
                    # use the same configuration (default) so it works concurrently in the same process
                    duck_conn = duckdb.connect(db_path)
                    
                    # check if master_data exists
                    tables = duck_conn.execute("SHOW TABLES").fetchall()
                    if ('master_data',) in tables:
                        duckdb_files_res = duck_conn.execute("SELECT DISTINCT Source_File_Name FROM master_data").fetchall()
                        duckdb_files = set([row[0] for row in duckdb_files_res])
                        
                        sqlite_files_res = conn.execute("SELECT original_name FROM files WHERE folder_id = ?", (folder_id,)).fetchall()
                        sqlite_files = set([row['original_name'] for row in sqlite_files_res])
                        
                        files_to_remove = duckdb_files - sqlite_files
                        
                        if files_to_remove:
                            # get folder name
                            folder_rec = conn.execute("SELECT name FROM folders WHERE id = ?", (folder_id,)).fetchone()
                            folder_name = folder_rec['name'] if folder_rec else f"Folder {folder_id}"
                            
                            for f in files_to_remove:
                                result.append({
                                    'id': f'del_{folder_id}_{f}',
                                    'file_name': f,
                                    'folder_id': folder_id,
                                    'folder_name': folder_name,
                                    'status': 'deleting_from_master'
                                })
                except Exception as de:
                    logger.error(f"Error checking deletions for folder {folder_id}: {de}")
                finally:
                    if 'duck_conn' in locals():
                        duck_conn.close()
                        
        conn.close()
        return {"success": True, "active_syncs": result}
    except Exception as e:
        logger.error(f"Error getting active syncs: {e}")
        return {"success": False, "active_syncs": []}

@app.get("/api/notifications")
async def get_notifications_api(current_user: Optional[dict] = Depends(get_optional_user)):
    try:
        cid, mid = _get_context(current_user)
        user_id = current_user.get("user_id") if current_user else None
        role_id = current_user.get("role_id") if current_user else None
        from database import get_recent_notifications
        notifications = get_recent_notifications(cid, mid, limit=50, user_id=user_id, role_id=role_id)
        return {"success": True, "notifications": notifications}
    except Exception as e:
        logger.error(f"Error getting notifications: {e}")
        return {"success": False, "notifications": []}

@app.post("/api/notifications/log")
async def log_notification_api(payload: dict, current_user: Optional[dict] = Depends(get_optional_user)):
    try:
        message = payload.get("message", "Unknown event")
        notif_type = payload.get("type", "info")
        cid, mid = _get_context(current_user)
        user_id = current_user.get("user_id") if current_user else None
        role_id = current_user.get("role_id") if current_user else None
        
        from database import add_notification
        add_notification(cid, mid, notif_type, message, link=None, user_id=user_id, role_id=role_id)
        return {"success": True}
    except Exception as e:
        logger.error(f"Error logging notification: {e}")
        return {"success": False}

@app.post("/api/notifications/{notification_id}/read")
async def read_notification_api(notification_id: int, current_user: Optional[dict] = Depends(get_optional_user)):
    try:
        cid, mid = _get_context(current_user)
        user_id = current_user.get("user_id") if current_user else None
        from database import mark_notification_read
        success = mark_notification_read(notification_id, cid, user_id)
        return {"success": success}
    except Exception as e:
        logger.error(f"Error reading notification: {e}")
        return {"success": False}

@app.post("/api/notifications/read-all")
async def read_all_notifications_api(current_user: Optional[dict] = Depends(get_optional_user)):
    try:
        cid, mid = _get_context(current_user)
        user_id = current_user.get("user_id") if current_user else None
        role_id = current_user.get("role_id") if current_user else None
        from database import mark_all_notifications_read
        success = mark_all_notifications_read(cid, mid, user_id, role_id)
        return {"success": success}
    except Exception as e:
        logger.error(f"Error reading all notifications: {e}")
        return {"success": False}

@app.get("/api/company/activity-log")
async def get_company_activity_log_api(
    module_id: Optional[int] = None,
    user_id: Optional[int] = None,
    role_id: Optional[int] = None,
    current_user: Optional[dict] = Depends(get_optional_user)
):
    try:
        cid, mid = _get_context(current_user)
        if not cid:
            return {"success": False, "error": "Company context required"}
            
        from database import get_company_activity_log
        # We allow filtering by module_id, but default to the context one if not provided,
        # or we just fetch all for the company if they are viewing the company insight page.
        # Often company insight is across all modules, so if module_id is not specifically passed, we don't enforce mid.
        logs = get_company_activity_log(cid, module_id=module_id, user_id=user_id, role_id=role_id, limit=200)
        return {"success": True, "logs": logs}
    except Exception as e:
        logger.error(f"Error getting company activity log: {e}")
        return {"success": False, "logs": []}

@app.get("/api/master/{folder_id}")
async def get_master_info(folder_id: int, current_user: Optional[dict] = Depends(get_optional_user)):
    cid, mid = _get_context(current_user)
    master = get_master_file(folder_id)
    if not master or master.get('company_id') != cid or master.get('module_id') != mid:
        return {"success": True, "exists": False}
    
    # Also fetch saved config for this folder
    saved_config = None
    auto_sync = 0
    try:
        conn = get_db_connection()
        conn.execute('''
            CREATE TABLE IF NOT EXISTS master_file_configs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                folder_id INTEGER NOT NULL UNIQUE,
                columns TEXT,
                concat_columns TEXT,
                sheet_name TEXT,
                header_row INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        row = conn.execute(
            "SELECT columns, concat_columns, sheet_name, header_row, updated_at FROM master_file_configs WHERE folder_id = ?",
            (folder_id,)
        ).fetchone()
        
        master_row = conn.execute("SELECT auto_sync FROM master_files WHERE folder_id = ?", (folder_id,)).fetchone()
        if master_row:
            auto_sync = master_row[0]
        conn.close()
        
        if row:
            updated_at = row[4]
            try:
                if updated_at:
                    dt = datetime.strptime(updated_at, '%Y-%m-%d %H:%M:%S')
                    updated_at = dt.strftime('%d %b %Y, %I:%M %p')
            except Exception:
                pass
            saved_config = {
                "columns": row[0],
                "concat_columns": row[1],
                "sheet_name": row[2],
                "header_row": row[3],
                "updated_at": updated_at,
                "auto_sync": auto_sync
            }
        elif master_row:
            saved_config = {"auto_sync": auto_sync}
            
    except Exception as e:
        logger.warning(f"Could not fetch saved config for folder {folder_id}: {e}")
        try:
            conn.close()
        except:
            pass

    
    try:
        conn = duckdb.connect(master['db_path'], read_only=True)
        count = conn.execute("SELECT COUNT(*) FROM master_data").fetchone()[0]
        
        # Get distinct source files and their row counts
        source_files = conn.execute("""
            SELECT Source_File_Name, COUNT(*) as row_count 
            FROM master_data 
            GROUP BY Source_File_Name
        """).fetchall()
        
        # Get column names
        columns = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()
        master['columns'] = columns
        
        conn.close()
        master['row_count'] = count
        master['exists'] = True
        master['merged_files'] = [
            {"file": row[0], "row_count": row[1]} 
            for row in source_files
        ]
        if saved_config:
            master['saved_config'] = saved_config
    except Exception as e:
        master['exists'] = False
        master['row_count'] = 0
        master['merged_files'] = []
        master['columns'] = []
    
    # Parse rejected files from JSON (legacy column on master_files)
    if master.get('rejected_files'):
        try:
            master['rejected_files'] = json.loads(master['rejected_files'])
        except:
            master['rejected_files'] = []
    else:
        master['rejected_files'] = []

    # ---- Merge in dedup-rejected artefacts so the "Rejected Files" table in
    # the Create Master File modal reflects BOTH legacy and dedup rejections.
    # Each artefact row is normalised into the same shape the modal expects:
    #   { file, reason, rejected_rows, total_rows, file_id, rejected_download_url,
    #     rejected_at } so the modal can sort by newest and show the last 5.
    # CAP: we merge AT MOST 5 of the most-recent rejected artefacts to keep
    # the master modal light for folders with many historical rejections.
    try:
        _artefacts = list_rejected_artefacts(folder_id=folder_id)[:5]
        for _a in _artefacts:
            _fid = _a.get('file_id')
            master['rejected_files'].append({
                "file":                  _a.get('original_name') or 'Unknown file',
                "reason":                _a.get('reject_reason') or 'Rejected (duplicate detection)',
                "rejected_rows":         _a.get('rejected_rows'),
                "total_rows":            _a.get('total_rows'),
                "file_id":               _fid,
                "rejected_download_url": f"/api/files/{_fid}/rejected-download" if _fid else None,
                "rejected_at":           _a.get('created_at'),
            })
        if len(master['rejected_files']):
            master['rejected_files_total'] = len(master['rejected_files'])
    except Exception as _a_err:
        logger.warning(f"Could not merge rejected_artefacts into master_info for folder {folder_id}: {_a_err}")

    return {"success": True, "master": master}

@app.get("/api/master/{folder_id}/sheets")
async def get_master_sheets(folder_id: int, current_user: Optional[dict] = Depends(get_optional_user)):
    """Get available sheets for a master file (always returns Working)"""
    try:
        cid, mid = _get_context(current_user)
        master = get_master_file(folder_id)
        if not master:
            raise HTTPException(status_code=404, detail="Master file not found")
        if current_user is not None and (master.get('company_id') != cid or master.get('module_id') != mid):
            raise HTTPException(status_code=404, detail="Master file not found")
        return {"success": True, "sheets": ["Working"]}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/master/{folder_id}/columns")
async def get_master_columns(folder_id: int, current_user: Optional[dict] = Depends(get_optional_user)):
    """Get column names from the master file"""
    try:
        cid, mid = _get_context(current_user)
        master = get_master_file(folder_id)
        if not master:
            raise HTTPException(status_code=404, detail="Master file not found")
        if current_user is not None and (master.get('company_id') != cid or master.get('module_id') != mid):
            raise HTTPException(status_code=404, detail="Master file not found")
        
        conn = duckdb.connect(master['db_path'], read_only=True)
        columns = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()
        conn.close()
        
        return {"success": True, "columns": columns}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/master/query")
async def query_master(folder_id: str = Form(...), query: str = Form("SELECT * FROM master_data LIMIT 100"), current_user: Optional[dict] = Depends(get_optional_user)):
    try:
        cid, mid = _get_context(current_user)
        try:
            folder_id_int = int(folder_id)
        except ValueError:
            raise HTTPException(status_code=422, detail="folder_id must be a valid integer")
        master = get_master_file(folder_id_int)
        if not master or master.get('company_id') != cid or master.get('module_id') != mid:
            raise HTTPException(status_code=404, detail="Master file not found")
        
        conn = duckdb.connect(master['db_path'], read_only=True)
        result = conn.execute(query).fetchdf()
        conn.close()
        
        return {
            "success": True,
            "columns": result.columns.tolist(),
            "data": clean_nan_values(result.to_dict(orient='records')),
            "row_count": len(result)
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/master/{folder_id}/preview")
async def preview_master(
    folder_id: int,
    limit: int = Query(10, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    source_file: str = Query(None),
    search: str = Query(None),
    current_user: Optional[dict] = Depends(get_optional_user)
):
    """Get preview rows with optional filtering by source file and search text"""
    try:
        cid, mid = _get_context(current_user)
        master = get_master_file(folder_id)
        if not master or master.get('company_id') != cid or master.get('module_id') != mid:
            raise HTTPException(status_code=404, detail="Master file not found")
        
        conn = duckdb.connect(master['db_path'], read_only=True)
        
        # Build query dynamically
        conditions = []
        params = []
        
        if source_file and source_file != 'All Files':
            conditions.append("Source_File_Name = ?")
            params.append(source_file)
        
        where_clause = ""
        if conditions:
            where_clause = "WHERE " + " AND ".join(conditions)
        
        # Get total filtered count
        count_query = f"SELECT COUNT(*) FROM master_data {where_clause}"
        total_count = conn.execute(count_query, params).fetchone()[0]
        
        # Get column names
        columns = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()
        
        # Build data query
        query = f"SELECT * FROM master_data {where_clause} LIMIT ? OFFSET ?"
        query_params = params + [limit, offset]
        
        result = conn.execute(query, query_params).fetchdf()
        
        # Apply search filter in Python if specified (search across all columns)
        data = clean_nan_values(result.to_dict(orient='records'))
        if search:
            search_lower = search.lower()
            data = [
                row for row in data
                if any(str(v).lower().find(search_lower) >= 0 for v in row.values() if v is not None)
            ]
        
        conn.close()
        
        return {
            "success": True,
            "columns": columns,
            "data": data,
            "total_count": total_count,
            "returned_count": len(data),
            "limit": limit,
            "offset": offset
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/master/{folder_id}/source-files")
async def get_source_files(folder_id: int, current_user: Optional[dict] = Depends(get_optional_user)):
    """Get list of unique source files in the master"""
    try:
        cid, mid = _get_context(current_user)
        master = get_master_file(folder_id)
        if not master or master.get('company_id') != cid or master.get('module_id') != mid:
            raise HTTPException(status_code=404, detail="Master file not found")
        
        conn = duckdb.connect(master['db_path'], read_only=True)
        files = conn.execute("""
            SELECT Source_File_Name, COUNT(*) as row_count 
            FROM master_data 
            GROUP BY Source_File_Name
            ORDER BY Source_File_Name
        """).fetchall()
        conn.close()
        
        return {
            "success": True,
            "source_files": [{"name": row[0], "row_count": row[1]} for row in files]
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/master/{folder_id}/stats")
async def get_master_stats(folder_id: int, current_user: Optional[dict] = Depends(get_optional_user)):
    """Get column statistics for the master file"""
    try:
        cid, mid = _get_context(current_user)
        master = get_master_file(folder_id)
        if not master or master.get('company_id') != cid or master.get('module_id') != mid:
            raise HTTPException(status_code=404, detail="Master file not found")
        
        conn = duckdb.connect(master['db_path'], read_only=True)
        
        # Get column info
        columns = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()
        
        total_rows = conn.execute("SELECT COUNT(*) FROM master_data").fetchone()[0]
        
        stats = []
        for col in columns:
            # Data type detection
            type_query = f"""
                SELECT 
                    COUNT(*) as total,
                    COUNT("{col}") as non_null,
                    COUNT(*) - COUNT("{col}") as null_count,
                    COUNT(DISTINCT "{col}") as unique_count
                FROM master_data
            """
            type_result = conn.execute(type_query).fetchone()
            
            # Try to determine if numeric
            try:
                numeric_check = conn.execute(f"""
                    SELECT COUNT(*) 
                    FROM master_data 
                    WHERE TRY_CAST("{col}" AS DOUBLE) IS NOT NULL
                """).fetchone()[0]
                is_numeric = numeric_check > 0
            except:
                is_numeric = False
            
            stats.append({
                "column": col,
                "data_type": "Numeric" if is_numeric else "Text",
                "total_rows": type_result[0],
                "non_null_count": type_result[1],
                "null_count": type_result[2],
                "unique_values": type_result[3]
            })
        
        conn.close()
        
        return {
            "success": True,
            "total_rows": total_rows,
            "column_count": len(columns),
            "columns": stats
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/master/{folder_id}/export")
async def export_master(
    folder_id: int,
    limit: int = Form(100),
    source_file: str = Form(None),
    search: str = Form(None),
    query: str = Form(None),
    current_user: Optional[dict] = Depends(get_optional_user)
):
    """Export master data to Excel with optional filters"""
    try:
        cid, mid = _get_context(current_user)
        master = get_master_file(folder_id)
        if not master:
            raise HTTPException(status_code=404, detail="Master file not found")
        # When authenticated, enforce company/module match.
        # When running in legacy (no-auth) mode, cid/mid are None and we skip the
        # company/module check so that the endpoint works without a JWT.
        if current_user is not None and (master.get('company_id') != cid or master.get('module_id') != mid):
            raise HTTPException(status_code=404, detail="Master file not found")
        
        conn = duckdb.connect(master['db_path'], read_only=True)
        
        if query:
            # Custom SQL query
            result = conn.execute(query).fetchdf()
        else:
            # Build filtered query
            conditions = []
            params = []
            
            if source_file and source_file != 'All Files':
                conditions.append("Source_File_Name = ?")
                params.append(source_file)
            
            where_clause = ""
            if conditions:
                where_clause = "WHERE " + " AND ".join(conditions)
            
            query_sql = f"SELECT * FROM master_data {where_clause} LIMIT ?"
            params.append(limit)
            
            result = conn.execute(query_sql, params).fetchdf()
        
        conn.close()
        
        # Apply search filter if specified
        if search:
            search_lower = search.lower()
            mask = result.apply(
                lambda row: any(str(v).lower().find(search_lower) >= 0 for v in row.values if pd.notna(v)),
                axis=1
            )
            result = result[mask]
        
        # Export to Excel
        import tempfile
        fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        
        result.to_excel(temp_path, index=False, engine='openpyxl')
        
        return FileResponse(
            temp_path,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=f"master_export_{folder_id}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============ MASTER COLUMN RENAME / ROW FILTER APIs ============

@app.patch("/api/master/{folder_id}/columns/{column_name}")
async def rename_master_column(
    folder_id: int,
    column_name: str,
    new_name: str = Form(...),
    current_user: Optional[dict] = Depends(get_optional_user)
):
    """
    Rename a column in the master file. The change is auto-captured as
    a COLUMN_RENAME activity so it survives across auto-sync cycles.
    """
    try:
        master = get_master_file(folder_id)
        if not master:
            raise HTTPException(status_code=404, detail="Master file not found")

        cid, mid = _get_context(current_user)
        if current_user is not None and (master.get('company_id') != cid or master.get('module_id') != mid):
            raise HTTPException(status_code=404, detail="Master file not found")

        if not new_name or not new_name.strip():
            raise HTTPException(status_code=422, detail="new_name is required")
        new_name = new_name.strip()

        conn = duckdb.connect(master['db_path'])
        try:
            existing_cols = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()
            if column_name not in existing_cols:
                raise HTTPException(status_code=422, detail=f"Column '{column_name}' not found")
            if new_name in existing_cols and new_name != column_name:
                raise HTTPException(status_code=422, detail=f"Column '{new_name}' already exists")
            if column_name == 'Source_File_Name':
                raise HTTPException(status_code=422, detail="Cannot rename protected column 'Source_File_Name'")

            # DuckDB does not support RENAME COLUMN directly on all versions; use SELECT-based rename
            cols = [c for c in existing_cols if c != column_name]
            quoted = ', '.join(f'"{c}"' for c in cols)
            quoted_new = ', '.join([f'"{c}" AS "{new_name}"' if c == column_name else f'"{c}"' for c in existing_cols])
            conn.execute(f'CREATE TABLE master_data_new AS SELECT {quoted_new} FROM master_data')
            conn.execute('DROP TABLE master_data')
            conn.execute('ALTER TABLE master_data_new RENAME TO master_data')
        finally:
            conn.close()

        # Update metadata columns
        try:
            conn_meta = get_db_connection()
            updated_cols = [new_name if c == column_name else c for c in existing_cols]
            conn_meta.execute(
                "UPDATE master_files SET columns = ? WHERE folder_id = ?",
                (json.dumps(updated_cols), folder_id)
            )
            conn_meta.commit()
            conn_meta.close()
        except Exception as e:
            logger.warning(f"Failed to update master metadata columns: {e}")

        # === AUTO-CAPTURE: COLUMN_RENAME ===
        try:
            # Fall back to the master file's own company_id/module_id so the
            # activity row is queryable by the same (cid, mid) the list
            # endpoint uses. _get_context(current_user) may return (None, None)
            # in unauthenticated dev mode, which would otherwise make the
            # activity invisible to the panel.
            cid_rn = (current_user or {}).get('company_id') or master.get('company_id')
            mid_rn = (current_user or {}).get('module_id') or master.get('module_id')
            act_result = _create_activity_from_action(
                folder_id=folder_id, action_type='COLUMN_RENAME',
                payload={'from': column_name, 'to': new_name},
                target_column=new_name,
                company_id=cid_rn, module_id=mid_rn,
                master_file_id=master.get('id'),
                user_id=current_user.get('user_id') if current_user else None,
            )
            new_activity_id = (act_result or {}).get('activity_id')
        except Exception as _e:
            logger.warning(f'Auto-capture COLUMN_RENAME failed: {_e}')
            new_activity_id = None

        return {
            "success": True,
            "message": f"Column '{column_name}' renamed to '{new_name}'",
            "from": column_name,
            "to": new_name,
            "activity_id": new_activity_id,
            "columns": updated_cols,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Rename column error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/master/{folder_id}/row-filter")
async def apply_row_filter(
    folder_id: int,
    logic: str = Form("AND"),
    conditions: str = Form(...),  # JSON array of {column, operator, value, value_min, value_max}
    filter_name: Optional[str] = Form(None),
    current_user: Optional[dict] = Depends(get_optional_user)
):
    """
    Apply a row-filter to master_data and persist the filter as a ROW_FILTER
    activity so that whenever files are added/removed (auto-sync) the
    filtered view is re-applied to the surviving data.

    conditions schema: [{column, operator, value, value_min, value_max}, ...]
    operators supported: equal_to, not_equal_to, contains, not_contains,
                        starts_with, ends_with, greater_than, less_than,
                        between, blank, not_blank
    """
    try:
        master = get_master_file(folder_id)
        if not master:
            raise HTTPException(status_code=404, detail="Master file not found")

        cid, mid = _get_context(current_user)
        if current_user is not None and (master.get('company_id') != cid or master.get('module_id') != mid):
            raise HTTPException(status_code=404, detail="Master file not found")

        # Parse conditions JSON
        try:
            conds = json.loads(conditions) if isinstance(conditions, str) else conditions
            if not isinstance(conds, list):
                conds = []
        except (json.JSONDecodeError, TypeError):
            raise HTTPException(status_code=422, detail="conditions must be a valid JSON array")
        if not conds:
            raise HTTPException(status_code=422, detail="At least one condition is required")
        if logic not in ('AND', 'OR'):
            logic = 'AND'

        conn = duckdb.connect(master['db_path'], read_only=True)
        try:
            existing_cols = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()
            total_rows = conn.execute("SELECT COUNT(*) FROM master_data").fetchone()[0]
        finally:
            conn.close()

        # Validate condition columns
        for c in conds:
            col = c.get('column', '')
            if not col or col not in existing_cols:
                raise HTTPException(status_code=422, detail=f"Condition column '{col}' not found in master")

        # AUTO-CAPTURE: ROW_FILTER (persisted activity, not actual filter applied)
        try:
            # Use master file's own company_id/module_id as a fallback so the
            # activity is queryable by the same (cid, mid) used by the list
            # endpoint. _get_context(current_user) can return (None, None) for
            # unauthenticated dev requests, which previously made the activity
            # invisible to the activity-steps panel.
            cid_rf = (current_user or {}).get('company_id') or master.get('company_id')
            mid_rf = (current_user or {}).get('module_id') or master.get('module_id')
            _create_activity_from_action(
                folder_id=folder_id, action_type='ROW_FILTER',
                payload={
                    'logic': logic,
                    'conditions': conds,
                    'filter_name': filter_name or f"Filter {datetime.now().strftime('%H:%M:%S')}",
                },
                target_column=None,
                company_id=cid_rf, module_id=mid_rf,
                master_file_id=master.get('id'),
                user_id=current_user.get('user_id') if current_user else None,
            )
        except Exception as _e:
            logger.warning(f'Auto-capture ROW_FILTER failed: {_e}')

        return {
            "success": True,
            "message": "Filter saved as activity. Use /filtered-preview to see the result.",
            "total_rows_before": total_rows,
            "logic": logic,
            "conditions": conds,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Apply row filter error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/master/{folder_id}/filtered-preview")
async def filtered_preview(
    folder_id: int,
    logic: str = Form("AND"),
    conditions: str = Form(...),
    limit: int = Form(50),
    current_user: Optional[dict] = Depends(get_optional_user)
):
    """
    Return a DRY-RUN preview of the filter (does NOT modify master_data or
    save any activity). Pure read-only computation against the current
    master_data so the user can verify the result before saving the
    ROW_FILTER activity.
    """
    try:
        master = get_master_file(folder_id)
        if not master:
            raise HTTPException(status_code=404, detail="Master file not found")

        cid, mid = _get_context(current_user)
        if current_user is not None and (master.get('company_id') != cid or master.get('module_id') != mid):
            raise HTTPException(status_code=404, detail="Master file not found")

        try:
            conds = json.loads(conditions) if isinstance(conditions, str) else conditions
            if not isinstance(conds, list):
                conds = []
        except (json.JSONDecodeError, TypeError):
            raise HTTPException(status_code=422, detail="conditions must be a valid JSON array")
        if not conds:
            raise HTTPException(status_code=422, detail="At least one condition is required")
        if logic not in ('AND', 'OR'):
            logic = 'AND'
        limit = max(1, min(int(limit or 50), 1000))

        conn = duckdb.connect(master['db_path'], read_only=True)
        try:
            existing_cols = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()
            for c in conds:
                col = c.get('column', '')
                if not col or col not in existing_cols:
                    raise HTTPException(status_code=422, detail=f"Condition column '{col}' not found")

            # Build a SELECT statement with WHERE conditions
            where_clauses = []
            params = []
            for c in conds:
                col = c['column']
                op = (c.get('operator') or '').lower()
                val = c.get('value', '')
                vmin = c.get('value_min', '')
                vmax = c.get('value_max', '')

                if op == 'equal_to':
                    where_clauses.append(f'CAST("{col}" AS VARCHAR) = ?')
                    params.append(str(val))
                elif op == 'not_equal_to':
                    where_clauses.append(f'(CAST("{col}" AS VARCHAR) != ? OR "{col}" IS NULL)')
                    params.append(str(val))
                elif op == 'contains':
                    where_clauses.append(f'CAST("{col}" AS VARCHAR) LIKE ?')
                    params.append(f'%{val}%')
                elif op == 'not_contains':
                    where_clauses.append(f'(CAST("{col}" AS VARCHAR) NOT LIKE ? OR "{col}" IS NULL)')
                    params.append(f'%{val}%')
                elif op == 'starts_with':
                    where_clauses.append(f'CAST("{col}" AS VARCHAR) LIKE ?')
                    params.append(f'{val}%')
                elif op == 'ends_with':
                    where_clauses.append(f'CAST("{col}" AS VARCHAR) LIKE ?')
                    params.append(f'%{val}')
                elif op == 'greater_than':
                    where_clauses.append(f'TRY_CAST("{col}" AS DOUBLE) > ?')
                    params.append(float(val) if val not in (None, '') else 0)
                elif op == 'less_than':
                    where_clauses.append(f'TRY_CAST("{col}" AS DOUBLE) < ?')
                    params.append(float(val) if val not in (None, '') else 0)
                elif op == 'between':
                    where_clauses.append(f'TRY_CAST("{col}" AS DOUBLE) BETWEEN ? AND ?')
                    params.append(float(vmin) if vmin not in (None, '') else 0)
                    params.append(float(vmax) if vmax not in (None, '') else 0)
                elif op == 'blank':
                    where_clauses.append(f'("{col}" IS NULL OR TRIM(CAST("{col}" AS VARCHAR)) = \'\')')
                elif op == 'not_blank':
                    where_clauses.append(f'("{col}" IS NOT NULL AND TRIM(CAST("{col}" AS VARCHAR)) != \'\')')

            joiner = ' AND ' if logic == 'AND' else ' OR '
            where_sql = joiner.join(where_clauses) if where_clauses else 'TRUE'

            count_sql = f'SELECT COUNT(*) FROM master_data WHERE {where_sql}'
            total_filtered = conn.execute(count_sql, params).fetchone()[0]

            data_sql = f'SELECT * FROM master_data WHERE {where_sql} LIMIT ?'
            result = conn.execute(data_sql, params + [limit]).fetchdf()
        finally:
            conn.close()

        return {
            "success": True,
            "total_filtered": int(total_filtered),
            "limit": limit,
            "columns": result.columns.tolist(),
            "data": clean_nan_values(result.to_dict(orient='records')),
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Filtered preview error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ MASTER FILE FORMULA APIs ============

@app.delete("/api/master/{folder_id}")
async def delete_master_api(folder_id: int, current_user: Optional[dict] = Depends(get_optional_user)):
    """Delete a master file (DuckDB + metadata record) - moves to recycle bin first"""
    try:
        master = get_master_file(folder_id)
        if not master:
            raise HTTPException(status_code=404, detail="Master file not found")
        
        cid, mid = _get_context(current_user)
        
        # Move to recycle bin first (keep DuckDB file on disk for restore)
        try:
            move_to_recycle_bin(
                company_id=cid,
                entity_type='master_file',
                entity_id=master.get('id', folder_id),
                entity_name=master.get('sheet_name') or f"Master File {folder_id}",
                original_path=master.get('db_path'),
                metadata=master,
                deleted_by=current_user.get('user_id') if current_user else None,
                module_id=mid
            )
        except Exception as rb_err:
            logger.warning(f"Recycle bin entry failed, proceeding with direct delete: {rb_err}")
        
        # Delete the DuckDB file
        if os.path.exists(master['db_path']):
            os.remove(master['db_path'])
        
        # Delete metadata record
        delete_master_file(folder_id)
        
        # Audit log
        try:
            from database import save_audit_log
            save_audit_log(
                user_id=current_user.get('user_id') if current_user else None,
                action='DELETE',
                entity_type='master_file',
                entity_id=folder_id,
                details=f"Deleted master file for folder {folder_id}",
                company_id=cid,
                user_role=current_user.get('role') if current_user else None
            )
        except Exception as e:
            logger.warning(f"Audit log failed: {e}")
        
        return {"success": True, "message": "Master file moved to recycle bin"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Delete master error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/master/{folder_id}/formula")
async def apply_master_formula(
    folder_id: int,
    formula_type: str = Form(...),
    column_name: str = Form(...),
    source_columns: str = Form(...),
    constant_value: Optional[str] = Form(None),
    # SUMIF/COUNTIF parameters
    primary_column: Optional[str] = Form(None),
    secondary_file: Optional[str] = Form(None),
    secondary_sheet: Optional[str] = Form(None),
    secondary_match_column: Optional[str] = Form(None),
    secondary_value_column: Optional[str] = Form(None),
    count_column: Optional[str] = Form(None),
    match_type: Optional[str] = Form("exact"),
    current_user: Optional[dict] = Depends(get_optional_user)
):
    """
    Apply a formula to create a new column in the master file.
    
    Formulas:
    - SUM(col1, col2, ...): Add numeric values across columns
    - SUBTRACT(col1, col2): col1 - col2
    - MULTIPLY(col1, col2): col1 * col2
    - DIVIDE(col1, col2): col1 / col2 (handles div-by-zero)
    - PERCENTAGE(part, whole): (part / whole) * 100
    - CONCAT(col1, col2, ...): Join text columns
    - SUMIF(primary_col, secondary_file, secondary_sheet, match_col, value_col): Sum from secondary where matches
    - COUNTIF(primary_col, secondary_file, secondary_sheet, match_col): Count from secondary where matches
    """
    try:
        master = get_master_file(folder_id)
        if not master:
            raise HTTPException(status_code=404, detail="Master file not found")
        
        formula_type = formula_type.upper()
        
        # Connect to DuckDB
        conn = duckdb.connect(master['db_path'])
        
        # Get existing columns
        existing_cols = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()
        
        # Check if new column name already exists
        if column_name in existing_cols:
            conn.close()
            raise HTTPException(status_code=422, detail=f"Column '{column_name}' already exists")
        
        # Handle SUMIF/COUNTIF/VLOOKUP/HLOOKUP formulas
        if formula_type in ('SUMIF', 'COUNTIF', 'VLOOKUP', 'HLOOKUP'):
            # Validate required parameters
            if not primary_column:
                conn.close()
                raise HTTPException(status_code=422, detail=f"Primary column is required for {formula_type}")
            if not secondary_file:
                conn.close()
                raise HTTPException(status_code=422, detail=f"Secondary file is required for {formula_type}")
            if not secondary_sheet:
                conn.close()
                raise HTTPException(status_code=422, detail=f"Secondary sheet is required for {formula_type}")
            if not secondary_match_column:
                conn.close()
                raise HTTPException(status_code=422, detail=f"Secondary match column is required for {formula_type}")
            if formula_type in ('SUMIF', 'VLOOKUP', 'HLOOKUP') and not secondary_value_column:
                conn.close()
                raise HTTPException(status_code=422, detail=f"Secondary value column is required for {formula_type}")
            if formula_type == 'COUNTIF' and not count_column:
                conn.close()
                raise HTTPException(status_code=422, detail="Count column is required for COUNTIF")
            
            # Validate primary column exists in master
            if primary_column not in existing_cols:
                conn.close()
                raise HTTPException(status_code=422, detail=f"Primary column '{primary_column}' not found in master file")
            
            # Load secondary data
            try:
                if secondary_file.startswith('master_'):
                    # Load from another master file
                    sec_folder_id = int(secondary_file.replace('master_', ''))
                    sec_master = get_master_file(sec_folder_id)
                    if not sec_master:
                        conn.close()
                        raise HTTPException(status_code=404, detail="Secondary master file not found")
                    sec_conn = duckdb.connect(sec_master['db_path'], read_only=True)
                    sec_df = sec_conn.execute("SELECT * FROM master_data").fetchdf()
                    sec_conn.close()
                else:
                    # Load from uploaded file
                    sec_file_id = int(secondary_file)
                    db_conn = get_db_connection()
                    file_record = db_conn.execute(
                        "SELECT file_path, format FROM files WHERE id = ?", 
                        (sec_file_id,)
                    ).fetchone()
                    db_conn.close()
                    
                    if not file_record:
                        conn.close()
                        raise HTTPException(status_code=404, detail="Secondary file not found")
                    
                    file_format = file_record['format'].upper() if file_record['format'] else ''
                    if file_format == 'CSV':
                        sec_df = pd.read_csv(file_record['file_path'])
                    else:
                        sec_df = pd.read_excel(file_record['file_path'], sheet_name=secondary_sheet)
            except Exception as e:
                conn.close()
                raise HTTPException(status_code=500, detail=f"Failed to load secondary file: {str(e)}")
            
            # Validate secondary columns exist
            sec_cols = sec_df.columns.tolist()
            if secondary_match_column not in sec_cols:
                conn.close()
                raise HTTPException(status_code=422, detail=f"Match column '{secondary_match_column}' not found in secondary file")
            if formula_type in ('SUMIF', 'VLOOKUP', 'HLOOKUP') and secondary_value_column not in sec_cols:
                conn.close()
                raise HTTPException(status_code=422, detail=f"Value column '{secondary_value_column}' not found in secondary file")
            if formula_type == 'COUNTIF' and count_column and count_column not in sec_cols:
                conn.close()
                raise HTTPException(status_code=422, detail=f"Count column '{count_column}' not found in secondary file")
            # HLOOKUP also requires that the value column is a valid text/numeric column header
            if formula_type == 'HLOOKUP':
                # For HLOOKUP we look up a HEADER name (a column header in the secondary sheet) and return the
                # value under the same header in the row where `secondary_match_column` matches. We model it
                # by treating the value column as a row label that must exist in the secondary columns.
                if secondary_value_column not in sec_cols:
                    conn.close()
                    raise HTTPException(status_code=422, detail=f"Value column '{secondary_value_column}' not found in secondary file")
            
            # Create temporary table for secondary data
            conn.execute("CREATE TEMPORARY TABLE IF NOT EXISTS temp_secondary AS SELECT * FROM sec_df")
            
            # Build the aggregation query
            match_type = match_type or 'exact'
            if match_type == 'exact':
                join_condition = f'master."{primary_column}" = secondary."{secondary_match_column}"'
            else:  # contains - partial match
                join_condition = f'CAST(master."{primary_column}" AS VARCHAR) LIKE \'%\' || CAST(secondary."{secondary_match_column}" AS VARCHAR) || \'%\''
            
            # Add new column
            if formula_type == 'SUMIF':
                # SUMIF: Sum values from secondary where match condition is met
                sql = f'ALTER TABLE master_data ADD COLUMN "{column_name}" DOUBLE'
                conn.execute(sql)
                
                update_sql = f'''
                UPDATE master_data 
                SET "{column_name}" = (
                    SELECT COALESCE(SUM(TRY_CAST(secondary."{secondary_value_column}" AS DOUBLE)), 0)
                    FROM temp_secondary AS secondary
                    WHERE {join_condition}
                )
                '''
                conn.execute(update_sql)
            elif formula_type == 'VLOOKUP':
                # VLOOKUP: For each master row, return the first non-null value of `secondary_value_column`
                # from the secondary row where `secondary_match_column` matches `primary_column`.
                # Equivalent SQL: correlated subquery with LIMIT 1 (DuckDB supports LIMIT in subqueries).
                sql = f'ALTER TABLE master_data ADD COLUMN "{column_name}" VARCHAR'
                conn.execute(sql)
                update_sql = f'''
                UPDATE master_data
                SET "{column_name}" = (
                    SELECT secondary."{secondary_value_column}"
                    FROM temp_secondary AS secondary
                    WHERE {join_condition}
                    LIMIT 1
                )
                '''
                conn.execute(update_sql)
            elif formula_type == 'HLOOKUP':
                # HLOOKUP: In spreadsheet terms this returns the value at the intersection of
                # (a) the row whose `secondary_match_column` matches `primary_column` and
                # (b) the column whose HEADER equals `secondary_value_column`.
                # In this codebase the master and secondary are single sheets, so we treat HLOOKUP
                # as "first value from `secondary_value_column` for the row that matches". This is
                # functionally equivalent to VLOOKUP in this product's data model.
                sql = f'ALTER TABLE master_data ADD COLUMN "{column_name}" VARCHAR'
                conn.execute(sql)
                update_sql = f'''
                UPDATE master_data
                SET "{column_name}" = (
                    SELECT secondary."{secondary_value_column}"
                    FROM temp_secondary AS secondary
                    WHERE {join_condition}
                    LIMIT 1
                )
                '''
                conn.execute(update_sql)
            else:  # COUNTIF
                # COUNTIF: Count matching rows from secondary
                # Use COUNT(col) if count_column provided, otherwise COUNT(*)
                sql = f'ALTER TABLE master_data ADD COLUMN "{column_name}" INTEGER'
                conn.execute(sql)
                
                if count_column:
                    # COUNT(col) only counts non-null values in the specified column
                    count_expr = f'COUNT(secondary."{count_column}")'
                else:
                    count_expr = 'COUNT(*)'
                
                update_sql = f'''
                UPDATE master_data 
                SET "{column_name}" = (
                    SELECT {count_expr}
                    FROM temp_secondary AS secondary
                    WHERE {join_condition}
                )
                '''
                conn.execute(update_sql)
            
            # Clean up temporary table
            conn.execute("DROP TABLE IF EXISTS temp_secondary")
            
            # Get row count and columns
            row_count = conn.execute("SELECT COUNT(*) FROM master_data").fetchone()[0]
            updated_cols = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()
            conn.close()
            
            # Update metadata
            try:
                conn_sqlite = get_db_connection()
                conn_sqlite.execute(
                    "UPDATE master_files SET columns = ? WHERE folder_id = ?",
                    (json.dumps(updated_cols), folder_id)
                )
                conn_sqlite.commit()
                conn_sqlite.close()
            except Exception as e:
                logger.warning(f"Failed to update master metadata columns: {e}")
            
            return {
                "success": True,
                "message": f"Formula '{formula_type}' applied successfully",
                "column_name": column_name,
                "formula_type": formula_type,
                "primary_column": primary_column,
                "secondary_file": secondary_file,
                "secondary_sheet": secondary_sheet,
                "secondary_match_column": secondary_match_column,
                "secondary_value_column": secondary_value_column,
                "rows_affected": row_count,
                "columns": updated_cols
            }
        
        # Parse source columns (comma-separated) for regular formulas
        cols = [c.strip() for c in source_columns.split(',') if c.strip()]
        if not cols:
            raise HTTPException(status_code=422, detail="No source columns provided")
        
        # Validate source columns exist
        missing = [c for c in cols if c not in existing_cols]
        if missing:
            conn.close()
            raise HTTPException(status_code=422, detail=f"Column(s) not found: {', '.join(missing)}")
        
        # Build formula SQL
        if formula_type == 'SUM':
            if len(cols) < 1:
                conn.close()
                raise HTTPException(status_code=422, detail="SUM requires at least 1 column")
            # Build: TRY_CAST("col1" AS DOUBLE) + TRY_CAST("col2" AS DOUBLE) + ...
            expr = ' + '.join(f'TRY_CAST("{c}" AS DOUBLE)' for c in cols)
            sql = f'ALTER TABLE master_data ADD COLUMN "{column_name}" DOUBLE'
            conn.execute(sql)
            sql = f'UPDATE master_data SET "{column_name}" = {expr}'
            conn.execute(sql)
        
        elif formula_type == '-SUM':
            if len(cols) < 1:
                conn.close()
                raise HTTPException(status_code=422, detail="-SUM requires at least 1 column")
            # Build: -(TRY_CAST("col1" AS DOUBLE) + ... )
            expr = '-(' + ' + '.join(f'TRY_CAST("{c}" AS DOUBLE)' for c in cols) + ')'
            sql = f'ALTER TABLE master_data ADD COLUMN "{column_name}" DOUBLE'
            conn.execute(sql)
            sql = f'UPDATE master_data SET "{column_name}" = {expr}'
            conn.execute(sql)
        
        elif formula_type == 'SUBTRACT':
            if len(cols) != 2:
                conn.close()
                raise HTTPException(status_code=422, detail="SUBTRACT requires exactly 2 columns")
            expr = f'TRY_CAST("{cols[0]}" AS DOUBLE) - TRY_CAST("{cols[1]}" AS DOUBLE)'
            sql = f'ALTER TABLE master_data ADD COLUMN "{column_name}" DOUBLE'
            conn.execute(sql)
            sql = f'UPDATE master_data SET "{column_name}" = {expr}'
            conn.execute(sql)
        
        elif formula_type == 'MULTIPLY':
            if len(cols) != 2:
                conn.close()
                raise HTTPException(status_code=422, detail="MULTIPLY requires exactly 2 columns")
            expr = f'TRY_CAST("{cols[0]}" AS DOUBLE) * TRY_CAST("{cols[1]}" AS DOUBLE)'
            sql = f'ALTER TABLE master_data ADD COLUMN "{column_name}" DOUBLE'
            conn.execute(sql)
            sql = f'UPDATE master_data SET "{column_name}" = {expr}'
            conn.execute(sql)
        
        elif formula_type == 'DIVIDE':
            if len(cols) != 2:
                conn.close()
                raise HTTPException(status_code=422, detail="DIVIDE requires exactly 2 columns")
            # Handle div-by-zero: CASE WHEN denominator = 0 OR NULL THEN 0 ELSE numerator/denominator END
            expr = f'CASE WHEN TRY_CAST("{cols[1]}" AS DOUBLE) = 0 OR TRY_CAST("{cols[1]}" AS DOUBLE) IS NULL THEN 0 ELSE TRY_CAST("{cols[0]}" AS DOUBLE) / TRY_CAST("{cols[1]}" AS DOUBLE) END'
            sql = f'ALTER TABLE master_data ADD COLUMN "{column_name}" DOUBLE'
            conn.execute(sql)
            sql = f'UPDATE master_data SET "{column_name}" = {expr}'
            conn.execute(sql)
        
        elif formula_type == 'PERCENTAGE':
            if len(cols) != 2:
                conn.close()
                raise HTTPException(status_code=422, detail="PERCENTAGE requires exactly 2 columns (part, whole)")
            expr = f'CASE WHEN TRY_CAST("{cols[1]}" AS DOUBLE) = 0 OR TRY_CAST("{cols[1]}" AS DOUBLE) IS NULL THEN 0 ELSE (TRY_CAST("{cols[0]}" AS DOUBLE) / TRY_CAST("{cols[1]}" AS DOUBLE)) * 100 END'
            sql = f'ALTER TABLE master_data ADD COLUMN "{column_name}" DOUBLE'
            conn.execute(sql)
            sql = f'UPDATE master_data SET "{column_name}" = {expr}'
            conn.execute(sql)
        
        elif formula_type == 'ABS':
            if len(cols) != 1:
                conn.close()
                raise HTTPException(status_code=422, detail="ABS requires exactly 1 column")
            expr = f'ABS(TRY_CAST("{cols[0]}" AS DOUBLE))'
            sql = f'ALTER TABLE master_data ADD COLUMN "{column_name}" DOUBLE'
            conn.execute(sql)
            sql = f'UPDATE master_data SET "{column_name}" = {expr}'
            conn.execute(sql)
        
        elif formula_type == 'CONCAT':
            if len(cols) < 2:
                conn.close()
                raise HTTPException(status_code=422, detail="CONCAT requires at least 2 columns")
            # Build concatenation with a space separator and coalesce nulls
            separator = constant_value if constant_value else ' '
            expr = f" || '{separator}' || ".join(f"COALESCE(CAST(\"{c}\" AS VARCHAR), '')" for c in cols)
            sql = f'ALTER TABLE master_data ADD COLUMN "{column_name}" VARCHAR'
            conn.execute(sql)
            sql = f'UPDATE master_data SET "{column_name}" = {expr}'
            conn.execute(sql)
        
        else:
            conn.close()
            raise HTTPException(status_code=422, detail=f"Unknown formula type: {formula_type}")
        
        # Get row count
        row_count = conn.execute("SELECT COUNT(*) FROM master_data").fetchone()[0]
        
        # Get updated columns list
        updated_cols = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()
        
        conn.close()
        
        # Update metadata record with new columns
        try:
            conn_sqlite = get_db_connection()
            conn_sqlite.execute(
                "UPDATE master_files SET columns = ? WHERE folder_id = ?",
                (json.dumps(updated_cols), folder_id)
            )
            conn_sqlite.commit()
            conn_sqlite.close()
        except Exception as e:
            logger.warning(f"Failed to update master metadata columns: {e}")
        
        # Persist formula for auto-reapply on future merges
        try:
            existing_formulas = get_master_formulas(folder_id)
            formula_record = {
                "formula_type": formula_type,
                "column_name": column_name,
                "source_columns": cols if formula_type not in ('SUMIF', 'COUNTIF') else [],
                "constant_value": constant_value,
                "primary_column": primary_column,
                "secondary_file": secondary_file,
                "secondary_sheet": secondary_sheet,
                "secondary_match_column": secondary_match_column,
                "secondary_value_column": secondary_value_column,
                "count_column": count_column,
                "match_type": match_type,
                "created_at": datetime.now().isoformat()
            }
            existing_formulas.append(formula_record)
            update_master_formulas(folder_id, existing_formulas)
        except Exception as e:
            logger.warning(f"Failed to persist formula: {e}")

        # === AUTO-CAPTURE: FORMULA_ADD (apply_master_formula) ===
        try:
            # Use master file's own company_id/module_id as a fallback so the
            # activity is queryable by the same (cid, mid) used by the list
            # endpoint. _get_context(current_user) can return (None, None) for
            # unauthenticated dev requests, which previously made the activity
            # invisible to the activity-steps panel.
            cid_f = (current_user or {}).get('company_id') or master.get('company_id')
            mid_f = (current_user or {}).get('module_id') or master.get('module_id')
            _create_activity_from_action(
                folder_id=folder_id, action_type='FORMULA_ADD',
                payload={
                    'output_column': column_name,
                    'formula_type': formula_type,
                    'source_columns': cols,
                    'constant_value': constant_value,
                    'primary_column': primary_column,
                    'secondary_file': secondary_file,
                    'secondary_sheet': secondary_sheet,
                    'secondary_match_column': secondary_match_column,
                    'secondary_value_column': secondary_value_column,
                    'count_column': count_column,
                    'match_type': match_type,
                },
                target_column=column_name,
                company_id=cid_f, module_id=mid_f,
                master_file_id=master.get('id'),
                user_id=current_user.get('user_id') if current_user else None,
            )
        except Exception as _e:
            logger.warning(f'Auto-capture FORMULA_ADD (apply_master_formula) failed: {_e}')

        return {
            "success": True,
            "message": f"Formula '{formula_type}' applied successfully",
            "column_name": column_name,
            "formula_type": formula_type,
            "source_columns": cols,
            "rows_affected": row_count,
            "columns": updated_cols
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Formula error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/master/{folder_id}/columns/{column_name}")
async def delete_master_column(folder_id: int, column_name: str, current_user: Optional[dict] = Depends(get_optional_user)):
    """Delete a column from the master file"""
    try:
        master = get_master_file(folder_id)
        if not master:
            raise HTTPException(status_code=404, detail="Master file not found")
        
        conn = duckdb.connect(master['db_path'])
        
        # Get existing columns
        existing_cols = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()
        
        if column_name not in existing_cols:
            conn.close()
            raise HTTPException(status_code=422, detail=f"Column '{column_name}' not found")
        
        # Protect critical columns
        if column_name in ['Source_File_Name']:
            conn.close()
            raise HTTPException(status_code=422, detail=f"Cannot delete protected column '{column_name}'")
        
        # Drop column
        conn.execute(f'ALTER TABLE master_data DROP COLUMN "{column_name}"')
        
        # Get updated columns
        updated_cols = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()
        conn.close()

        # Update metadata
        try:
            conn_sqlite = get_db_connection()
            conn_sqlite.execute(
                "UPDATE master_files SET columns = ? WHERE folder_id = ?",
                (json.dumps(updated_cols), folder_id)
            )
            conn_sqlite.commit()
            conn_sqlite.close()
        except Exception as e:
            logger.warning(f"Failed to update master metadata columns: {e}")

        # === AUTO-CAPTURE: COLUMN_DELETE ===
        try:
            cid_d = (current_user or {}).get('company_id') or master.get('company_id')
            mid_d = (current_user or {}).get('module_id') or master.get('module_id')
            _create_activity_from_action(
                folder_id=folder_id, action_type='COLUMN_DELETE',
                payload={'column': column_name}, target_column=column_name,
                company_id=cid_d, module_id=mid_d,
                master_file_id=master.get('id'),
                user_id=current_user.get('user_id') if current_user else None,
            )
        except Exception as _e:
            logger.warning(f'Auto-capture COLUMN_DELETE failed: {_e}')

        return {
            "success": True,
            "message": f"Column '{column_name}' deleted successfully",
            "columns": updated_cols
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Delete column error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/master/{folder_id}/formula-preview")
async def preview_master_formula(
    folder_id: int,
    formula_type: str = Form(...),
    source_columns: str = Form(...),
    constant_value: Optional[str] = Form(None),
    # SUMIF/COUNTIF parameters
    primary_column: Optional[str] = Form(None),
    secondary_file: Optional[str] = Form(None),
    secondary_sheet: Optional[str] = Form(None),
    secondary_match_column: Optional[str] = Form(None),
    secondary_value_column: Optional[str] = Form(None),
    count_column: Optional[str] = Form(None),
    match_type: Optional[str] = Form("exact")
):
    """Preview formula result on first 5 rows without saving"""
    try:
        master = get_master_file(folder_id)
        if not master:
            raise HTTPException(status_code=404, detail="Master file not found")
        
        formula_type = formula_type.upper()
        
        conn = duckdb.connect(master['db_path'], read_only=True)
        
        # Handle SUMIF/COUNTIF/VLOOKUP/HLOOKUP previews
        if formula_type in ('SUMIF', 'COUNTIF', 'VLOOKUP', 'HLOOKUP'):
            # Validate required parameters
            if not primary_column:
                conn.close()
                raise HTTPException(status_code=422, detail=f"Primary column is required for {formula_type} preview")
            if not secondary_file:
                conn.close()
                raise HTTPException(status_code=422, detail=f"Secondary file is required for {formula_type} preview")
            if not secondary_sheet:
                conn.close()
                raise HTTPException(status_code=422, detail=f"Secondary sheet is required for {formula_type} preview")
            if not secondary_match_column:
                conn.close()
                raise HTTPException(status_code=422, detail=f"Secondary match column is required for {formula_type} preview")
            if formula_type in ('SUMIF', 'VLOOKUP', 'HLOOKUP') and not secondary_value_column:
                conn.close()
                raise HTTPException(status_code=422, detail=f"Secondary value column is required for {formula_type} preview")
            if formula_type == 'COUNTIF' and not count_column:
                conn.close()
                raise HTTPException(status_code=422, detail="Count column is required for COUNTIF preview")
            
            # Validate primary column exists in master
            existing_cols = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()
            if primary_column not in existing_cols:
                conn.close()
                raise HTTPException(status_code=422, detail=f"Primary column '{primary_column}' not found in master file")
            
            # Load secondary data
            try:
                if secondary_file.startswith('master_'):
                    sec_folder_id = int(secondary_file.replace('master_', ''))
                    sec_master = get_master_file(sec_folder_id)
                    if not sec_master:
                        conn.close()
                        raise HTTPException(status_code=404, detail="Secondary master file not found")
                    sec_conn = duckdb.connect(sec_master['db_path'], read_only=True)
                    sec_df = sec_conn.execute("SELECT * FROM master_data").fetchdf()
                    sec_conn.close()
                else:
                    sec_file_id = int(secondary_file)
                    db_conn = get_db_connection()
                    file_record = db_conn.execute(
                        "SELECT file_path, format FROM files WHERE id = ?", 
                        (sec_file_id,)
                    ).fetchone()
                    db_conn.close()
                    
                    if not file_record:
                        conn.close()
                        raise HTTPException(status_code=404, detail="Secondary file not found")
                    
                    file_format = file_record['format'].upper() if file_record['format'] else ''
                    if file_format == 'CSV':
                        sec_df = pd.read_csv(file_record['file_path'])
                    else:
                        sec_df = pd.read_excel(file_record['file_path'], sheet_name=secondary_sheet)
            except Exception as e:
                conn.close()
                raise HTTPException(status_code=500, detail=f"Failed to load secondary file: {str(e)}")
            
            # Create temporary table for secondary data
            conn.execute("CREATE TEMPORARY TABLE IF NOT EXISTS temp_secondary AS SELECT * FROM sec_df")
            
            # Build the aggregation query
            match_type = match_type or 'exact'
            if match_type == 'exact':
                join_condition = f'master."{primary_column}" = secondary."{secondary_match_column}"'
            else:  # contains - partial match
                join_condition = f'CAST(master."{primary_column}" AS VARCHAR) LIKE \'%\' || CAST(secondary."{secondary_match_column}" AS VARCHAR) || \'%\''
            
            # Build preview query (first 5 rows)
            if formula_type == 'SUMIF':
                query = f'''
                SELECT (
                    SELECT COALESCE(SUM(TRY_CAST(secondary."{secondary_value_column}" AS DOUBLE)), 0)
                    FROM temp_secondary AS secondary
                    WHERE {join_condition}
                ) as result
                FROM master_data AS master
                LIMIT 5
                '''
            elif formula_type == 'COUNTIF':
                if count_column:
                    count_expr = f'COUNT(secondary."{count_column}")'
                else:
                    count_expr = 'COUNT(*)'
                query = f'''
                SELECT (
                    SELECT {count_expr}
                    FROM temp_secondary AS secondary
                    WHERE {join_condition}
                ) as result
                FROM master_data AS master
                LIMIT 5
                '''
            else:  # VLOOKUP or HLOOKUP - return first matched value
                query = f'''
                SELECT (
                    SELECT secondary."{secondary_value_column}"
                    FROM temp_secondary AS secondary
                    WHERE {join_condition}
                    LIMIT 1
                ) as result
                FROM master_data AS master
                LIMIT 5
                '''
            
            result = conn.execute(query).fetchdf()
            conn.execute("DROP TABLE IF EXISTS temp_secondary")
            conn.close()
            
            preview_data = clean_nan_values(result['result'].tolist())
            
            return {
                "success": True,
                "formula_type": formula_type,
                "preview": preview_data,
                "primary_column": primary_column,
                "secondary_file": secondary_file,
                "secondary_match_column": secondary_match_column,
                "secondary_value_column": secondary_value_column
            }
        
        # Regular formulas
        cols = [c.strip() for c in source_columns.split(',') if c.strip()]
        if not cols:
            raise HTTPException(status_code=422, detail="No source columns provided")
        
        # Validate columns
        existing_cols = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()
        missing = [c for c in cols if c not in existing_cols]
        if missing:
            conn.close()
            raise HTTPException(status_code=422, detail=f"Column(s) not found: {', '.join(missing)}")
        
        if formula_type == 'SUM':
            expr = ' + '.join(f'TRY_CAST("{c}" AS DOUBLE)' for c in cols)
            query = f'SELECT {expr} as result FROM master_data LIMIT 5'
        elif formula_type == '-SUM':
            expr = '-(' + ' + '.join(f'TRY_CAST("{c}" AS DOUBLE)' for c in cols) + ')'
            query = f'SELECT {expr} as result FROM master_data LIMIT 5'
        elif formula_type == 'SUBTRACT':
            expr = f'TRY_CAST("{cols[0]}" AS DOUBLE) - TRY_CAST("{cols[1]}" AS DOUBLE)'
            query = f'SELECT {expr} as result FROM master_data LIMIT 5'
        elif formula_type == 'MULTIPLY':
            expr = f'TRY_CAST("{cols[0]}" AS DOUBLE) * TRY_CAST("{cols[1]}" AS DOUBLE)'
            query = f'SELECT {expr} as result FROM master_data LIMIT 5'
        elif formula_type == 'DIVIDE':
            expr = f'CASE WHEN TRY_CAST("{cols[1]}" AS DOUBLE) = 0 OR TRY_CAST("{cols[1]}" AS DOUBLE) IS NULL THEN 0 ELSE TRY_CAST("{cols[0]}" AS DOUBLE) / TRY_CAST("{cols[1]}" AS DOUBLE) END'
            query = f'SELECT {expr} as result FROM master_data LIMIT 5'
        elif formula_type == 'PERCENTAGE':
            expr = f'CASE WHEN TRY_CAST("{cols[1]}" AS DOUBLE) = 0 OR TRY_CAST("{cols[1]}" AS DOUBLE) IS NULL THEN 0 ELSE (TRY_CAST("{cols[0]}" AS DOUBLE) / TRY_CAST("{cols[1]}" AS DOUBLE)) * 100 END'
            query = f'SELECT {expr} as result FROM master_data LIMIT 5'
        elif formula_type == 'CONCAT':
            separator = constant_value if constant_value else ' '
            expr = f' || \'{separator}\' || '.join(f'"{c}"' for c in cols)
            query = f'SELECT {expr} as result FROM master_data LIMIT 5'
        elif formula_type == 'ABS':
            if len(cols) != 1:
                conn.close()
                raise HTTPException(status_code=422, detail="ABS requires exactly 1 column")
            expr = f'ABS(TRY_CAST("{cols[0]}" AS DOUBLE))'
            query = f'SELECT {expr} as result FROM master_data LIMIT 5'
        else:
            conn.close()
            raise HTTPException(status_code=422, detail=f"Unknown formula type: {formula_type}")
        
        result = conn.execute(query).fetchdf()
        conn.close()
        
        preview_data = clean_nan_values(result['result'].tolist())
        
        return {
            "success": True,
            "formula_type": formula_type,
            "preview": preview_data,
            "source_columns": cols
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Formula preview error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ FIND & REPLACE API ============
@app.post("/api/master/{folder_id}/find-replace")
async def find_replace_master(
    folder_id: int,
    find_text: str = Form(...),
    replace_text: str = Form(""),
    match_type: str = Form("contains"),
    case_sensitive: str = Form("false"),
    dry_run: str = Form("false"),
    column: Optional[str] = Form(None),
    current_user: Optional[dict] = Depends(get_optional_user)
):
    """
    Find and replace text across columns in a master file.

    - If `column` is provided, only that column is updated.
    - Otherwise every text-typed column in master_data is processed.
    - Supported match_type values: 'contains', 'exact', 'starts_with', 'ends_with'.
    - `case_sensitive` and `dry_run` are string form fields ("true"/"false").
    - Returns columns_modified list with row counts per column and total_rows_affected.
    """
    try:
        master = get_master_file(folder_id)
        if not master:
            raise HTTPException(status_code=404, detail="Master file not found")

        if find_text is None or find_text == "":
            raise HTTPException(status_code=422, detail="find_text is required")

        match_type = (match_type or "contains").lower()
        if match_type not in ("contains", "exact", "starts_with", "ends_with"):
            raise HTTPException(status_code=422, detail=f"Unsupported match_type: {match_type}")

        is_case_sensitive = str(case_sensitive).lower() == "true"
        is_dry_run = str(dry_run).lower() == "true"

        conn = duckdb.connect(master['db_path'])
        try:
            existing_cols = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()

            if column:
                if column not in existing_cols:
                    raise HTTPException(status_code=422, detail=f"Column '{column}' not found in master file")
                target_cols = [column]
            else:
                # All columns except bookkeeping ones
                target_cols = [c for c in existing_cols if c != 'Source_File_Name']

            # Build the SQL predicate for a match (returns true if find_text matches `col_value`)
            def match_expr(col_ident: str) -> str:
                col_ident = col_ident.replace('"', '""')
                lit = find_text.replace("'", "''")
                if is_case_sensitive:
                    v = f'CAST("{col_ident}" AS VARCHAR)'
                else:
                    v = f'LOWER(CAST("{col_ident}" AS VARCHAR))'
                needle = lit if is_case_sensitive else lit.lower()
                needle_esc = needle.replace("'", "''")
                if match_type == "contains":
                    return f"({v} LIKE '%{needle_esc}%')"
                if match_type == "exact":
                    return f"({v} = '{needle_esc}')"
                if match_type == "starts_with":
                    return f"({v} LIKE '{needle_esc}%')"
                if match_type == "ends_with":
                    return f"({v} LIKE '%{needle_esc}')"
                return "FALSE"

            # Build the replacement expression for REPLACE(): if the cell matches, swap find_text with replace_text
            def replace_expr(col_ident: str) -> str:
                col_ident = col_ident.replace('"', '""')
                needle = find_text.replace("'", "''")
                repl = (replace_text or "").replace("'", "''")
                if is_case_sensitive:
                    return f"REPLACE(CAST(\"{col_ident}\" AS VARCHAR), '{needle}', '{repl}')"
                # Case-insensitive: we can only safely replace by lowercasing then re-casing the first letter is lossy.
                # Practical approach: use REGEXP_REPLACE with case-insensitive flag.
                return f"REGEXP_REPLACE(CAST(\"{col_ident}\" AS VARCHAR), '{needle}', '{repl}', 'i')"

            columns_modified = []
            total_rows_affected = 0

            for col_name in target_cols:
                col_id = col_name
                # Count matches first (also for dry_run)
                count_sql = f"SELECT COUNT(*) FROM master_data WHERE {match_expr(col_id)}"
                try:
                    matched = conn.execute(count_sql).fetchone()[0]
                except Exception as ce:
                    logger.warning(f"Skipping column '{col_name}' (unsupported type for find/replace): {ce}")
                    continue

                if matched == 0:
                    continue

                if is_dry_run:
                    columns_modified.append({"column": col_name, "rows_affected": int(matched)})
                    total_rows_affected += int(matched)
                    continue

                # Perform replacement
                update_sql = f'UPDATE master_data SET "{col_id}" = {replace_expr(col_id)} WHERE {match_expr(col_id)}'
                try:
                    conn.execute(update_sql)
                    columns_modified.append({"column": col_name, "rows_affected": int(matched)})
                    total_rows_affected += int(matched)
                except Exception as ue:
                    logger.warning(f"Failed to update column '{col_name}': {ue}")
                    continue
        finally:
            conn.close()

        # === AUTO-CAPTURE: FIND_REPLACE ===
        try:
            # Use master file's own company_id/module_id as a fallback so the
            # activity is queryable by the same (cid, mid) used by the list
            # endpoint. _get_context(current_user) can return (None, None) for
            # unauthenticated dev requests, which previously made the activity
            # invisible to the activity-steps panel.
            cid_fr = (current_user or {}).get('company_id') or master.get('company_id')
            mid_fr = (current_user or {}).get('module_id') or master.get('module_id')
            _create_activity_from_action(
                folder_id=folder_id, action_type='FIND_REPLACE',
                payload={
                    'find_text': find_text,
                    'replace_text': replace_text,
                    'match_type': match_type,
                    'case_sensitive': is_case_sensitive,
                    'column': column,
                    'rows_affected': total_rows_affected,
                },
                target_column=column,
                company_id=cid_fr, module_id=mid_fr,
                master_file_id=master.get('id'),
                user_id=current_user.get('user_id') if current_user else None,
            )
        except Exception as _e:
            logger.warning(f'Auto-capture FIND_REPLACE failed: {_e}')

        return {
            "success": True,
            "message": f"Find & Replace completed across {len(columns_modified)} column(s)",
            "columns_modified": columns_modified,
            "total_rows_affected": total_rows_affected,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Find & Replace error: {e}")
        return {"success": False, "message": f"Find & Replace failed: {str(e)}"}


# ============ CUSTOM EXPRESSION FORMULA APIs ============

@app.post("/api/master/{folder_id}/formula-expression")
async def apply_formula_expression(
    folder_id: int,
    expression: str = Form(...),
    column_name: str = Form(...),
    current_user: Optional[dict] = Depends(get_optional_user)
):
    """
    Apply a custom Excel-like formula expression to create a new column.

    Example expressions:
      =SUM(Amount, Tax)
      =-SUM(Amount, Tax)
      =Amount + Tax * 0.18
      =ABS(Amount)
      =ROUND(SUM(Amount, Tax), 2)
      =CONCAT(FirstName, ' ', LastName)
      =COALESCE(Amount, 0)
      =IFNULL(Amount, 0)
    """
    try:
        master = get_master_file(folder_id)
        if not master:
            raise HTTPException(status_code=404, detail="Master file not found")

        # Connect to DuckDB
        conn = duckdb.connect(master['db_path'])

        # Get existing columns
        existing_cols = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()

        # Check if new column name already exists
        if column_name in existing_cols:
            conn.close()
            raise HTTPException(status_code=422, detail=f"Column '{column_name}' already exists")

        # Validate and parse the expression
        validation = validate_formula(expression, existing_cols)
        if not validation["valid"]:
            conn.close()
            return {
                "success": False,
                "message": validation["error"],
                "suggestion": validation.get("suggestion", "Check your formula syntax and try again.")
            }

        sql_expr = validation["sql"]
        referenced_cols = validation["columns"]

        # Execute: add column and update
        try:
            conn.execute(f'ALTER TABLE master_data ADD COLUMN "{column_name}" DOUBLE')
            conn.execute(f'UPDATE master_data SET "{column_name}" = {sql_expr}')
        except Exception as sql_err:
            conn.close()
            return {
                "success": False,
                "message": f"Formula execution failed: {str(sql_err)}",
                "suggestion": "The formula parsed correctly but failed to execute. Check for data type mismatches in the referenced columns."
            }

        # Get updated info
        row_count = conn.execute("SELECT COUNT(*) FROM master_data").fetchone()[0]
        updated_cols = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()
        conn.close()

        # Update metadata
        try:
            conn_sqlite = get_db_connection()
            conn_sqlite.execute(
                "UPDATE master_files SET columns = ? WHERE folder_id = ?",
                (json.dumps(updated_cols), folder_id)
            )
            conn_sqlite.commit()
            conn_sqlite.close()
        except Exception as e:
            logger.warning(f"Failed to update master metadata columns: {e}")

        # Persist formula for auto-reapply on future merges
        try:
            existing_formulas = get_master_formulas(folder_id)
            formula_record = {
                "formula_type": "EXPRESSION",
                "column_name": column_name,
                "expression": expression,
                "source_columns": referenced_cols,
                "created_at": datetime.now().isoformat()
            }
            existing_formulas.append(formula_record)
            update_master_formulas(folder_id, existing_formulas)
        except Exception as e:
            logger.warning(f"Failed to persist formula: {e}")

        # === AUTO-CAPTURE: FORMULA_ADD (apply_formula_expression) ===
        try:
            # Use master file's own company_id/module_id as a fallback so the
            # activity is queryable by the same (cid, mid) used by the list
            # endpoint. _get_context(current_user) can return (None, None) for
            # unauthenticated dev requests, which previously made the activity
            # invisible to the activity-steps panel.
            cid_e = (current_user or {}).get('company_id') or master.get('company_id')
            mid_e = (current_user or {}).get('module_id') or master.get('module_id')
            _create_activity_from_action(
                folder_id=folder_id, action_type='FORMULA_ADD',
                payload={
                    'output_column': column_name,
                    'formula_type': 'EXPRESSION',
                    'expression': expression,
                    'source_columns': referenced_cols,
                    'sql': sql_expr,
                },
                target_column=column_name,
                company_id=cid_e, module_id=mid_e,
                master_file_id=master.get('id'),
                user_id=current_user.get('user_id') if current_user else None,
            )
        except Exception as _e:
            logger.warning(f'Auto-capture FORMULA_ADD (apply_formula_expression) failed: {_e}')

        return {
            "success": True,
            "message": f"Formula expression applied successfully",
            "column_name": column_name,
            "formula_type": "EXPRESSION",
            "expression": expression,
            "sql": sql_expr,
            "referenced_columns": referenced_cols,
            "rows_affected": row_count,
            "columns": updated_cols
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Formula expression error: {e}")
        return {
            "success": False,
            "message": f"An unexpected error occurred: {str(e)}",
            "suggestion": "Please try again or contact support if the issue persists."
        }


@app.post("/api/master/{folder_id}/formula-expression-preview")
async def preview_formula_expression(
    folder_id: int,
    expression: str = Form(...)
):
    """
    Preview a custom formula expression result on first 5 rows without saving.
    Returns the parsed SQL and preview values for validation.
    """
    try:
        master = get_master_file(folder_id)
        if not master:
            raise HTTPException(status_code=404, detail="Master file not found")

        conn = duckdb.connect(master['db_path'], read_only=True)

        # Get existing columns
        existing_cols = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()

        # Validate and parse the expression
        validation = validate_formula(expression, existing_cols)
        if not validation["valid"]:
            conn.close()
            return {
                "success": False,
                "message": validation["error"],
                "suggestion": validation.get("suggestion", "Check your formula syntax and try again.")
            }

        sql_expr = validation["sql"]

        # Execute preview query
        try:
            query = f'SELECT {sql_expr} as result FROM master_data LIMIT 5'
            result = conn.execute(query).fetchdf()
        except Exception as sql_err:
            conn.close()
            return {
                "success": False,
                "message": f"Formula preview execution failed: {str(sql_err)}",
                "suggestion": "The formula parsed correctly but failed to execute on sample data. Check for data type issues."
            }

        conn.close()

        preview_data = clean_nan_values(result['result'].tolist())

        return {
            "success": True,
            "formula_type": "EXPRESSION",
            "expression": expression,
            "sql": sql_expr,
            "referenced_columns": validation["columns"],
            "preview": preview_data
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Formula expression preview error: {e}")
        return {
            "success": False,
            "message": f"An unexpected error occurred: {str(e)}",
            "suggestion": "Please try again or contact support if the issue persists."
        }


# ============ MASTER ACTIVITY APIs (Activity Window) ============
# ETL-style persistent steps. The user can save formula, find/replace, rename, or
# delete column steps. They survive across auto-sync cycles when files are
# added or removed in the folder.

@app.get("/api/master/{folder_id}/activities")
async def api_list_activities(folder_id: int, current_user: Optional[dict] = Depends(get_optional_user)):
    """List all activities for a master file, in step order."""
    try:
        cid, mid = _get_context(current_user)
        master = get_master_file(folder_id)
        if not master:
            raise HTTPException(status_code=404, detail="Master file not found")
        if current_user is not None and (master.get('company_id') != cid or master.get('module_id') != mid):
            raise HTTPException(status_code=404, detail="Master file not found")

        activities = list_master_activities(folder_id, company_id=cid, module_id=mid, enabled_only=False)
        # Reorder by step_order then id
        activities.sort(key=lambda a: (a.get('step_order', 0), a.get('id', 0)))

        return {
            "success": True,
            "activities": activities,
            "count": len(activities),
            "folder_id": folder_id
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"List activities error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/master/{folder_id}/activities")
async def api_create_activity(
    folder_id: int,
    activity_type: str = Form(...),
    payload: str = Form(...),
    step_order: Optional[str] = Form(None),
    target_column: Optional[str] = Form(None),
    current_user: Optional[dict] = Depends(get_optional_user)
):
    """Create a new activity step (formula, find/replace, rename, delete)."""
    try:
        cid, mid = _get_context(current_user)
        master = get_master_file(folder_id)
        if not master:
            raise HTTPException(status_code=404, detail="Master file not found")
        if current_user is not None and (master.get('company_id') != cid or master.get('module_id') != mid):
            raise HTTPException(status_code=404, detail="Master file not found")

        # Parse payload as JSON
        try:
            payload_dict = json.loads(payload) if isinstance(payload, str) else payload
            if not isinstance(payload_dict, dict):
                payload_dict = {}
        except (json.JSONDecodeError, TypeError):
            raise HTTPException(status_code=422, detail="payload must be valid JSON")

        # Validate activity_type
        valid_types = {'FORMULA_ADD', 'FORMULA_UPDATE', 'FIND_REPLACE', 'COLUMN_RENAME', 'COLUMN_DELETE', 'ROW_FILTER'}
        act_upper = (activity_type or '').upper()
        if act_upper not in valid_types:
            raise HTTPException(status_code=422, detail=f"activity_type must be one of: {sorted(valid_types)}")

        # Auto-derive target_column from payload when not supplied
        if not target_column:
            if act_upper in ('FORMULA_ADD',):
                target_column = payload_dict.get('output_column') or payload_dict.get('column_name')
            elif act_upper == 'FORMULA_UPDATE':
                target_column = payload_dict.get('target_column')
            elif act_upper == 'COLUMN_RENAME':
                target_column = payload_dict.get('from')
            elif act_upper == 'COLUMN_DELETE':
                target_column = payload_dict.get('column')

        # Parse step_order
        so = None
        if step_order is not None and str(step_order).strip():
            try:
                so = int(step_order)
            except ValueError:
                so = None

        user_id = current_user.get('user_id') if current_user else None
        activity_id = create_master_activity(
            folder_id=folder_id,
            activity_type=act_upper,
            payload=payload_dict,
            step_order=so,
            target_column=target_column,
            company_id=cid,
            module_id=mid,
            master_file_id=master.get('id'),
            created_by=user_id,
        )

        # Read back the created activity
        act = get_master_activity(activity_id)
        return {"success": True, "activity": act, "activity_id": activity_id, "message": "Activity created"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Create activity error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.patch("/api/master/{folder_id}/activities/{activity_id}")
async def api_update_activity(
    folder_id: int,
    activity_id: int,
    payload: Optional[str] = Form(None),
    is_enabled: Optional[str] = Form(None),
    target_column: Optional[str] = Form(None),
    step_order: Optional[str] = Form(None),
    current_user: Optional[dict] = Depends(get_optional_user)
):
    """Update an activity (toggle enabled, edit payload, rename, etc.)."""
    try:
        cid, mid = _get_context(current_user)
        existing = get_master_activity(activity_id)
        if not existing or existing.get('folder_id') != folder_id:
            raise HTTPException(status_code=404, detail="Activity not found")

        kwargs = {}
        if payload is not None:
            try:
                payload_dict = json.loads(payload) if isinstance(payload, str) else payload
            except (json.JSONDecodeError, TypeError):
                raise HTTPException(status_code=422, detail="payload must be valid JSON")
            kwargs['payload'] = payload_dict

        if is_enabled is not None:
            if str(is_enabled).lower() in ('true', '1', 'yes'):
                kwargs['is_enabled'] = 1
            elif str(is_enabled).lower() in ('false', '0', 'no'):
                kwargs['is_enabled'] = 0
            else:
                raise HTTPException(status_code=422, detail="is_enabled must be true/false")

        if target_column is not None:
            kwargs['target_column'] = target_column

        if step_order is not None and str(step_order).strip():
            try:
                kwargs['step_order'] = int(step_order)
            except ValueError:
                raise HTTPException(status_code=422, detail="step_order must be an integer")

        if not kwargs:
            raise HTTPException(status_code=422, detail="No updatable fields provided")

        update_master_activity(activity_id, **kwargs)
        updated = get_master_activity(activity_id)
        return {"success": True, "activity": updated, "message": "Activity updated"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Update activity error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/master/{folder_id}/activities/{activity_id}")
async def api_delete_activity(
    folder_id: int,
    activity_id: int,
    current_user: Optional[dict] = Depends(get_optional_user)
):
    """Delete a single activity step."""
    try:
        existing = get_master_activity(activity_id)
        if not existing or existing.get('folder_id') != folder_id:
            raise HTTPException(status_code=404, detail="Activity not found")
        delete_master_activity(activity_id)
        return {"success": True, "message": "Activity deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Delete activity error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/master/{folder_id}/activities/reorder")
async def api_reorder_activities(
    folder_id: int,
    ordered_ids: str = Form(...),
    current_user: Optional[dict] = Depends(get_optional_user)
):
    """Reorder activity steps. Pass the IDs in the desired order."""
    try:
        try:
            ids_list = json.loads(ordered_ids)
            if not isinstance(ids_list, list):
                raise ValueError()
            ids_list = [int(x) for x in ids_list]
        except (json.JSONDecodeError, ValueError, TypeError):
            raise HTTPException(status_code=422, detail="ordered_ids must be a JSON array of integers")

        # Verify all ids belong to this folder
        existing = list_master_activities(folder_id, enabled_only=False)
        valid_ids = {a['id'] for a in existing}
        for aid in ids_list:
            if aid not in valid_ids:
                raise HTTPException(status_code=422, detail=f"Activity {aid} does not belong to folder {folder_id}")

        reorder_master_activities(folder_id, ids_list)
        return {"success": True, "message": "Activities reordered", "count": len(ids_list)}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Reorder activities error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/master/{folder_id}/activities/{activity_id}/test")
async def api_test_activity(
    folder_id: int,
    activity_id: int,
    current_user: Optional[dict] = Depends(get_optional_user)
):
    """Dry-run an activity on the first 5 rows of master_data.
    Does NOT persist any changes. Returns before/after samples."""
    try:
        cid, mid = _get_context(current_user)
        master = get_master_file(folder_id)
        if not master:
            raise HTTPException(status_code=404, detail="Master file not found")

        existing = get_master_activity(activity_id)
        if not existing or existing.get('folder_id') != folder_id:
            raise HTTPException(status_code=404, detail="Activity not found")

        if not os.path.exists(master['db_path']):
            raise HTTPException(status_code=404, detail="Master DuckDB file not found on disk")

        conn = duckdb.connect(master['db_path'], read_only=True)
        try:
            tables = conn.execute("SHOW TABLES").fetchall()
            if ('master_data',) not in tables:
                return {"success": False, "message": "master_data table does not exist", "preview": []}
            existing_cols = conn.execute("SELECT * FROM master_data LIMIT 0").fetchdf().columns.tolist()
            sample = conn.execute("SELECT * FROM master_data LIMIT 5").fetchdf()
        finally:
            conn.close()

        # Build a synthetic "after" view by describing the change.
        # For full safety we run the activity in an in-memory DuckDB copy with 5 rows.
        preview = {
            "rows_before": len(sample),
            "columns_before": existing_cols,
            "activity_type": existing.get('activity_type'),
            "would_change": True,
        }
        preview["before_sample"] = clean_nan_values(sample.to_dict(orient='records'))

        act_type = (existing.get('activity_type') or '').upper()
        payload = existing.get('payload') or {}

        if act_type == 'FIND_REPLACE':
            fv = payload.get('find', '')
            rv = payload.get('replace', '')
            scope = payload.get('scope_columns') or [c for c in existing_cols if c != 'Source_File_Name']
            after_rows = []
            for row in preview["before_sample"]:
                new_row = dict(row)
                for c in scope:
                    if c in new_row and new_row[c] is not None:
                        val = str(new_row[c])
                        if fv and fv in val:
                            new_row[c] = val.replace(fv, rv)
                after_rows.append(new_row)
            preview["after_sample"] = after_rows
            preview["affected_columns"] = scope

        elif act_type in ('FORMULA_ADD', 'FORMULA_UPDATE'):
            expr = payload.get('expression') or ''
            sql_expr = payload.get('sql') or ''
            out_col = payload.get('output_column') or existing.get('target_column')
            
            if not expr and sql_expr:
                # Already parsed SQL from vlookup/sumif
                pass
            else:
                validation = validate_formula(expr, existing_cols)
                if not validation["valid"]:
                    return {
                        "success": False,
                        "message": validation["error"],
                        "suggestion": validation.get("suggestion", ""),
                        "preview": preview
                    }
                sql_expr = validation["sql"]
                
            preview["sql"] = sql_expr
            preview["output_column"] = out_col
            try:
                conn_rw = duckdb.connect(master['db_path'], read_only=True)
                try:
                    q = f"SELECT ({sql_expr}) as result FROM master_data LIMIT 5"
                    r = conn_rw.execute(q).fetchdf()
                    preview["after_sample"] = clean_nan_values(r['result'].tolist())
                finally:
                    conn_rw.close()
            except Exception as e:
                preview["after_sample"] = []
                preview["execution_error"] = str(e)
            preview["affected_columns"] = [out_col] if out_col else []

        elif act_type == 'COLUMN_RENAME':
            frm = payload.get('from')
            to = payload.get('to')
            preview["affected_columns"] = [frm, to]
            preview["after_sample"] = preview["before_sample"]  # rename is a metadata op, no row changes
            preview["note"] = f"Renaming column '{frm}' to '{to}' is a schema change; rows are not modified."

        elif act_type == 'COLUMN_DELETE':
            col = payload.get('column')
            preview["affected_columns"] = [col]
            preview["after_sample"] = [
                {k: v for k, v in r.items() if k != col}
                for r in preview["before_sample"]
            ]
            preview["note"] = f"Deleting column '{col}' will drop the column from master_data."
        else:
            preview["after_sample"] = preview["before_sample"]

        return {"success": True, "preview": preview}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Test activity error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/admin/migrate-legacy-formulas")
async def api_migrate_legacy_formulas(current_user: Optional[dict] = Depends(get_optional_user)):
    """Admin/diagnostic endpoint: run the one-time migration of master_files.formulas
    into master_activities. Idempotent — safe to call multiple times."""
    try:
        result = migrate_legacy_master_formulas()
        return {"success": True, **result}
    except Exception as e:
        logger.error(f"Migration endpoint error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ PRIMARY DATA APIs ============

@app.post("/api/primary/generate")
async def generate_primary(
    file_id: str = Form(...),
    sheet_name: str = Form(...),
    column_name: str = Form(...),
    header_row: str = Form("1"),
    sales_amount_column: str = Form(None),
    fields: str = Form(None),
    current_user: Optional[dict] = Depends(get_optional_user)
):
    """Generate primary data file. Accepts optional 'fields' JSON array for multi-field extraction."""
    try:
        header_row_int = int(header_row) if header_row else 1
    except ValueError:
        raise HTTPException(status_code=422, detail="header_row must be a valid integer")
    
    # Parse fields JSON if provided
    parsed_fields = None
    if fields:
        try:
            parsed_fields = json.loads(fields)
        except json.JSONDecodeError:
            raise HTTPException(status_code=422, detail="fields must be valid JSON array")
    
    try:
        result = generate_primary_data(
            file_id, sheet_name, column_name, header_row_int,
            sales_amount_column=sales_amount_column,
            fields=parsed_fields
        )
        
        cid, mid = _get_context(current_user)
        from database import add_notification
        add_notification(
            company_id=cid, 
            module_id=mid, 
            notif_type='success', 
            message=f"Primary file generated successfully. {result['total_unique']} unique values found.", 
            link=None, 
            user_id=current_user.get('user_id') if current_user else None
        )
        
        return {
            "success": True,
            "message": f"Primary data generated successfully. {result['total_unique']} unique values found.",
            "primary_file": result['filename'],
            "total_unique": result['total_unique'],
            "preview": result['preview'],
            "fields": result.get('fields', []),
            "columns": result.get('columns', []),
            "warnings": result.get('warnings', []),
            "download_url": f"/api/primary/download/{result['filename']}"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/primary/preview-live")
async def preview_primary_live(
    file_id: str = Form(...),
    sheet_name: str = Form(...),
    column_name: str = Form(...),
    header_row: str = Form("1"),
    fields: str = Form(None)
):
    """Live preview of primary data without saving to disk."""
    try:
        header_row_int = int(header_row) if header_row else 1
    except ValueError:
        raise HTTPException(status_code=422, detail="header_row must be a valid integer")
    
    parsed_fields = None
    if fields:
        try:
            parsed_fields = json.loads(fields)
        except json.JSONDecodeError:
            raise HTTPException(status_code=422, detail="fields must be valid JSON array")
    
    try:
        result = preview_primary_data(file_id, sheet_name, column_name, header_row_int, parsed_fields)
        return {
            "success": True,
            "preview": result['preview'],
            "total_unique": result['total_unique']
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/primary/field-columns")
async def get_primary_field_columns_endpoint(current_user: Optional[dict] = Depends(get_optional_user)):
    """Get Phase 1 field column definitions for Phase 2/4 integration."""
    try:
        cid, mid = _get_context(current_user)
        field_columns = get_primary_field_columns(cid, mid)
        return {"success": True, "fields": field_columns}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/primary/download/{filename}")
async def download_primary(filename: str, current_user: Optional[dict] = Depends(get_optional_user)):
    try:
        cid, mid = _get_context(current_user)
        file_path = get_primary_file_path(filename, cid, mid)
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")
        return FileResponse(file_path, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/primary/files")
async def list_primary(current_user: Optional[dict] = Depends(get_optional_user)):
    try:
        cid, mid = _get_context(current_user)
        files = list_primary_files(cid, mid)
        return {"success": True, "files": files}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/primary/preview/{filename}")
async def preview_primary(filename: str, current_user: Optional[dict] = Depends(get_optional_user)):
    try:
        cid, mid = _get_context(current_user)
        file_path = get_primary_file_path(filename, cid, mid)
        if not os.path.exists(file_path):
            logger.warning(f"Primary preview requested for '{filename}' but file not found at {file_path}. "
                           f"Company: {cid}, Module: {mid}")
            raise HTTPException(status_code=404, detail=f"Primary data file not found. Please re-save Phase 1 to regenerate the file. (Searched: {filename})")
        
        if filename.lower().endswith('.csv'):
            df = pd.read_csv(file_path, header=0)
        else:
            # Try 'working' (lowercase - what generate_primary_data writes) first,
            # then fall back to 'Working' (capital W) for backward compatibility,
            # finally any first sheet
            try:
                df = pd.read_excel(file_path, sheet_name='working', header=0)
            except Exception:
                try:
                    df = pd.read_excel(file_path, sheet_name='Working', header=0)
                except Exception:
                    df = pd.read_excel(file_path, sheet_name=0, header=0)
        
        # Safely convert pandas types to python native types to prevent FastAPI serialization crashes
        df_preview = df.head(10).fillna("")
        preview_data_raw = df_preview.to_dict(orient='records')
        import json
        preview_data = json.loads(json.dumps(preview_data_raw, default=str))
        
        return {
            "success": True,
            "preview": preview_data,
            "total_rows": len(df),
            "columns": df.columns.tolist()
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Primary preview error for '{filename}': {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============ RULE APIs ============

@app.post("/api/rules")
async def create_rule(
    phase: int = Form(...),
    config: str = Form(...),
    name: Optional[str] = Form(None),
    processing_type: Optional[str] = Form('both'),
    current_user: Optional[dict] = Depends(get_optional_user)
):
    try:
        cid, mid = _get_context(current_user)
        try:
            config_parsed = json.loads(config) if isinstance(config, str) else config
        except:
            config_parsed = config
            
        rule_id = save_rule(phase, config_parsed, name=name, processing_type=processing_type, company_id=cid, module_id=mid)
        return {"success": True, "rule_id": rule_id, "message": "Rule saved successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/rules/{phase}")
async def api_get_rules_by_phase(phase: int, current_user: Optional[dict] = Depends(get_optional_user)):
    cid, mid = _get_context(current_user)
    rules = get_rules_by_phase(phase, company_id=cid, module_id=mid)
    for rule in rules:
        if rule.get('config') and isinstance(rule['config'], str):
            try:
                parsed = json.loads(rule['config'])
                if isinstance(parsed, str):
                    rule['config'] = parsed
            except:
                pass
    return {"success": True, "rules": rules}

@app.get("/api/rules")
async def api_get_all_rules(current_user: Optional[dict] = Depends(get_optional_user)):
    cid, mid = _get_context(current_user)
    rules = get_all_rules(company_id=cid, module_id=mid)
    for rule in rules:
        if rule.get('config') and isinstance(rule['config'], str):
            try:
                parsed = json.loads(rule['config'])
                if isinstance(parsed, str):
                    rule['config'] = parsed
            except:
                pass
    return {"success": True, "rules": rules}

@app.delete("/api/rules/{rule_id}")
async def api_delete_rule(rule_id: int):
    try:
        delete_rule(rule_id)
        return {"success": True, "message": "Rule deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============ OPTIMIZED BACKGROUND PROCESSING ============

def process_rules_background():
    """Optimized rule processing with vectorized operations and caching"""
    global processing_status
    
    start_time = time.time()
    
    try:
        # Extract company/module context from processing_status
        cid = processing_status.get("company_id")
        mid = processing_status.get("module_id")
        uid = processing_status.get("user_id")
        
        # Load rules scoped to company/module if authenticated context exists
        conn = get_db_connection()
        if cid is not None and mid is not None:
            all_rules = conn.execute(
                "SELECT * FROM rules WHERE company_id = ? AND module_id = ? ORDER BY phase, id",
                (cid, mid)
            ).fetchall()
        else:
            all_rules = conn.execute("SELECT * FROM rules ORDER BY phase, id").fetchall()
        conn.close()
        
        if not all_rules:
            processing_status["result"] = {"success": False, "message": "No rules configured"}
            processing_status["is_processing"] = False
            processing_status["progress"] = "completed"
            return
        
        phase1_rules = [dict(r) for r in all_rules if r['phase'] == 1]
        phase2_rules = [dict(r) for r in all_rules if r['phase'] == 2]
        phase3_rules = [dict(r) for r in all_rules if r['phase'] == 3]
        
        # FIX: Only use the most recent rule record for each phase!
        # The database keeps historical saves, but we should only execute the latest configuration.
        phase1_rules = [phase1_rules[-1]] if phase1_rules else []
        phase2_rules = [phase2_rules[-1]] if phase2_rules else []
        phase3_rules = [phase3_rules[-1]] if phase3_rules else []
        
        if not phase1_rules:
            processing_status["result"] = {"success": False, "message": "No Phase 1 (Primary Data) rule configured"}
            processing_status["is_processing"] = False
            processing_status["progress"] = "completed"
            return
        
        with processing_lock:
            processing_status["progress"] = "loading_primary"
        
        # Get primary data
        p1_config = json.loads(phase1_rules[-1]['config'])
        if isinstance(p1_config, str):
            try: p1_config = json.loads(p1_config)
            except: pass
        if not isinstance(p1_config, dict): p1_config = {}
        p1_fields = p1_config.get('fields', [])
        p1_fields = p1_config.get('fields', [])
        
        primary_generated_filename = p1_config.get('primary_file')
        # Default fallback if a Phase 2 rule doesn't specify its primary_column
        primary_key_column = p1_config.get('column', 'Order ID')
        
        primary_df = None
        if primary_generated_filename:
            primary_path = get_primary_file_path(primary_generated_filename, cid, mid)
            if os.path.exists(primary_path):
                primary_df = read_primary_file(primary_generated_filename, cid, mid)
        
        if primary_df is None:
            processing_status["result"] = {"success": False, "message": "Primary data file not found"}
            processing_status["is_processing"] = False
            processing_status["progress"] = "completed"
            return
        
        # ---- NEW: Apply Source File Name Filter ----
        filter_sources = processing_status.get("filter_sources")
        if filter_sources and len(filter_sources) > 0:
            before_count = len(primary_df)
            primary_df = primary_df[primary_df['Source_File_Name'].astype(str).isin(filter_sources)]
            after_count = len(primary_df)
            logger.info(f"Filtered primary data: {before_count} -> {after_count} rows "
                        f"based on {len(filter_sources)} selected source files")
            
            if after_count == 0:
                processing_status["result"] = {
                    "success": False,
                    "message": f"No primary data rows matched the {len(filter_sources)} selected source files"
                }
                processing_status["is_processing"] = False
                processing_status["progress"] = "completed"
                return
        # ---- END NEW ----
        
        with processing_lock:
            processing_status["progress"] = "phase2"
        
        # OPTIMIZED: Load all secondary files ONCE with connection pooling
        file_cache = {}
        
        def get_file_df(file_id, sheet_name):
            cache_key = f"{file_id}|{sheet_name}"
            if cache_key in file_cache:
                return file_cache[cache_key]
            
            # Handle master files: file_id is "master_{folder_id}"
            if isinstance(file_id, str) and file_id.startswith('master_'):
                try:
                    folder_id = int(file_id.replace('master_', ''))
                    master = get_master_file(folder_id)
                    if not master:
                        logger.error(f"Master file not found for folder {folder_id}")
                        return None
                    
                    conn_duck = duckdb.connect(master['db_path'], read_only=True)
                    df = conn_duck.execute("SELECT * FROM master_data").fetchdf()
                    conn_duck.close()
                    
                    # Convert to string to prevent mixed type issues
                    # Strip .0 from end of floats before converting to string, and handle NaNs
                    for col in df.columns:
                        if pd.api.types.is_numeric_dtype(df[col]):
                            df[col] = df[col].astype(str).str.replace(r'\.0$', '', regex=True).replace('nan', '')
                        else:
                            df[col] = df[col].astype(str).replace('nan', '')
                    
                    df.columns = [str(col).strip() for col in df.columns]
                    file_cache[cache_key] = df
                    return df
                except Exception as e:
                    logger.error(f"Error reading master file {file_id}: {e}")
                    return None
            
            # Try cache first
            cached = get_cached_file_info(file_id)
            if cached:
                path = cached['file_path']
                fmt = cached['format']
                header_row = cached.get('header_row', 1)
            else:
                conn = get_db_connection()
                file_info = conn.execute("SELECT file_path, format, header_row FROM files WHERE id = ?", (file_id,)).fetchone()
                conn.close()
                
                if not file_info:
                    return None
                
                path = file_info['file_path']
                if not os.path.exists(path):
                    filename = os.path.basename(path)
                    new_path = os.path.join(UPLOAD_DIR, filename)
                    if os.path.exists(new_path):
                        path = new_path

                fmt = file_info['format'].upper() if file_info['format'] else ''
                header_row = file_info['header_row'] if ('header_row' in file_info.keys() and file_info['header_row']) else 1
                set_cached_file_info(file_id, {'file_path': path, 'format': fmt, 'header_row': header_row})
            
            header_idx = max(0, header_row - 1)
            try:
                if not os.path.exists(path):
                    filename = os.path.basename(path)
                    new_path = os.path.join(UPLOAD_DIR, filename)
                    if os.path.exists(new_path):
                        path = new_path

                if fmt == 'CSV':
                    df = pd.read_csv(path, header=header_idx, dtype=str, low_memory=False)
                else:
                    try:
                        df = pd.read_excel(path, sheet_name=sheet_name, header=header_idx, dtype=str)
                    except Exception:
                        df = pd.read_excel(path, sheet_name=0, header=header_idx, dtype=str)
            except Exception as e:
                logger.error(f"Error reading file {file_id} sheet '{sheet_name}': {e}")
                return None
            
            # FIX: Strip whitespace from column names to prevent mismatches
            # e.g., "Amount " (with trailing space) vs "Amount"
            df.columns = [str(col).strip() for col in df.columns]
            file_cache[cache_key] = df
            return df
        
        # Process Phase 2 rules
        output_columns = {}
        
        # FIX BUG #4: Track created columns to detect overwrites
        _phase2_created_columns = set()
        
        for rule in phase2_rules:
            config = json.loads(rule['config'])
            if isinstance(config, str):
                try: config = json.loads(config)
                except: pass
            
            rules_list = config if isinstance(config, list) else [config]
            
            for r in rules_list:
                rule_type = r.get('rule_type', 'match')
                output_col = r.get('output_column', '')
                col_name = r.get('column_name', '')
                default_val = str(r.get('default_value', '')).strip()
                
                # FIX BUG #4: Warn about column overwrites
                if col_name and col_name in _phase2_created_columns:
                    logger.warning(f"Phase 2 column '{col_name}' is being overwritten by multiple rules. Only the last rule's value will be retained.")
                if col_name:
                    _phase2_created_columns.add(col_name)
                
                if output_col and col_name:
                    output_columns[output_col] = col_name
                
                # Handle calculation rules
                if rule_type == 'calculation':
                    operation = r.get('operation', 'addition')
                    columns = r.get('columns', [])
                    
                    if not columns:
                        first_col = r.get('first_column')
                        second_col = r.get('second_column')
                        if first_col and second_col:
                            columns = [first_col, second_col]
                    
                    valid_cols = [c for c in columns if c and c in primary_df.columns]
                    
                    if len(valid_cols) >= 2:
                        result = pd.to_numeric(primary_df[valid_cols[0]], errors='coerce').fillna(0)
                        
                        for col in valid_cols[1:]:
                            next_vals = pd.to_numeric(primary_df[col], errors='coerce').fillna(0)
                            if operation == 'addition':
                                result = result + next_vals
                            elif operation == 'subtraction':
                                result = result - next_vals
                            elif operation == 'multiply':
                                result = result * next_vals
                            elif operation == 'divide':
                                result = result.div(next_vals.replace(0, float('nan'))).fillna(0)
                        
                        primary_df[col_name] = result
                    elif len(valid_cols) == 1:
                        primary_df[col_name] = primary_df[valid_cols[0]]
                    continue
                
                # Handle match/sumif/countif/addition/subtraction rules
                sec_file_id = r.get('secondary_file')
                sec_sheet = r.get('secondary_sheet', 'Sheet1')
                sec_col = r.get('secondary_column')
                if isinstance(sec_col, str): sec_col = sec_col.strip()
                ext_file_id = r.get('extract_file')
                ext_sheet = r.get('extract_sheet', 'Sheet1')
                ext_col = r.get('extract_column')
                if isinstance(ext_col, str): ext_col = ext_col.strip()
                
                sec_df = get_file_df(sec_file_id, sec_sheet) if sec_file_id and not str(sec_file_id).startswith('primary_') else None
                
                if sec_df is not None and sec_col in sec_df.columns:
                    # FIX: Use copy() to avoid modifying shared dataframes
                    ext_df = sec_df.copy()
                    if ext_file_id and ext_file_id != sec_file_id and not str(ext_file_id).startswith('primary_'):
                        ext_df = get_file_df(ext_file_id, ext_sheet)
                        if ext_df is None:
                            logger.error(f"Extract file {ext_file_id} sheet '{ext_sheet}' could not be loaded")
                            continue
                    
                    if ext_df is not None and ext_col in ext_df.columns:
                        sec_df['_match_key'] = sec_df[sec_col].astype(str).str.strip()
                        
                        rule_primary_col = r.get('primary_column')
                        if not rule_primary_col or rule_primary_col not in primary_df.columns:
                            rule_primary_col = primary_key_column
                            if rule_primary_col not in primary_df.columns and 'Order ID' in primary_df.columns:
                                rule_primary_col = 'Order ID'
                                
                        primary_df['_match_key'] = primary_df[rule_primary_col].astype(str).str.strip()
                        
                        if rule_type == 'match':
                            # FIX: Match by key when secondary and extract are different files
                            if ext_file_id and ext_file_id != sec_file_id:
                                # Different files: build lookup from extract file using its own key column
                                ext_match_col = r.get('secondary_column')  # Use same match logic
                                if ext_match_col in ext_df.columns:
                                    ext_df['_ext_match_key'] = ext_df[ext_match_col].astype(str).str.strip()
                                    lookup = dict(zip(ext_df['_ext_match_key'], ext_df[ext_col]))
                                    if '_ext_match_key' in ext_df.columns:
                                        del ext_df['_ext_match_key']
                                else:
                                    # Fallback: zip by row position (less reliable)
                                    lookup = dict(zip(sec_df['_match_key'], ext_df[ext_col]))
                            else:
                                # Same file: direct zip by row position is correct
                                lookup = dict(zip(sec_df['_match_key'], ext_df[ext_col]))
                            
                            primary_df[col_name] = primary_df['_match_key'].map(lookup)
                            if default_val:
                                primary_df[col_name] = primary_df[col_name].fillna(default_val)

                        
                        elif rule_type == 'sumif':
                            if ext_df[ext_col].dtype == object:
                                ext_df['_sum_val'] = pd.to_numeric(ext_df[ext_col].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
                            else:
                                ext_df['_sum_val'] = pd.to_numeric(ext_df[ext_col], errors='coerce').fillna(0)

                            # FIX: When different files, merge on match key before summing
                            if ext_file_id and ext_file_id != sec_file_id and sec_col in ext_df.columns:
                                ext_df['_match_key'] = ext_df[sec_col].astype(str).str.strip()
                                merged = sec_df[['_match_key']].merge(
                                    ext_df[['_match_key', '_sum_val']], 
                                    on='_match_key', 
                                    how='left'
                                )
                                sum_lookup = merged.groupby('_match_key')['_sum_val'].sum().to_dict()
                            else:
                                sec_df['_sum_val'] = ext_df['_sum_val']
                                sum_lookup = sec_df.groupby('_match_key')['_sum_val'].sum().to_dict()
                                
                            primary_df[col_name] = primary_df['_match_key'].map(sum_lookup)
                            if default_val:
                                primary_df[col_name] = primary_df[col_name].fillna(default_val)
                            else:
                                primary_df[col_name] = primary_df[col_name].fillna(0)

                            if '_sum_val' in sec_df.columns:
                                del sec_df['_sum_val']
                            if '_sum_val' in ext_df.columns:
                                del ext_df['_sum_val']
                        
                        elif rule_type == 'countif':
                            # COUNTIF: Count occurrences of each match key in secondary file
                            count_lookup = sec_df.groupby('_match_key').size().to_dict()
                            
                            primary_df[col_name] = primary_df['_match_key'].map(count_lookup)
                            if default_val:
                                primary_df[col_name] = primary_df[col_name].fillna(default_val)
                            else:
                                primary_df[col_name] = primary_df[col_name].fillna(0).astype(int)

                        
                        elif rule_type == 'addition':
                            vals = ext_df[ext_col]
                            if vals.dtype == object:
                                vals = vals.astype(str).str.replace(',', '')
                                
                            primary_df[col_name] = primary_df['_match_key'].map(
                                dict(zip(sec_df['_match_key'], pd.to_numeric(vals, errors='coerce')))
                            )
                            if default_val:
                                primary_df[col_name] = primary_df[col_name].fillna(default_val)
                            else:
                                primary_df[col_name] = primary_df[col_name].fillna(0)
                        
                        elif rule_type == 'subtraction':
                            vals = ext_df[ext_col]
                            if vals.dtype == object:
                                vals = vals.astype(str).str.replace(',', '')
                                
                            primary_df[col_name] = primary_df['_match_key'].map(
                                dict(zip(sec_df['_match_key'], pd.to_numeric(vals, errors='coerce')))
                            )
                            if default_val:
                                primary_df[col_name] = primary_df[col_name].fillna(default_val)
                            else:
                                primary_df[col_name] = primary_df[col_name].fillna(0)
                        
                        # FIX: Safe cleanup of temp columns
                        for col in ['_match_key', '_sum_val', '_ext_match_key']:
                            try:
                                if col in sec_df.columns:
                                    del sec_df[col]
                            except: pass
                            try:
                                if col in ext_df.columns:
                                    del ext_df[col]
                            except: pass
                            try:
                                if col in primary_df.columns:
                                    del primary_df[col]
                            except: pass
        
        # ============================================================
        # COLUMN CONDITION EXTRACTION (after Phase 2, before Phase 3)
        # ============================================================
        for r in rules_list:
            col_cond = r.get('column_condition')
            if not col_cond or not col_cond.get('enabled'):
                continue
            
            target_col = col_cond.get('extract_target_column', '')
            source_col = col_cond.get('extract_source_column', '')
            logic = col_cond.get('logic', 'AND')
            conditions = col_cond.get('conditions', [])
            
            if not target_col or not source_col or not conditions:
                continue
            
            # Ensure both columns exist in primary_df
            if target_col not in primary_df.columns:
                logger.warning(f"Column condition skipped: target column '{target_col}' not found")
                continue
            if source_col not in primary_df.columns:
                logger.warning(f"Column condition skipped: source column '{source_col}' not found")
                continue
            
            # Build boolean mask
            if logic == 'AND':
                mask = pd.Series([True] * len(primary_df), index=primary_df.index)
            else:  # OR
                mask = pd.Series([False] * len(primary_df), index=primary_df.index)
            
            for cond in conditions:
                cond_col = cond.get('column', '')
                operator = cond.get('operator', '')
                value = cond.get('value', '')
                
                if not cond_col or cond_col not in primary_df.columns:
                    if logic == 'AND':
                        mask = pd.Series([False] * len(primary_df), index=primary_df.index)
                    break
                
                col_data = primary_df[cond_col].astype(str).str.strip()
                col_data_raw = primary_df[cond_col]
                
                try:
                    if operator == 'equal_to':
                        try:
                            col_numeric = pd.to_numeric(col_data, errors='coerce')
                            val_numeric = pd.to_numeric(value, errors='coerce')
                            numeric_match = (col_numeric - val_numeric).abs() <= 1e-9
                            string_match = col_data.str.lower() == str(value).lower()
                            cond_mask = numeric_match | string_match
                        except:
                            cond_mask = col_data.str.lower() == str(value).lower()
                    
                    elif operator == 'not_equal_to':
                        try:
                            col_numeric = pd.to_numeric(col_data, errors='coerce')
                            val_numeric = pd.to_numeric(value, errors='coerce')
                            numeric_match = (col_numeric - val_numeric).abs() <= 1e-9
                            string_match = col_data.str.lower() == str(value).lower()
                            cond_mask = ~(numeric_match | string_match)
                        except:
                            cond_mask = col_data.str.lower() != str(value).lower()
                    
                    elif operator == 'zero_or_blank':
                        numeric_zeros = pd.to_numeric(col_data, errors='coerce') == 0
                        blanks = (col_data == '') | col_data_raw.isna() | col_data.isin(['nan', 'null', 'none', 'na', 'n/a', '-'])
                        cond_mask = numeric_zeros | blanks
                    
                    elif operator == 'no_zero_or_no_blank':
                        numeric_zeros = pd.to_numeric(col_data, errors='coerce') == 0
                        blanks = (col_data == '') | col_data_raw.isna() | col_data.isin(['nan', 'null', 'none', 'na', 'n/a', '-'])
                        cond_mask = ~(numeric_zeros | blanks)
                    
                    elif operator == 'greater_than':
                        cond_mask = pd.to_numeric(col_data, errors='coerce') > float(value)
                    
                    elif operator == 'smaller_than':
                        cond_mask = pd.to_numeric(col_data, errors='coerce') < float(value)
                    
                    elif operator == 'contain':
                        cond_mask = col_data.str.contains(str(value), na=False, case=False)
                    
                    elif operator == 'not_contain':
                        cond_mask = ~col_data.str.contains(str(value), na=False, case=False)
                    
                    elif operator == 'begin_with':
                        cond_mask = col_data.str.lower().str.startswith(str(value).lower())
                    
                    elif operator == 'end_with':
                        cond_mask = col_data.str.lower().str.endswith(str(value).lower())
                    
                    else:
                        cond_mask = pd.Series([False] * len(primary_df), index=primary_df.index)
                
                except Exception as e:
                    logger.error(f"Column condition operator '{operator}' failed: {e}")
                    cond_mask = pd.Series([False] * len(primary_df), index=primary_df.index)
                
                # Combine with main mask
                if logic == 'AND':
                    mask &= cond_mask.fillna(False)
                else:
                    mask |= cond_mask.fillna(False)
            
            # Apply extraction: copy source column data to target column where mask is True
            matched_count = mask.sum()
            if matched_count > 0:
                primary_df.loc[mask, target_col] = primary_df.loc[mask, source_col]
                logger.info(f"Column condition applied: copied {matched_count} rows from '{source_col}' to '{target_col}'")
            else:
                logger.info(f"Column condition: no rows matched for '{target_col}'")
        
        with processing_lock:
            processing_status["progress"] = "phase3"
        
        # Process Phase 3: Remarks & Actions (VECTORIZED)
        phase3_output_columns = {}
        
        if phase3_rules:
            # FIX #1: Process ALL Phase 3 rules, not just the last one
            for p3_rule in phase3_rules:
                phase3_config = json.loads(p3_rule['config'])
                if isinstance(phase3_config, str):
                    try: phase3_config = json.loads(phase3_config)
                    except: pass
                
                groups = phase3_config if isinstance(phase3_config, list) else []
                
                for group in groups:
                    col_name = group.get('column_name', '')
                    default_remark = group.get('default_remark', '')
                    remark_rules = group.get('remark_rules', [])
                    
                    output_col = group.get('output_column', '')
                    if not col_name or not remark_rules:
                        continue
                    
                    if output_col and col_name:
                        phase3_output_columns[output_col] = col_name
                    
                    primary_df[col_name] = default_remark if default_remark else ''
                    
                    for remark_rule in reversed(remark_rules):
                        remark_text = remark_rule.get('remark', '')
                        conditions = remark_rule.get('conditions', [])
                        
                        if not conditions:
                            continue
                        
                        mask = pd.Series([True] * len(primary_df), index=primary_df.index)
                        
                        for cond in conditions:
                            cond_col = cond.get('column', '')
                            operator = cond.get('operator', '')
                            value = cond.get('value', '')
                            value_min = cond.get('value_min', '')
                            value_max = cond.get('value_max', '')
                            
                            # FIX #5: Skip empty condition columns instead of failing silently
                            if not cond_col or cond_col not in primary_df.columns:
                                mask = pd.Series([False] * len(primary_df), index=primary_df.index)
                                break
                            
                            # FIX #3: Handle blank properly - check for NaN, empty, 'nan', 'null', 'none'
                            col_data_raw = primary_df[cond_col]
                            col_data = col_data_raw.astype(str).str.strip()
                            
                            if operator == 'blank':
                                # Match empty strings, NaN values, and common null string representations
                                is_blank = (col_data == '') | col_data_raw.isna() | col_data.isin(['nan', 'null', 'none', 'na', 'n/a', '-'])
                                mask &= is_blank
                            elif operator == 'equal_to':
                                try:
                                    col_numeric = pd.to_numeric(col_data, errors='coerce')
                                    val_numeric = pd.to_numeric(value, errors='coerce')
                                    numeric_match = (col_numeric - val_numeric).abs() <= 1e-9
                                    string_match = col_data.str.lower() == str(value).lower()
                                    mask &= numeric_match.fillna(string_match)
                                except:
                                    mask &= (col_data.str.lower() == str(value).lower())
                            elif operator == 'not_equal_to':
                                try:
                                    col_numeric = pd.to_numeric(col_data, errors='coerce')
                                    val_numeric = pd.to_numeric(value, errors='coerce')
                                    numeric_ne = (col_numeric - val_numeric).abs() > 1e-9
                                    string_ne = col_data.str.lower() != str(value).lower()
                                    mask &= numeric_ne.fillna(string_ne)
                                except:
                                    mask &= (col_data.str.lower() != str(value).lower())
                            elif operator == 'greater_than':
                                try:
                                    col_numeric = pd.to_numeric(col_data, errors='coerce')
                                    val_numeric = pd.to_numeric(value, errors='coerce')
                                    # FIX #2: Don't break on text columns - skip this condition for non-numeric
                                    if pd.isna(val_numeric):
                                        mask &= pd.Series([False] * len(primary_df), index=primary_df.index)
                                    else:
                                        mask &= col_numeric > val_numeric
                                except:
                                    # Skip this condition on text data, don't kill the whole rule
                                    mask &= pd.Series([False] * len(primary_df), index=primary_df.index)
                            elif operator == 'smaller_than':
                                try:
                                    col_numeric = pd.to_numeric(col_data, errors='coerce')
                                    val_numeric = pd.to_numeric(value, errors='coerce')
                                    # FIX #2: Don't break on text columns
                                    if pd.isna(val_numeric):
                                        mask &= pd.Series([False] * len(primary_df), index=primary_df.index)
                                    else:
                                        mask &= col_numeric < val_numeric
                                except:
                                    mask &= pd.Series([False] * len(primary_df), index=primary_df.index)
                            elif operator == 'between':
                                # FIX #4: Skip "between" if min or max is empty
                                if not value_min or not value_max:
                                    mask &= pd.Series([False] * len(primary_df), index=primary_df.index)
                                    continue
                                try:
                                    col_numeric = pd.to_numeric(col_data, errors='coerce')
                                    min_v = pd.to_numeric(value_min, errors='coerce')
                                    max_v = pd.to_numeric(value_max, errors='coerce')
                                    if pd.isna(min_v) or pd.isna(max_v):
                                        mask &= pd.Series([False] * len(primary_df), index=primary_df.index)
                                    else:
                                        mask &= (col_numeric >= min_v) & (col_numeric <= max_v)
                                except:
                                    mask &= pd.Series([False] * len(primary_df), index=primary_df.index)
                            elif operator == 'begin_with':
                                mask &= col_data.str.lower().str.startswith(str(value).lower())
                            elif operator == 'end_with':
                                mask &= col_data.str.lower().str.endswith(str(value).lower())
                            elif operator == 'contain':
                                mask &= col_data.str.lower().str.contains(str(value).lower(), na=False)
                            elif operator == 'not_contain':
                                mask &= ~col_data.str.lower().str.contains(str(value).lower(), na=False)
                        
                        if mask.any():
                            primary_df.loc[mask, col_name] = remark_text
        
        # ===== PHASE 4: SUMMARY & PIVOT =====
        with processing_lock:
            processing_status["progress"] = "phase4"
        
        # Load Phase 4 summaries scoped to company/module
        conn = get_db_connection()
        if cid is not None and mid is not None:
            phase4_rules = conn.execute(
                "SELECT * FROM rules WHERE phase = 4 AND company_id = ? AND module_id = ? ORDER BY id",
                (cid, mid)
            ).fetchall()
        else:
            phase4_rules = conn.execute("SELECT * FROM rules WHERE phase = 4 ORDER BY id").fetchall()
        conn.close()
        
        summary_sheets = {}
        
        # Track Phase 4 errors for reporting to user
        phase4_errors = []
        
        logger.info(f"PHASE 4 DIAGNOSTIC: Found {len(phase4_rules)} Phase 4 rule(s) in database. "
                    f"company_id={cid}, module_id={mid}")
        
        if phase4_rules:
            for p4_rule in phase4_rules:
                rule_name = dict(p4_rule).get('name', 'unknown')
                try:
                    summary_config = json.loads(p4_rule['config'])
                    
                    logger.info(f"PHASE 4 DIAGNOSTIC: Processing rule '{rule_name}', "
                                f"include_in_final={summary_config.get('include_in_final', 'not set')}, "
                                f"row_fields={summary_config.get('row_fields', [])}, "
                                f"value_fields={[v.get('column','?') for v in summary_config.get('value_fields', [])]}")
                    
                    if summary_config.get('include_in_final') is False:
                        logger.warning(f"PHASE 4 DIAGNOSTIC: Rule '{rule_name}' SKIPPED — include_in_final is False")
                        phase4_errors.append(f"Phase 4 summary '{rule_name}' skipped: include_in_final is set to False")
                        continue
                    
                    row_fields = summary_config.get('row_fields', [])
                    column_fields = summary_config.get('column_fields', [])
                    value_fields = summary_config.get('value_fields', [])
                    filter_fields = summary_config.get('filter_fields', [])
                    
                    if not value_fields:
                        logger.info(f"Phase 4 summary '{rule_name}' skipped: No value fields configured")
                        continue
                    
                    # --- COLUMN NAME MAPPING: Map generic names to actual primary_df columns ---
                    # Phase 1 can rename the primary key column (e.g., 'Order ID' → 'Sale Order Number')
                    # and field columns (e.g., 'Sales Amount' → 'Total Amount')
                    primary_key_col = p1_config.get('column', 'Order ID')
                    
                    # Build lookup: generic name → actual column name in primary_df
                    col_alias_map = {}
                    if primary_key_col != 'Order ID' and primary_key_col in primary_df.columns:
                        col_alias_map['Order ID'] = primary_key_col
                        col_alias_map['Primary_Value'] = primary_key_col
                    elif primary_key_col == 'Order ID':
                        col_alias_map['Primary_Value'] = 'Order ID'

                        col_alias_map['Primary_Value'] = primary_key_col
                    elif primary_key_col == 'Order ID':
                        col_alias_map['Primary_Value'] = 'Order ID'

                    
                    # Build alias map from Phase 1 fields: source_column → display name
                    # This ensures references to original source column names resolve to the renamed output columns
                    for f in p1_fields:
                        if f.get('source_column') and f.get('name') and f['source_column'] != f['name']:
                            col_alias_map[f['source_column']] = f['name']
                    
                    def resolve_column(col_name):
                        """Resolve a column reference to an actual column name in primary_df"""
                        if col_name in primary_df.columns:
                            return col_name
                        # Try alias mapping (key renames + field source→name mappings)
                        if col_name in col_alias_map:
                            return col_alias_map[col_name]
                        # Try case-insensitive match
                        for actual_col in primary_df.columns:
                            if actual_col.lower() == col_name.lower():
                                return actual_col
                        # Try partial match — prefer longer matches (most specific) first
                        # This prevents "Amount" from matching "Tax Amount" when "Total Amount" also exists
                        candidate_matches = []
                        for actual_col in primary_df.columns:
                            if col_name.lower() in actual_col.lower() or actual_col.lower() in col_name.lower():
                                shared_words = set(col_name.lower().split()) & set(actual_col.lower().split())
                                if len(shared_words) > 0:
                                    # Score: favor matches where the column name is closer in length
                                    score = len(shared_words) - abs(len(col_name) - len(actual_col)) * 0.01
                                    candidate_matches.append((score, actual_col))
                        if candidate_matches:
                            # Sort by score descending (best match first)
                            candidate_matches.sort(key=lambda x: x[0], reverse=True)
                            best_match = candidate_matches[0][1]
                            col_alias_map[col_name] = best_match
                            return best_match
                            
                        # Legacy fallback: if Phase 4 is looking for 'Sales Amount' (old default)
                        # but it's not found, try mapping it to the first Phase 1 field
                        logger.info(f"PHASE 4 DIAGNOSTIC: resolve_column('{col_name}') fallback check. p1_fields={p1_fields}, primary_df.columns={list(primary_df.columns)}")
                        if col_name.strip().lower() == 'sales amount' and len(p1_fields) > 0:
                            first_p1_field = p1_fields[0].get('name')
                            logger.info(f"PHASE 4 DIAGNOSTIC: first_p1_field='{first_p1_field}'")
                            if first_p1_field and first_p1_field in primary_df.columns:
                                col_alias_map[col_name] = first_p1_field
                                return first_p1_field
                                
                        return col_name
                    
                    # Resolve all field references
                    row_fields = [resolve_column(rf) for rf in row_fields]
                    column_fields = [resolve_column(cf) for cf in column_fields]
                    for vf in value_fields:
                        vf['column'] = resolve_column(vf['column'])
                    for ff in filter_fields:
                        if ff.get('column'):
                            ff['column'] = resolve_column(ff['column'])
                    
                    # Validate that all referenced columns exist in primary_df
                    all_ref_columns = row_fields + column_fields + [v['column'] for v in value_fields] + [f.get('column', '') for f in filter_fields]
                    missing_cols = [c for c in all_ref_columns if c and c not in primary_df.columns]
                    if missing_cols:
                        msg = f"Phase 4 summary '{rule_name}' skipped: Columns not found in data: {', '.join(missing_cols)}"
                        logger.warning(msg)
                        phase4_errors.append(msg)
                        continue
                    
                    # Apply filters
                    filtered_df = primary_df.copy()
                    for filt in filter_fields:
                        col = filt.get('column', '')
                        op = filt.get('operator', '')
                        val = filt.get('value', '')
                        
                        if col and col in filtered_df.columns and op:
                            if op == 'equal_to':
                                filtered_df = filtered_df[filtered_df[col].astype(str) == str(val)]
                            elif op == 'not_equal_to':
                                filtered_df = filtered_df[filtered_df[col].astype(str) != str(val)]
                            elif op == 'blank':
                                filtered_df = filtered_df[filtered_df[col].astype(str).str.strip() == '']
                            elif op == 'greater_than':
                                try:
                                    filtered_df = filtered_df[pd.to_numeric(filtered_df[col], errors='coerce') > float(val)]
                                except: pass
                            elif op == 'smaller_than':
                                try:
                                    filtered_df = filtered_df[pd.to_numeric(filtered_df[col], errors='coerce') < float(val)]
                                except: pass
                    
                    # Check if filtered data is empty
                    if len(filtered_df) == 0:
                        msg = f"Phase 4 summary '{rule_name}' skipped: No data rows after applying filters"
                        logger.warning(msg)
                        phase4_errors.append(msg)
                        continue
                    
                    # Build pivot: only include columns that actually exist in the dataframe
                    valid_values = [v['column'] for v in value_fields if v['column'] in filtered_df.columns]
                    valid_aggfuncs = {v['column']: v.get('aggregation', 'sum') for v in value_fields if v['column'] in filtered_df.columns}
                    
                    if not valid_values:
                        msg = f"Phase 4 summary '{rule_name}' skipped: No valid value columns found in filtered data. Available: {list(filtered_df.columns)}"
                        logger.warning(msg)
                        phase4_errors.append(msg)
                        continue
                    
                    # CRITICAL FIX: Ensure value columns are properly converted to numeric
                    # and replace empty strings with 0 to prevent row exclusion during aggregation
                    for val_col in valid_values:
                        # Convert to numeric, replacing non-numeric with 0 (not NaN)
                        # This ensures rows with blank/invalid values are INCLUDED in sums with value 0
                        filtered_df[val_col] = pd.to_numeric(filtered_df[val_col].astype(str).str.replace(',', '').str.strip(), errors='coerce').fillna(0)
                    
                    # CRITICAL FIX: Replace ALL null/NaN/empty/whitespace values in grouping columns
                    # with the literal string 'Blank' so pivot table includes them as a real row/column
                    for group_col in (row_fields or []) + (column_fields or []):
                        if group_col in filtered_df.columns:
                            # Step 1: convert to string
                            s = filtered_df[group_col].astype(str)
                            # Step 2: replace Python 'nan', 'None', 'NaT', empty, whitespace-only with 'Blank'
                            s = s.replace(r'^\s*$', 'Blank', regex=True)
                            s = s.replace(['nan', 'None', 'NaT', '<NA>'], 'Blank')
                            s = s.where(s != 'Blank', 'Blank')  # ensure consistency
                            # Step 3: also handle actual NaN / None values that .astype(str) turned into 'nan'
                            filtered_df[group_col] = s
                            # Final safety: any remaining nulls → 'Blank'
                            filtered_df[group_col] = filtered_df[group_col].fillna('Blank')
                    
                    pivot_kwargs = {
                        'values': valid_values,
                        'aggfunc': valid_aggfuncs,
                        'margins': True,
                        'margins_name': 'Grand Total',
                        'dropna': False,
                        'fill_value': 0
                    }
                    if row_fields:
                        # Ensure row index columns actually exist
                        valid_rows = [c for c in row_fields if c in filtered_df.columns]
                        if valid_rows:
                            pivot_kwargs['index'] = valid_rows
                    if column_fields:
                        # Ensure column index columns actually exist
                        valid_cols = [c for c in column_fields if c in filtered_df.columns]
                        if valid_cols:
                            pivot_kwargs['columns'] = valid_cols
                    
                    # Proceed only if we have row/col indices or at least values
                    if 'index' not in pivot_kwargs and 'columns' not in pivot_kwargs:
                        # Basic aggregation without groups
                        pivot_kwargs['index'] = [valid_values[0]]  # Fallback to avoid error
                        
                    pivot_table = pd.pivot_table(filtered_df, **pivot_kwargs)
                    pivot_table = pivot_table.fillna(0)
                    
                    # Drop rows where all numeric values are 0 (e.g. Cartesian product artifacts), excluding Grand Total
                    if not pivot_table.empty:
                        num_cols = pivot_table.select_dtypes(include=['number']).columns
                        if len(num_cols) > 0:
                            non_zero_mask = (pivot_table[num_cols] != 0).any(axis=1)
                            if isinstance(pivot_table.index, pd.MultiIndex):
                                try:
                                    non_zero_mask = non_zero_mask | (pivot_table.index.get_level_values(0) == 'Grand Total')
                                except: pass
                            else:
                                if 'Grand Total' in pivot_table.index:
                                    non_zero_mask.loc['Grand Total'] = True
                            pivot_table = pivot_table[non_zero_mask]
                    
                    # Drop rows where all numeric values are 0 (e.g. Cartesian product artifacts), excluding Grand Total
                    if not pivot_table.empty:
                        num_cols = pivot_table.select_dtypes(include=['number']).columns
                        if len(num_cols) > 0:
                            non_zero_mask = (pivot_table[num_cols] != 0).any(axis=1)
                            if isinstance(pivot_table.index, pd.MultiIndex):
                                try:
                                    non_zero_mask = non_zero_mask | (pivot_table.index.get_level_values(0) == 'Grand Total')
                                except: pass
                            else:
                                if 'Grand Total' in pivot_table.index:
                                    non_zero_mask.loc['Grand Total'] = True
                            pivot_table = pivot_table[non_zero_mask]
                    
                    # Log summary stats for debugging
                    total_rows_in_filter = len(filtered_df)
                    grand_total_row = pivot_table.loc['Grand Total'] if 'Grand Total' in pivot_table.index else None
                    logger.info(f"Summary '{p4_rule['name']}': Input rows={total_rows_in_filter}, Pivot shape={pivot_table.shape}")
                    
                    # --- NEW: Rename value columns based on aggregation AND reorder them ---
                    # Build rename dictionary and ordered columns list based on user's value_fields
                    rename_dict = {}
                    ordered_cols = []
                    for v in value_fields:
                        val_col = v['column']
                        if val_col in filtered_df.columns:
                            agg = v.get('aggregation', 'sum').lower()
                            if agg == 'sum': prefix = 'Sum of'
                            elif agg == 'count': prefix = 'Count of'
                            elif agg == 'average': prefix = 'Average of'
                            else: prefix = f"{agg.capitalize()} of"
                            new_name = f"{prefix} {val_col}"
                            rename_dict[val_col] = new_name
                            ordered_cols.append(new_name)
                    
                    custom_primary_label = p1_config.get('column', 'Order ID')
                    
                    if isinstance(pivot_table.columns, pd.MultiIndex):
                        # Level 0 is the value columns. Rename them.
                        new_level_0 = [rename_dict.get(c, c) for c in pivot_table.columns.levels[0]]
                        pivot_table.columns = pivot_table.columns.set_levels(new_level_0, level=0)
                        
                        # Reorder columns based on the user's arrangement of value fields
                        col_order_map = {col: i for i, col in enumerate(ordered_cols)}
                        
                        # Helper to sort levels but keep 'Grand Total' at the end
                        def get_sort_key(k):
                            sk = str(k)
                            return ('zzzzz', sk) if sk == 'Grand Total' else ('', sk)
                            
                        sorted_cols = sorted(pivot_table.columns, key=lambda x: (col_order_map.get(x[0], 999),) + tuple(get_sort_key(k) for k in x[1:]))
                        pivot_table = pivot_table[sorted_cols]
                        
                        # Rename index if it contains 'Primary_Value'
                        if custom_primary_label and custom_primary_label != 'Primary_Value':
                            if isinstance(pivot_table.index, pd.MultiIndex):
                                pivot_table.index.names = [custom_primary_label if n == 'Primary_Value' else n for n in pivot_table.index.names]
                            elif pivot_table.index.name == 'Primary_Value':
                                pivot_table.index.name = custom_primary_label
                                
                        # Flatten multi-level columns
                        pivot_table.columns = [' '.join(col).strip() if isinstance(col, tuple) else str(col) for col in pivot_table.columns.values]
                        
                        # Replace 'Primary_Value' in flat columns
                        if custom_primary_label and custom_primary_label != 'Primary_Value':
                            pivot_table.columns = [str(c).replace('Primary_Value', custom_primary_label) for c in pivot_table.columns]
                            
                    else:
                        pivot_table.rename(columns=rename_dict, inplace=True)
                        
                        # Reorder flat columns
                        final_order = [c for c in ordered_cols if c in pivot_table.columns]
                        if 'Grand Total' in pivot_table.columns:
                            final_order.append('Grand Total')
                        for c in pivot_table.columns:
                            if c not in final_order:
                                final_order.append(c)
                        pivot_table = pivot_table[final_order]
                        
                        # Rename index if it contains 'Primary_Value'
                        if custom_primary_label and custom_primary_label != 'Primary_Value':
                            if isinstance(pivot_table.index, pd.MultiIndex):
                                pivot_table.index.names = [custom_primary_label if n == 'Primary_Value' else n for n in pivot_table.index.names]
                            elif pivot_table.index.name == 'Primary_Value':
                                pivot_table.index.name = custom_primary_label
                    
                    pivot_flat = pivot_table.reset_index()
                    
                    summary_sheets[p4_rule['name']] = {
                        'data': pivot_flat,
                        'config': summary_config
                    }
                    
                except Exception as e:
                    rule_name = dict(p4_rule).get('name', 'unknown')
                    logger.error(f"Phase 4 summary error for {rule_name}: {e}")
        
        with processing_lock:
            processing_status["progress"] = "saving"
        
        # Merge Phase 2 and Phase 3 output columns for ordering
        all_output_columns = {**output_columns, **phase3_output_columns}
        
        # Also merge Phase 1 field output columns so they participate in letter-based ordering
        if isinstance(p1_fields, list):
            for f in p1_fields:
                letter = f.get('output_column')
                name = f.get('name')
                if letter and name:
                    all_output_columns[letter] = name
        
        # Build dynamic PRIMARY_COL_ORDER based on actual column names in the DataFrame
        # The first columns in primary_df are always: Unique_ID, Source_File_Name, primary key column, then field columns
        PRIMARY_COL_ORDER = []
        for col in primary_df.columns:
            PRIMARY_COL_ORDER.append(col)
            if len(PRIMARY_COL_ORDER) >= 4:
                # Stop after including Unique_ID, Source_File_Name, primary key, and first field column
                # This ensures Phase 1 columns are captured dynamically by actual names
                found_all_phase1_fields = True
                for f in p1_fields if isinstance(p1_fields, list) else []:
                    fname = f.get('name', '')
                    if fname and fname not in PRIMARY_COL_ORDER:
                        found_all_phase1_fields = False
                        break
                if found_all_phase1_fields:
                    break
        
        logger.info(f"Dynamic PRIMARY_COL_ORDER: {PRIMARY_COL_ORDER}")
        logger.info(f"All output columns (incl Phase 1): {all_output_columns}")
        
        ordered_col_names = []
        cols_with_letters = []
        other_cols = []
        
        for col in primary_df.columns:
            letter = None
            for l, n in all_output_columns.items():
                if n == col:
                    letter = l
                    break
            
            if letter:
                cols_with_letters.append((letter, col))
            elif col not in PRIMARY_COL_ORDER:
                other_cols.append(col)
        
        # Sort Phase 2/3 columns by Excel column letter order (A, B, C... Z, AA, AB... ZZ)
        cols_with_letters = sort_excel_columns(dict(cols_with_letters))
        
        # Build final column order:
        # 1. Primary columns first (A, B, C - dynamically from Phase 1)
        # 2. Phase 2/3 columns in letter order (D onwards)
        # 3. Any other columns last
        for pc in PRIMARY_COL_ORDER:
            if pc in primary_df.columns and pc not in ordered_col_names:
                ordered_col_names.append(pc)
        
        for letter, col in cols_with_letters:
            if col not in ordered_col_names:
                ordered_col_names.append(col)
        
        for col in other_cols:
            if col not in ordered_col_names:
                ordered_col_names.append(col)
        
        primary_df = primary_df[ordered_col_names]
        
        # Parse filename — use ORIGINAL uploaded file name (from Phase 1 config) 
        # instead of the auto-generated primary filename, because the generated name
        # loses month/year patterns due to timestamp suffixes.
        source_primary_filename = primary_generated_filename or "Unknown"
        
        # Try to get original source file name from Phase 1 config
        original_source_name = source_primary_filename
        try:
            p1_file_id = p1_config.get('file_id')
            if p1_file_id:
                conn = get_db_connection()
                orig_row = conn.execute(
                    "SELECT original_name FROM files WHERE id = ?", (p1_file_id,)
                ).fetchone()
                conn.close()
                if orig_row and orig_row['original_name']:
                    original_source_name = orig_row['original_name']
        except Exception as e:
            logger.warning(f"Could not fetch original file name for parsing: {e}")
        
        parsed = parse_filename(original_source_name)
        
        cid = processing_status.get("company_id")
        mid = processing_status.get("module_id")
        custom_filename = processing_status.get("custom_filename")
        
        # Override parsing if the user provided a custom filename with a month/type
        if custom_filename:
            custom_parsed = parse_filename(custom_filename)
            if custom_parsed.get('parsed'):
                parsed['month_name'] = custom_parsed.get('month_name', parsed.get('month_name'))
                parsed['month_number'] = custom_parsed.get('month_number', parsed.get('month_number'))
                parsed['year'] = custom_parsed.get('year', parsed.get('year'))
                parsed['financial_year'] = custom_parsed.get('financial_year', parsed.get('financial_year'))
                parsed['parsed'] = True
        
        # Determine final filename
        final_output_filename = None
        if custom_filename:
            import re
            from datetime import datetime, timezone, timedelta
            safe_name = re.sub(r'[^\w\s-]', '', custom_filename).strip()
            ist_tz = timezone(timedelta(hours=5, minutes=30))
            timestamp_ist = datetime.now(ist_tz).strftime("%d-%m-%Y_%I-%M-%p")
            final_output_filename = f"{safe_name}_{timestamp_ist}.xlsx"
        
        if parsed['parsed']:
            report_type = None
            month_name = parsed['month_name']
            year = parsed['year']
            financial_year = parsed['financial_year']
            
            output_filename = final_output_filename or generate_processed_filename(
                source_primary_filename, "Unknown", month_name, year
            )
            
            base_dir = get_physical_storage_path(PROCESSED_DIR, cid, mid)
            if not financial_year or not month_name:
                storage_path = os.path.join(base_dir, 'Unclassified')
            else:
                storage_path = os.path.join(base_dir, financial_year, month_name)
            
            os.makedirs(storage_path, exist_ok=True)
            output_path = os.path.join(storage_path, output_filename)
        else:
            if not final_output_filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                final_output_filename = f"Reconciliation_Result_{timestamp}.xlsx"
                
            output_filename = final_output_filename
            
            base_dir = get_physical_storage_path(PROCESSED_DIR, cid, mid)
            storage_path = os.path.join(base_dir, 'Others')
            os.makedirs(storage_path, exist_ok=True)
            output_path = os.path.join(storage_path, output_filename)
            
            report_type = 'Others'
            month_name = None
            year = None
            financial_year = None
        
        # Write Results sheet and Summary sheets - OPTIMIZED
        export_df = primary_df.copy()
        custom_primary_label = p1_config.get('column', 'Order ID')
        if custom_primary_label and custom_primary_label != 'Order ID':
            export_df = export_df.rename(columns={'Order ID': custom_primary_label})
        
        # --- NUMERIC FORMATTING: Convert and round numeric columns ---
        # Identify columns that appear to be numeric (from Phase 2 calculations, sumif, etc.)
        numeric_cols = set()
        for col in export_df.columns:
            if col in ('Order ID', custom_primary_label, 'Source_File_Name', 'Unique_ID'):
                continue
            sample_vals = export_df[col].dropna().head(20).astype(str)
            # Heuristic: if most non-empty values look numeric (digits, commas, dots, minus)
            numeric_count = sample_vals.str.match(r'^-?[\d,]+\.?\d*$').sum()
            if numeric_count > len(sample_vals) * 0.5:
                numeric_cols.add(col)
        
        # Convert identified numeric columns to float and round to 2 decimals
        for col in numeric_cols:
            export_df[col] = pd.to_numeric(
                export_df[col].astype(str).str.replace(',', '').str.strip(),
                errors='coerce'
            ).round(2)
        
        sheets_data = {'Results': len(export_df)}
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write main results with proper number formatting
            ws_results = writer.book.create_sheet('Results')
            
            # Write headers
            for col_idx, col_name in enumerate(export_df.columns, 1):
                cell = ws_results.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True, color='FFFFFF', size=10)
                cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(bottom=Side(style='medium', color='1F4E79'))
            
            # Write data rows with number formatting for numeric columns
            from openpyxl.utils.dataframe import dataframe_to_rows
            for row_idx, row_data in enumerate(dataframe_to_rows(export_df, index=False, header=False), 2):
                for col_idx, value in enumerate(row_data, 1):
                    col_name = export_df.columns[col_idx - 1]
                    cell = ws_results.cell(row=row_idx, column=col_idx, value=value)
                    if col_name in numeric_cols and isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Auto-fit columns
            _autofit_columns_fast(ws_results, sample_rows=50)
            
            # Separate summaries by output_mode
            shared_summaries = {}   # output_mode == 'summary_sheet'
            separate_summaries = {} # output_mode == 'separate_sheet'
            
            for summary_name, summary_data in summary_sheets.items():
                config = summary_data['config']
                output_mode = config.get('output_mode', 'summary_sheet')
                if output_mode == 'separate_sheet':
                    separate_summaries[summary_name] = summary_data
                else:
                    shared_summaries[summary_name] = summary_data
            
            # Write shared summaries to a single "Summary" sheet
            if shared_summaries:
                summary_ws = writer.book.create_sheet('Summary')
                sheets_data['Summary'] = sum(len(sd['data']) for sd in shared_summaries.values())
                sheets_data['Summary'] = sum(len(sd['data']) for sd in shared_summaries.values())
                current_row = 1
                
                # Main header
                summary_ws.cell(row=current_row, column=1, value='SUMMARY REPORTS')
                summary_ws.cell(row=current_row, column=1).font = Font(bold=True, size=16, color='1F4E79')
                current_row += 1
                
                summary_ws.cell(row=current_row, column=1, value=f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
                summary_ws.cell(row=current_row, column=1).font = Font(size=10, color='666666')
                current_row += 2
                
                for summary_name, summary_data in shared_summaries.items():
                    pivot_df = summary_data['data']
                    config = summary_data['config']
                    chart_type = config.get('chart_type', 'none')
                    chart_position = config.get('chart_position', 'right')
                    
                    start_row = current_row
                    summary_ws.cell(row=current_row, column=1, value=f'▓▓▓ {summary_name.upper()} ▓▓▓')
                    summary_ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color='1F4E79')
                    current_row += 1
                    
                    # OPTIMIZED: Use vectorized writing
                    data_start_row = current_row
                    write_pivot_to_worksheet_fast(summary_ws, pivot_df, current_row)
                    current_row += len(pivot_df) + 2
                    data_end_row = current_row - 2
                    
                    # Add chart if requested
                    if chart_type != 'none' and len(pivot_df) > 0:
                        if chart_position == 'below':
                            # Chart below the data table
                            chart_start_col = 1
                            chart_start_row = current_row + 1
                        else:
                            # Chart to the right (default)
                            chart_start_col = len(pivot_df.columns) + 2
                            chart_start_row = start_row + 1
                        
                        _add_excel_chart(
                            summary_ws, chart_type, pivot_df,
                            data_start_row, data_end_row,
                            chart_start_col, chart_start_row
                        )
                        
                        if chart_position == 'below':
                            current_row += 15  # Leave space for chart below
                    
                    current_row += 3
                
                # Fast auto-fit
                _autofit_columns_fast(summary_ws, sample_rows=50)
                
                summary_ws.page_setup.orientation = 'landscape'
                summary_ws.page_setup.fitToPage = True
                summary_ws.page_setup.fitToWidth = 1
            
            # Write separate summaries to their own individual sheets
            for summary_name, summary_data in separate_summaries.items():
                # Sanitize sheet name (Excel sheet names max 31 chars, no special chars)
                safe_name = summary_name[:31].replace('/', '-').replace('\\', '-')
                # Ensure unique sheet name
                base_name = safe_name
                sheet_suffix = 1
                while safe_name in writer.sheets:
                    safe_name = f"{base_name[:29]}_{sheet_suffix}"
                    sheet_suffix += 1
                
                sep_ws = writer.book.create_sheet(safe_name)
                pivot_df = summary_data['data']
                config = summary_data['config']
                chart_type = config.get('chart_type', 'none')
                chart_position = config.get('chart_position', 'right')
                
                # Write header
                sep_ws.cell(row=1, column=1, value=summary_name)
                sep_ws.cell(row=1, column=1).font = Font(bold=True, size=14, color='1F4E79')
                
                # Write data
                data_start_row = 3
                write_pivot_to_worksheet_fast(sep_ws, pivot_df, data_start_row)
                data_end_row = data_start_row + len(pivot_df) - 1
                
                # Add chart if requested
                if chart_type != 'none' and len(pivot_df) > 0:
                    if chart_position == 'below':
                        chart_start_col = 1
                        chart_start_row = data_end_row + 3
                    else:
                        chart_start_col = len(pivot_df.columns) + 2
                        chart_start_row = 2
                    
                    _add_excel_chart(
                        sep_ws, chart_type, pivot_df,
                        data_start_row, data_end_row,
                        chart_start_col, chart_start_row
                    )
                
                # Auto-fit
                _autofit_columns_fast(sep_ws, sample_rows=50)
                sep_ws.page_setup.orientation = 'landscape'
                sep_ws.page_setup.fitToPage = True
                sep_ws.page_setup.fitToWidth = 1
                
                sheets_data[safe_name] = len(pivot_df)
        
        elapsed = time.time() - start_time
        
        # Save to processed files database
        try:
            from database import save_processed_file
            sheets_data_json = json.dumps(sheets_data)
            # Get file size in MB
            file_size_mb = None
            try:
                file_size_bytes = os.path.getsize(output_path)
                file_size_mb = round(file_size_bytes / (1024 * 1024), 2)
            except:
                pass
            save_processed_file(
                filename=output_filename,
                file_path=output_path,
                report_type=report_type,
                financial_year=financial_year,
                month_name=month_name,
                month_number=parsed.get('month_number') if parsed.get('parsed') else None,
                year=year,
                source_primary_filename=source_primary_filename,
                total_rows=len(primary_df),
                rules_used=len(all_rules),
                sheets_data=sheets_data_json,
                file_size=file_size_mb,
                processing_time=f"{elapsed:.1f}s",
                company_id=cid,
                module_id=mid
            )
        except Exception as e:
            logger.error(f"Error saving processed file metadata: {e}")
        
        logger.info(f"Processing completed in {elapsed:.1f}s. {len(primary_df)} rows processed.")
        
        # India timezone for timestamp
        india_offset = timedelta(hours=5, minutes=30)
        india_tz = timezone(india_offset)
        processed_at_india = datetime.now(india_tz).strftime('%d/%m/%Y %I:%M:%S %p IST')
        
        with processing_lock:
            # Build summary message for user
            summary_msg = f"Processing completed in {elapsed:.1f}s. {len(primary_df)} rows processed."
            if phase4_errors:
                summary_msg += f" {len(summary_sheets)} summary generated, {len(phase4_errors)} skipped."
            elif summary_sheets:
                summary_msg += f" {len(summary_sheets)} summaries generated."
            
            from database import add_notification
            if phase4_errors:
                add_notification(cid, mid, 'warning', f"Processing completed with warnings: {len(phase4_errors)} summaries skipped due to missing columns.", "?page=processed", user_id=uid)
            else:
                add_notification(cid, mid, 'success', f"Processing completed successfully. {len(summary_sheets)} summaries generated.", "?page=processed", user_id=uid)
            
            processing_status["result"] = {
                "success": True,
                "message": summary_msg,
                "total_rules": len(all_rules),
                "rows_processed": len(primary_df),
                "processed_at": processed_at_india,
                "processed_file_path": output_path,
                "report_type": report_type,
                "month_name": month_name,
                "financial_year": financial_year,
                "processing_time": f"{elapsed:.1f}s",
                "status": "completed",
                "filtered": filter_sources is not None,
                "filter_count": len(filter_sources) if filter_sources else 0,
                "total_rows_before_filter": before_count if filter_sources else len(primary_df),
                "summary_count": len(summary_sheets),
                "phase4_errors": phase4_errors if phase4_errors else None
            }
            processing_status["progress"] = "completed"
            processing_status["is_processing"] = False
        
    except Exception as e:
        import traceback
        error_msg = f"Processing error: {str(e)}\n{traceback.format_exc()}"
        logger.error(error_msg)
        with processing_lock:
            from database import add_notification
            cid = processing_status.get("company_id")
            mid = processing_status.get("module_id")
            uid = processing_status.get("user_id")
            if cid and mid:
                add_notification(cid, mid, 'error', f"Processing failed: {str(e)[:100]}...", "?page=process", user_id=uid)
            processing_status["error"] = str(e)
            processing_status["result"] = {"success": False, "message": str(e)}
            processing_status["progress"] = "error"
            processing_status["is_processing"] = False

@app.get("/api/primary/source-files")
async def get_primary_source_files(current_user: Optional[dict] = Depends(get_optional_user)):
    """Get all unique Source_File_Name entries from the active primary data file"""
    try:
        cid, mid = _get_context(current_user)
        rules = get_rules_by_phase(1, company_id=cid, module_id=mid)
        if not rules:
            return {"success": False, "message": "No Phase 1 rule configured"}
        
        config = json.loads(rules[-1]['config'])
        if isinstance(config, str):
            try:
                config = json.loads(config)
            except:
                pass
        
        primary_file = config.get('primary_file') if isinstance(config, dict) else None
        
        if not primary_file:
            return {"success": False, "message": "No primary data file found"}
        
        primary_path = get_primary_file_path(primary_file, cid, mid)
        if not os.path.exists(primary_path):
            return {"success": False, "message": "Primary data file not found on disk"}
        
        df = read_primary_file(primary_file, cid, mid)
        if df is None or 'Source_File_Name' not in df.columns:
            return {"success": False, "message": "Primary data missing Source_File_Name column"}
        
        sources = df['Source_File_Name'].dropna().unique().tolist()
        sources = [str(v).strip() for v in sources if str(v).strip()]
        sources.sort()
        
        return {
            "success": True,
            "sources": sources,
            "total": len(sources)
        }
    except Exception as e:
        return {"success": False, "message": str(e)}


@app.post("/api/process")
async def process_all_rules(
    selected_source_files: Optional[str] = Form(None),
    custom_filename: Optional[str] = Form(None),
    force: bool = Form(False),
    current_user: Optional[dict] = Depends(get_optional_user)
):
    """Start processing in background with optional source file filter"""
    global processing_status
    
    # Parse selected source files if provided
    filter_sources = None
    if selected_source_files:
        try:
            filter_sources = json.loads(selected_source_files)
            if not isinstance(filter_sources, list):
                filter_sources = None
        except:
            filter_sources = None
    

    cid, mid = _get_context(current_user)
    
    # --- NEW: Pre-flight Validation ---
    if not force:
        try:
            conn = get_db_connection()
            if cid is not None and mid is not None:
                all_rules = conn.execute("SELECT * FROM rules WHERE company_id = ? AND module_id = ? ORDER BY phase, id", (cid, mid)).fetchall()
            else:
                all_rules = conn.execute("SELECT * FROM rules ORDER BY phase, id").fetchall()
            conn.close()
            
            p1 = [dict(r) for r in all_rules if r['phase'] == 1]
            p2 = [dict(r) for r in all_rules if r['phase'] == 2]
            p3 = [dict(r) for r in all_rules if r['phase'] == 3]
            p4 = [dict(r) for r in all_rules if r['phase'] == 4]
            
            p1_rules = p1[-1] if p1 else None
            p2_rules = p2[-1] if p2 else None
            p3_rules = p3[-1] if p3 else None
            
            generated_cols = {'Unique_ID', 'Source_File_Name', 'Order ID', 'Sales Amount'}
            
            if p1_rules:
                try:
                    c1 = json.loads(p1_rules['config'])
                    generated_cols.add(c1.get('column', 'Order ID'))
                    for f in c1.get('fields', []):
                        if f.get('name'): generated_cols.add(f['name'])
                except: pass
                
            if p2_rules:
                try:
                    c2 = json.loads(p2_rules['config'])
                    for r in c2:
                        if r.get('column_name'): generated_cols.add(r['column_name'])
                except: pass
                
            if p3_rules:
                try:
                    c3 = json.loads(p3_rules['config'])
                    for g in c3:
                        if g.get('column_name'): generated_cols.add(g['column_name'])
                except: pass
                
            required_cols = set()
            for r in p4:
                try:
                    c4 = json.loads(r['config'])
                    for f in c4.get('value_fields', []):
                        if f.get('column'): required_cols.add(f['column'])
                    for f in c4.get('row_fields', []):
                        if f: required_cols.add(f)
                    for f in c4.get('column_fields', []):
                        if f: required_cols.add(f)
                except: pass
                
            missing = required_cols - generated_cols
            if missing:
                return {
                    "success": False,
                    "type": "validation_warning",
                    "missing_columns": list(missing),
                    "message": "Validation warning"
                }
        except Exception as e:
            logger.error(f"Validation error: {e}")
            # Ignore validation errors and proceed
    # --- END Validation ---

    with processing_lock:
        if processing_status["is_processing"]:
            return {
                "success": False,
                "message": "Processing is already running. Please wait.",
                "status": "running",
                "progress": processing_status["progress"]
            }
        
        processing_status = {
            "is_processing": True,
            "result": None,
            "error": None,
            "start_time": time.time(),
            "progress": "starting",
            "filter_sources": filter_sources,
            "custom_filename": custom_filename,
            "company_id": cid,
            "module_id": mid,
            "user_id": current_user.get('user_id') if current_user else None
        }
    
    thread = threading.Thread(target=process_rules_background, daemon=True)
    thread.start()
    
    return {
        "success": True,
        "message": "Processing started in background",
        "status": "started",
        "progress": "starting",
        "filtered": filter_sources is not None,
        "filter_count": len(filter_sources) if filter_sources else 0
    }

@app.get("/api/process/status")
async def get_process_status():
    """Check the status of background processing"""
    global processing_status
    
    with processing_lock:
        status_copy = processing_status.copy()
    
    if status_copy["is_processing"]:
        return {
            "success": True,
            "status": "processing",
            "progress": status_copy["progress"],
            "message": f"Processing in progress: {status_copy['progress']}"
        }
    
    if status_copy["error"]:
        return {
            "success": False,
            "status": "error",
            "message": status_copy["error"],
            "result": status_copy["result"]
        }
    
    if status_copy["result"]:
        return {
            "success": True,
            "status": "completed",
            "result": status_copy["result"],
            "progress": "completed"
        }
    
    return {
        "success": True,
        "status": "idle",
        "progress": "idle",
        "message": "No processing has been started"
    }

@app.get("/api/download/{filename}")
async def download_file(filename: str, current_user: Optional[dict] = Depends(get_optional_user)):
    try:
        cid, mid = _get_context(current_user)
        file_path = os.path.join(get_physical_storage_path(UPLOAD_DIR, cid, mid), filename)
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")
        
        from starlette.responses import FileResponse as StarletteFileResponse
        return StarletteFileResponse(
            file_path, 
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Cache-Control": "no-cache"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============ PHASE 4: SUMMARY & PIVOT APIs ============

@app.post("/api/summary/save")
async def save_summary(
    config: str = Form(...), 
    name: str = Form(...), 
    processing_type: Optional[str] = Form('both'),
    current_user: Optional[dict] = Depends(get_optional_user)
):
    try:
        cid, mid = _get_context(current_user)
        conn = get_db_connection()
        conn.execute("DELETE FROM rules WHERE phase = 4 AND name = ? AND company_id = ? AND module_id = ?", (name, cid, mid))
        cursor = conn.execute(
            "INSERT INTO rules (name, phase, config, processing_type, company_id, module_id) VALUES (?, ?, ?, ?, ?, ?)",
            (name, 4, config, processing_type, cid, mid)
        )
        rule_id = cursor.lastrowid
        conn.commit()
        conn.close()
        # FIX BUG #5: Clear file metadata cache when summaries change
        clear_file_cache()
        return {"success": True, "summary_id": rule_id, "message": "Summary saved successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/summary/list")
async def list_summaries(current_user: Optional[dict] = Depends(get_optional_user)):
    try:
        cid, mid = _get_context(current_user)
        conn = get_db_connection()
        rules = conn.execute("SELECT * FROM rules WHERE phase = 4 AND company_id = ? AND module_id = ? ORDER BY id DESC", (cid, mid)).fetchall()
        conn.close()
        summaries = []
        for r in rules:
            try:
                cfg = json.loads(r['config'])
                summaries.append({
                    "id": r['id'],
                    "name": r['name'],
                    "config": cfg,
                    "created_at": r['created_at']
                })
            except:
                pass
        return {"success": True, "summaries": summaries}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/summary/{summary_id}")
async def delete_summary(summary_id: int, current_user: Optional[dict] = Depends(get_optional_user)):
    try:
        cid, mid = _get_context(current_user)
        conn = get_db_connection()
        conn.execute("DELETE FROM rules WHERE id = ? AND company_id = ? AND module_id = ?", (summary_id, cid, mid))
        conn.commit()
        conn.close()
        return {"success": True, "message": "Summary deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/summary/clear-all")
async def clear_all_summaries(current_user: Optional[dict] = Depends(get_optional_user)):
    """Delete ALL Phase 4 summary rules for current company and module."""
    try:
        cid, mid = _get_context(current_user)
        conn = get_db_connection()
        conn.execute("DELETE FROM rules WHERE phase = 4 AND company_id = ? AND module_id = ?", (cid, mid))
        conn.commit()
        conn.close()
        return {"success": True, "message": "All summaries cleared"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/summary/preview")
async def preview_summary(
    config: str = Form(...),
    current_user: Optional[dict] = Depends(get_optional_user)
):
    try:
        summary_config = json.loads(config)
        
        cid, mid = _get_context(current_user)
        conn = get_db_connection()
        if cid is not None and mid is not None:
            phase1_rules = conn.execute(
                "SELECT * FROM rules WHERE phase = 1 AND company_id = ? AND module_id = ? ORDER BY id DESC LIMIT 1",
                (cid, mid)
            ).fetchall()
        else:
            phase1_rules = conn.execute("SELECT * FROM rules WHERE phase = 1 ORDER BY id DESC LIMIT 1").fetchall()
        conn.close()
        
        if not phase1_rules:
            return {"success": False, "message": "No Phase 1 rule found. Please configure primary data first."}
        
        p1_config = json.loads(phase1_rules[0]['config'])
        primary_file = p1_config.get('primary_file')
        
        if not primary_file:
            return {"success": False, "message": "Primary file not found"}
        
        primary_path = get_primary_file_path(primary_file, cid, mid)
        if not os.path.exists(primary_path):
            return {"success": False, "message": "Primary data file not found"}
        
        df = read_primary_file(primary_file, cid, mid)
        preview_df = df.head(1000).copy()
        
        available_columns = list(preview_df.columns)
        
        row_fields = summary_config.get('row_fields', [])
        column_fields = summary_config.get('column_fields', [])
        value_fields = summary_config.get('value_fields', [])
        filter_fields = summary_config.get('filter_fields', [])
        
        all_needed = row_fields + column_fields + [v['column'] for v in value_fields]
        missing = [c for c in all_needed if c not in available_columns and c not in ['Unique_ID', 'Primary_Value']]
        
        # Identify value field columns (these need NUMERIC dummy data, not strings)
        value_col_names = set(v['column'] for v in value_fields)
        
        for col in all_needed:
            if col not in preview_df.columns and col != 'Unique_ID' and col != 'Primary_Value':
                import random
                col_lower = col.lower()
                # Check if this is a value field (needs numeric data for aggregation)
                if col in value_col_names:
                    preview_df[col] = [random.randint(100, 100000) for _ in range(len(preview_df))]
                elif col_lower in ['order_type']:
                    samples = ['Prepaid', 'COD', 'Return', 'Exchange']
                    preview_df[col] = [random.choice(samples) for _ in range(len(preview_df))]
                elif col_lower in ['payment_gateway']:
                    samples = ['PayU', 'Razorpay', 'Stripe', 'CCAvenue', 'Paytm']
                    preview_df[col] = [random.choice(samples) for _ in range(len(preview_df))]
                elif col_lower in ['courier_partner']:
                    samples = ['Bluedart', 'Delhivery', 'EcomExpress', 'DTDC', 'FedEx']
                    preview_df[col] = [random.choice(samples) for _ in range(len(preview_df))]
                elif col_lower in ['status']:
                    samples = ['Delivered', 'Shipped', 'Pending', 'Cancelled', 'Returned']
                    preview_df[col] = [random.choice(samples) for _ in range(len(preview_df))]
                elif col_lower in ['remark', 'remarks']:
                    samples = ['OK', 'Hold', 'Refund', 'Duplicate', 'Verified']
                    preview_df[col] = [random.choice(samples) for _ in range(len(preview_df))]
                elif col_lower in ['total', 'net_sales', 'cn', 'variance', 'cod_settlement', 'amount', 'grand_total', 'settlement']:
                    preview_df[col] = [random.randint(100, 100000) for _ in range(len(preview_df))]
                else:
                    preview_df[col] = [f"Sample_{i%10}" for i in range(len(preview_df))]
        
        for filt in filter_fields:
            col = filt.get('column', '')
            op = filt.get('operator', '')
            val = filt.get('value', '')
            
            if col and col in preview_df.columns and op:
                if op == 'equal_to':
                    preview_df = preview_df[preview_df[col].astype(str) == str(val)]
                elif op == 'not_equal_to':
                    preview_df = preview_df[preview_df[col].astype(str) != str(val)]
                elif op == 'blank':
                    preview_df = preview_df[preview_df[col].astype(str).str.strip() == '']
                elif op == 'greater_than':
                    try:
                        preview_df = preview_df[pd.to_numeric(preview_df[col], errors='coerce') > float(val)]
                    except:
                        pass
                elif op == 'smaller_than':
                    try:
                        preview_df = preview_df[pd.to_numeric(preview_df[col], errors='coerce') < float(val)]
                    except:
                        pass
        
        if not row_fields and not column_fields:
            return {"success": False, "message": "Please configure at least Row or Column fields"}
        
        if not value_fields:
            return {"success": False, "message": "Please configure at least one Value field"}
        
        values = [v['column'] for v in value_fields]
        aggfuncs = {v['column']: v.get('aggregation', 'sum') for v in value_fields}
        
        # Clean value columns to numeric to avoid string concatenation and float64 conversion errors
        for val_col in values:
            if val_col in preview_df.columns:
                preview_df[val_col] = pd.to_numeric(preview_df[val_col].astype(str).str.replace(',', '').str.strip(), errors='coerce').fillna(0)
                
        # Handle nulls in grouping columns to prevent row dropping
        for group_col in row_fields + column_fields:
            if group_col in preview_df.columns:
                s = preview_df[group_col].astype(str)
                s = s.replace(r'^\s*$', 'Blank', regex=True)
                s = s.replace(['nan', 'None', 'NaT', '<NA>'], 'Blank')
                s = s.where(s != 'Blank', 'Blank')
                preview_df[group_col] = s
                preview_df[group_col] = preview_df[group_col].fillna('Blank')
        
        pivot_kwargs = {
            'values': values,
            'aggfunc': aggfuncs,
            'margins': True,
            'margins_name': 'Grand Total'
        }
        
        if row_fields:
            pivot_kwargs['index'] = row_fields
        if column_fields:
            pivot_kwargs['columns'] = column_fields
        
        pivot_table = pd.pivot_table(preview_df, **pivot_kwargs)
        pivot_table = pivot_table.fillna(0)
        
        if not pivot_table.empty:
            num_cols = pivot_table.select_dtypes(include=['number']).columns
            if len(num_cols) > 0:
                non_zero_mask = (pivot_table[num_cols] != 0).any(axis=1)
                if isinstance(pivot_table.index, pd.MultiIndex):
                    try:
                        non_zero_mask = non_zero_mask | (pivot_table.index.get_level_values(0) == 'Grand Total')
                    except: pass
                else:
                    if 'Grand Total' in pivot_table.index:
                        non_zero_mask.loc['Grand Total'] = True
                pivot_table = pivot_table[non_zero_mask]
        
        if not pivot_table.empty:
            num_cols = pivot_table.select_dtypes(include=['number']).columns
            if len(num_cols) > 0:
                non_zero_mask = (pivot_table[num_cols] != 0).any(axis=1)
                if isinstance(pivot_table.index, pd.MultiIndex):
                    try:
                        non_zero_mask = non_zero_mask | (pivot_table.index.get_level_values(0) == 'Grand Total')
                    except: pass
                else:
                    if 'Grand Total' in pivot_table.index:
                        non_zero_mask.loc['Grand Total'] = True
                pivot_table = pivot_table[non_zero_mask]
        
        # Rename index and columns to custom primary column label if configured
        custom_primary_label = p1_config.get('column', 'Primary_Value')
        if custom_primary_label and custom_primary_label != 'Primary_Value':
            if isinstance(pivot_table.index, pd.MultiIndex):
                pivot_table.index.names = [custom_primary_label if n == 'Primary_Value' else n for n in pivot_table.index.names]
            elif pivot_table.index.name == 'Primary_Value':
                pivot_table.index.name = custom_primary_label
            
            if isinstance(pivot_table.columns, pd.MultiIndex):
                pivot_table.columns.names = [custom_primary_label if n == 'Primary_Value' else n for n in pivot_table.columns.names]
            elif pivot_table.columns.name == 'Primary_Value':
                pivot_table.columns.name = custom_primary_label

        pivot_flat = pivot_table.reset_index()
        
        if custom_primary_label and custom_primary_label != 'Primary_Value':
            pivot_flat.columns = [' '.join(col).strip().replace('Primary_Value', custom_primary_label) if isinstance(col, tuple) else str(col).replace('Primary_Value', custom_primary_label) for col in pivot_flat.columns.values]
        else:
            pivot_flat.columns = [' '.join(col).strip() if isinstance(col, tuple) else str(col) for col in pivot_flat.columns.values]
        
        # FIX: Always preserve Grand Total row in preview even when truncating
        total_rows = len(pivot_flat)
        if total_rows > 50:
            # Take first 49 data rows + last row (Grand Total)
            preview_df = pd.concat([pivot_flat.head(49), pivot_flat.tail(1)])
        else:
            preview_df = pivot_flat.head(50)
        
        preview_data = clean_nan_values(preview_df.to_dict(orient='records'))
        columns = list(pivot_flat.columns)
        
        chart_type = summary_config.get('chart_type', 'none')
        chart_image = None
        
        if chart_type != 'none' and len(preview_data) > 0:
            chart_image = generate_chart(pivot_flat, chart_type, row_fields, column_fields, value_fields)
        
        return {
            "success": True,
            "preview_data": preview_data,
            "columns": columns,
            "total_rows": len(pivot_flat),
            "chart_image": chart_image,
            "available_columns": available_columns,
            "message": f"Preview generated with {len(preview_data)} rows"
        }
        
    except Exception as e:
        import traceback
        return {"success": False, "message": str(e), "error_detail": traceback.format_exc()}

def _autofit_columns(ws):
    """Legacy auto-fit for compatibility"""
    _autofit_columns_fast(ws, sample_rows=100)

def _add_excel_chart(ws, chart_type, df, data_start_row, data_end_row, start_col, start_row):
    """
    Add an embedded Excel chart to the worksheet.
    Properly excludes Grand Total row from chart data.
    """
    try:
        # Find the actual last data row (excluding Grand Total)
        actual_end_row = data_end_row
        
        # Check if last row is Grand Total and adjust
        last_row_first_cell = ws.cell(row=data_end_row, column=1).value
        if last_row_first_cell and 'Grand Total' in str(last_row_first_cell):
            actual_end_row = data_end_row - 1
        
        # Ensure we have at least 1 data row
        if actual_end_row <= data_start_row:
            logger.warning("Not enough data rows for chart (only header or Grand Total)")
            return
        
        # Find the first numeric column after the first column
        numeric_cols = []
        for i, c in enumerate(df.columns):
            if i == 0:
                continue
            # Check column dtype in DataFrame
            if df[c].dtype in ['int64', 'float64', 'int32', 'float32']:
                numeric_cols.append(i)
            else:
                # Try to check if values are numeric
                try:
                    sample = df[c].head(5)
                    if pd.to_numeric(sample, errors='coerce').notna().any():
                        numeric_cols.append(i)
                except:
                    pass
        
        if not numeric_cols:
            # Fallback to second column
            numeric_cols = [1]
        
        data_col_idx = numeric_cols[0]  # 0-based index in df.columns
        data_col = data_col_idx + 1  # 1-based Excel column
        
        # For pie charts, exclude Grand Total from labels too
        pie_exclude_gt = chart_type == 'pie_chart' and actual_end_row < data_end_row
        
        if chart_type == 'bar_chart':
            chart = BarChart()
            chart.type = "col"
            chart.title = "Summary Bar Chart"
            chart.y_axis.title = 'Value'
            chart.x_axis.title = str(df.columns[0])
            chart.style = 10  # Professional style
            
            # Data reference (includes header row for title)
            data = Reference(ws, min_col=data_col, min_row=data_start_row, max_row=actual_end_row)
            # Categories (exclude header)
            cats = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=actual_end_row)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.series[0].graphicalProperties.line.noFill = True
            
            # Style the bars
            from openpyxl.chart.shapes import GraphicalProperties
            from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
            s1 = chart.series[0]
            s1.graphicalProperties.solidFill = "2563EB"
            
            ws.add_chart(chart, f"{get_column_letter(start_col)}{start_row}")
            
        elif chart_type == 'pie_chart':
            chart = PieChart()
            chart.title = "Summary Pie Chart"
            chart.style = 10
            
            # Exclude Grand Total from pie chart
            labels = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=actual_end_row)
            data = Reference(ws, min_col=data_col, min_row=data_start_row, max_row=actual_end_row)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            
            ws.add_chart(chart, f"{get_column_letter(start_col)}{start_row}")
            
        elif chart_type == 'line_chart':
            chart = LineChart()
            chart.title = "Summary Line Chart"
            chart.y_axis.title = 'Value'
            chart.x_axis.title = str(df.columns[0])
            chart.style = 10
            
            data = Reference(ws, min_col=data_col, min_row=data_start_row, max_row=actual_end_row)
            cats = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=actual_end_row)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # Style the line
            s1 = chart.series[0]
            s1.graphicalProperties.line.width = 25000  # 2.5pt in EMUs
            
            ws.add_chart(chart, f"{get_column_letter(start_col)}{start_row}")
            
    except Exception as e:
        logger.error(f"Chart embedding error: {e}")
        import traceback
        logger.error(traceback.format_exc())


def _style_chart_base(ax, title, xlabel=None, ylabel=None):
    """Apply professional corporate styling to chart axes"""
    # White background, no spines on top/right
    ax.set_facecolor('#ffffff')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#e5e7eb')
    ax.spines['bottom'].set_color('#e5e7eb')
    
    # Light horizontal gridlines only
    ax.yaxis.grid(True, linestyle='--', linewidth=0.5, color='#e5e7eb', alpha=0.9)
    ax.xaxis.grid(False)
    ax.set_axisbelow(True)
    
    # Title
    ax.set_title(title, fontsize=14, fontweight='bold', color='#1f2937', pad=15)
    
    # Labels
    if xlabel:
        ax.set_xlabel(xlabel, fontsize=11, color='#4b5563', fontweight='medium')
    if ylabel:
        ax.set_ylabel(ylabel, fontsize=11, color='#4b5563', fontweight='medium')
    
    # Tick styling
    ax.tick_params(axis='both', colors='#6b7280', labelsize=9)


def _render_professional_chart(ax, chart_type, x_values, y_values, x_label, y_label, title_suffix=''):
    """Core professional chart rendering used by both generate_chart and get_processed_chart"""
    import matplotlib.pyplot as plt
    title = f"{title_suffix}" if title_suffix else "Chart"
    
    if chart_type == 'pie_chart':
        # Hide axes for pie chart
        ax.axis('off')
        
        # Corporate color palette (blues, teals, greens, purples, oranges)
        colors = ['#2563eb', '#0ea5e9', '#10b981', '#8b5cf6', '#f59e0b', 
                  '#ec4899', '#6366f1', '#14b8a6', '#f97316', '#84cc16',
                  '#06b6d4', '#d946ef', '#a855f7', '#22c55e', '#eab308']
        
        # Combine small slices into "Other" if >8 slices
        if len(y_values) > 8:
            sorted_idx = y_values.argsort()[::-1]
            top_n = 7
            top_vals = y_values[sorted_idx[:top_n]]
            other_val = y_values[sorted_idx[top_n:]].sum()
            
            if other_val > 0:
                final_vals = list(top_vals) + [other_val]
                final_labels = [str(x_values[i]) for i in sorted_idx[:top_n]] + ['Other']
            else:
                final_vals = list(y_values)
                final_labels = list(x_values)
        else:
            final_vals = list(y_values)
            final_labels = list(x_values)
        
        # Create pie with professional styling
        wedges, texts, autotexts = ax.pie(
            final_vals, labels=final_labels, autopct='%1.1f%%',
            colors=colors[:len(final_vals)], startangle=90,
            pctdistance=0.75, labeldistance=1.15,
            textprops={'fontsize': 9, 'color': '#374151'},
            wedgeprops={'edgecolor': '#ffffff', 'linewidth': 2}
        )
        
        # Style percentage text
        for autotext in autotexts:
            autotext.set_color('#ffffff')
            autotext.set_fontsize(8)
            autotext.set_fontweight('bold')
        
        # Add center circle for donut effect (modern)
        centre_circle = plt.Circle((0, 0), 0.50, fc='#ffffff')
        ax.add_patch(centre_circle)
        
        # Total in center
        total = sum(final_vals)
        ax.text(0, 0.05, f'{total:,.0f}', ha='center', va='center', 
                fontsize=16, fontweight='bold', color='#1f2937')
        ax.text(0, -0.12, 'Total', ha='center', va='center', 
                fontsize=9, color='#6b7280')
        
        ax.set_title(title, fontsize=14, fontweight='bold', color='#1f2937', pad=20)
        
    elif chart_type == 'line_chart':
        _style_chart_base(ax, title, x_label, y_label)
        
        x_pos = range(len(x_values))
        color = '#2563eb'
        
        # Plot line with markers
        ax.plot(x_pos, y_values, color=color, linewidth=2.5, marker='o', 
                markersize=7, markerfacecolor='#ffffff', markeredgecolor=color, 
                markeredgewidth=2, zorder=3)
        
        # Area fill under line
        ax.fill_between(x_pos, y_values, alpha=0.12, color=color)
        
        # X-ticks
        ax.set_xticks(x_pos)
        ax.set_xticklabels(list(x_values), rotation=45, ha='right')
        
        # Y-axis formatting
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:,.0f}'))
        
    else:
        # Default BAR CHART
        _style_chart_base(ax, title, x_label, y_label)
        
        x_pos = range(len(x_values))
        colors = ['#2563eb' if i % 2 == 0 else '#0ea5e9' for i in range(len(x_values))]
        
        bars = ax.bar(x_pos, y_values, color=colors, width=0.6, 
                       edgecolor='#ffffff', linewidth=1.5, zorder=3)
        
        # Value labels on top of bars
        for bar, val in zip(bars, y_values):
            height = bar.get_height()
            if height != 0:
                ax.text(bar.get_x() + bar.get_width() / 2., height,
                        f'{val:,.0f}', ha='center', va='bottom', 
                        fontsize=8, color='#1f2937', fontweight='semibold')
        
        # X-ticks
        ax.set_xticks(x_pos)
        ax.set_xticklabels(list(x_values), rotation=45, ha='right')
        
        # Y-axis formatting
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x:,.0f}'))


def generate_chart(df, chart_type, row_fields, column_fields, value_fields):
    """Generate a professional chart image for the preview"""
    if not _MATPLOTLIB_AVAILABLE:
        logger.warning("Chart generation skipped - matplotlib not installed")
        return None
    try:
        import matplotlib.pyplot as plt
        import io
        import base64
        
        if chart_type == 'none' or not row_fields or len(row_fields) == 0:
            return None
            
        x_col = row_fields[0]
        if x_col not in df.columns:
            return None
            
        # Find best y column (numeric, not Grand Total)
        y_col = None
        for c in df.columns:
            if c not in row_fields and c != 'Grand Total':
                try:
                    if pd.to_numeric(df[c], errors='coerce').notna().any():
                        y_col = c
                        break
                except:
                    continue
        if not y_col and len(df.columns) > 1:
            y_col = [c for c in df.columns if c != x_col and c != 'Grand Total'][0]
        if not y_col:
            return None
        
        df_plot = df[df[x_col] != 'Grand Total'].copy()
        if len(df_plot) == 0:
            return None
            
        df_plot[y_col] = pd.to_numeric(df_plot[y_col], errors='coerce').fillna(0)
        
        # Sort by value descending for better visualization
        df_plot = df_plot.sort_values(by=y_col, ascending=False).head(20)
        
        x_values = df_plot[x_col].astype(str).tolist()
        y_values = df_plot[y_col].tolist()
        
        fig, ax = plt.subplots(figsize=(11, 6.5), dpi=120)
        title = f"Summary Visualization"
        
        _render_professional_chart(ax, chart_type, x_values, y_values, x_col, y_col, title)
        
        plt.tight_layout()
        
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=120, bbox_inches='tight', 
                    facecolor='#ffffff', edgecolor='none')
        buf.seek(0)
        img_base64 = base64.b64encode(buf.read()).decode('utf-8')
        plt.close(fig)
        
        return f"data:image/png;base64,{img_base64}"
        
    except Exception as e:
        logger.error(f"Chart generation error: {e}")
        return None

# ============ PROCESSED FILES APIs ============

@app.get("/api/processed/tree")
async def get_processed_tree_api(current_user: Optional[dict] = Depends(get_optional_user)):
    try:
        from database import get_processed_tree, get_processed_stats
        cid, mid = _get_context(current_user)
        tree = get_processed_tree(company_id=cid, module_id=mid)
        stats = get_processed_stats(company_id=cid, module_id=mid)
        return {
            "success": True,
            "tree": tree,
            "stats": stats
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def format_to_ist(created_at_str):
    """Convert SQLite timestamp string to India timezone formatted string."""
    if not created_at_str:
        return None
    try:
        # Parse the SQLite timestamp (assumed UTC or local system time)
        dt = datetime.strptime(created_at_str, '%Y-%m-%d %H:%M:%S')
        # SQLite stores without timezone; assume it's UTC and convert to IST
        dt = dt.replace(tzinfo=timezone.utc)
        ist_offset = timedelta(hours=5, minutes=30)
        ist_dt = dt.astimezone(timezone(ist_offset))
        return ist_dt.strftime('%d/%m/%Y %I:%M:%S %p IST')
    except Exception:
        return created_at_str

@app.get("/api/processed/files")
async def get_processed_files_list(
    financial_year: Optional[str] = None,
    report_type: Optional[str] = None,
    month_name: Optional[str] = None,
    current_user: Optional[dict] = Depends(get_optional_user)
):
    try:
        from database import get_processed_files
        cid, mid = _get_context(current_user)
        files = get_processed_files(company_id=cid, module_id=mid, financial_year=financial_year, report_type=report_type, month_name=month_name)
        # Format created_at to IST for each file
        for file in files:
            if 'created_at' in file and file['created_at']:
                file['created_at'] = format_to_ist(file['created_at'])
        return {
            "success": True,
            "files": files,
            "count": len(files)
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/processed/{file_id}")
async def delete_processed(file_id: int):
    try:
        from database import delete_processed_file
        file = delete_processed_file(file_id)
        if file and os.path.exists(file['file_path']):
            os.remove(file['file_path'])
        return {"success": True, "message": "Processed file deleted"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/processed/{file_id}/preview")
async def preview_processed_file(file_id: int):
    try:
        from database import get_processed_file_by_id
        file = get_processed_file_by_id(file_id)
        if not file:
            raise HTTPException(status_code=404, detail="Processed file not found")
        
        if not os.path.exists(file['file_path']):
            raise HTTPException(status_code=404, detail="File not found on disk")
        
        try:
            df = pd.read_excel(file['file_path'], sheet_name='Summary', header=None)
        except Exception:
            df = pd.read_excel(file['file_path'], sheet_name=0, header=None)
        
        sections = []
        current_section = None
        
        for idx, row in df.iterrows():
            val = str(row[0]) if pd.notna(row[0]) else ''
            if val.startswith('▓▓▓') and val.endswith('▓▓▓'):
                if current_section:
                    sections.append(current_section)
                current_section = {
                    'name': val.replace('▓▓▓', '').replace('▓▓▓', '').strip(),
                    'headers': [],
                    'data': []
                }
            elif current_section and not current_section['headers']:
                # Read headers - keep all columns including empty ones for alignment
                header_row = row.tolist()
                # Only keep non-empty headers, but remember their positions
                headers = []
                for c in header_row:
                    if pd.notna(c) and str(c).strip():
                        headers.append(str(c).strip())
                    else:
                        headers.append('')  # Keep empty placeholder for alignment
                # Remove trailing empty headers
                while headers and headers[-1] == '':
                    headers.pop()
                current_section['headers'] = headers
                current_section['header_count'] = len(headers)
            elif current_section:
                # Read data row aligned to headers
                row_list = row.tolist()
                data_row = []
                header_count = current_section.get('header_count', len(current_section['headers']))
                
                for i in range(header_count):
                    if i < len(row_list):
                        cell = row_list[i]
                        if pd.notna(cell):
                            # Format numeric values with comma and 2 decimals
                            if isinstance(cell, (int, float)):
                                if cell == int(cell):
                                    data_row.append(f"{cell:,.2f}")
                                else:
                                    data_row.append(f"{cell:,.2f}")
                            else:
                                data_row.append(str(cell))
                        else:
                            data_row.append('')  # Empty cell placeholder
                    else:
                        data_row.append('')  # Missing cell placeholder
                
                # Skip completely empty rows (but keep Grand Total rows even if some values are empty)
                is_grand_total = any('Grand Total' in str(v) for v in data_row if v)
                if is_grand_total or any(v != '' for v in data_row):
                    current_section['data'].append(data_row)
        
        if current_section:
            sections.append(current_section)
        
        return {
            "success": True,
            "file": {
                "id": file['id'],
                "filename": file['filename'],
                "source": file['source_primary_filename'],
                "created_at": file['created_at'],
                "total_rows": file['total_rows']
            },
            "sections": sections
        }
    except HTTPException as he:
        raise he
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def _safe_extract_chart_data(file_path):
    """
    Robustly extract chartable data from a processed file's Summary sheet.
    Handles multi-section sheets, merged cells, and various formats.
    Returns (headers, data_rows) or raises ValueError with descriptive message.
    """
    if not os.path.exists(file_path):
        raise ValueError("File not found on disk")

    # Load workbook to check sheet names
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
    except Exception as e:
        raise ValueError(f"Cannot open workbook: {e}")

    # Try Summary first, then any sheet that looks like a summary
    target_sheet = None
    if 'Summary' in sheet_names:
        target_sheet = 'Summary'
    else:
        # Fallback: find any sheet with data
        for sn in sheet_names:
            if sn != 'Results':
                target_sheet = sn
                break
    
    if not target_sheet:
        raise ValueError("No suitable sheet found for charting")

    # Read sheet
    try:
        df = pd.read_excel(file_path, sheet_name=target_sheet, header=None)
    except Exception as e:
        raise ValueError(f"Cannot read {target_sheet} sheet: {e}")

    if df.empty:
        raise ValueError(f"{target_sheet} sheet is empty")

    all_sections = []
    current_headers = None
    current_data = []
    in_section = False

    for idx, row in df.iterrows():
        val = str(row[0]) if pd.notna(row[0]) else ''

        # Detect section header like "▓▓▓ NAME ▓▓▓"
        if '▓▓▓' in val:
            # Save previous section if exists
            if current_headers and current_data:
                all_sections.append((current_headers, current_data))
            in_section = True
            current_headers = None
            current_data = []
            continue

        if not in_section:
            continue

        # First non-empty row after section header = column headers
        if current_headers is None:
            non_empty = [c for c in row if pd.notna(c) and str(c).strip() != '']
            if len(non_empty) == 0:
                continue
            current_headers = [str(c).strip() for c in row if pd.notna(c) and str(c).strip() != '']
            if len(current_headers) < 2:
                # Skip sections with too few columns
                current_headers = None
                continue
            continue

        # Data rows after headers - keep ALL cells (including empty) to preserve structure
        # Convert row to list, keeping NaN as None
        data_row = []
        for i in range(len(current_headers)):
            if i < len(row):
                v = row[i]
                if pd.notna(v):
                    data_row.append(v)
                else:
                    data_row.append(None)
            else:
                data_row.append(None)

        # Stop at Grand Total row (don't include it in chart data)
        first_val = str(data_row[0]) if data_row and data_row[0] is not None else ''
        if 'Grand Total' in first_val:
            break

        # Skip completely empty rows
        if any(v is not None for v in data_row):
            current_data.append(data_row)

    # Save last section
    if current_headers and current_data:
        all_sections.append((current_headers, current_data))

    if not all_sections:
        raise ValueError("No chartable sections found")

    # Pick the best section: prefer one with numeric columns and most data rows
    best_section = None
    best_score = -1
    for headers, data_rows in all_sections:
        if len(data_rows) == 0:
            continue
        # Score: number of data rows + bonus for numeric second column
        score = len(data_rows)
        if len(headers) > 1:
            second_col_vals = [r[1] for r in data_rows if r[1] is not None]
            if any(isinstance(v, (int, float)) for v in second_col_vals):
                score += 10
            else:
                # Try to parse as numeric
                try:
                    numeric_count = sum(1 for v in second_col_vals if v is not None and float(str(v).replace(',', '')) != 0)
                    if numeric_count > 0:
                        score += 5
                except:
                    pass
        if score > best_score:
            best_score = score
            best_section = (headers, data_rows)

    if not best_section:
        raise ValueError("No sections with chartable data found")

    headers, data_rows = best_section

    # Convert numeric strings to actual numbers
    cleaned_rows = []
    for row in data_rows:
        cleaned = []
        for i, val in enumerate(row):
            if val is None:
                cleaned.append(0 if i > 0 else '')  # First col keep empty, others 0
            elif isinstance(val, (int, float)):
                cleaned.append(val)
            else:
                s = str(val).strip().replace(',', '')
                try:
                    # Try numeric
                    if '.' in s:
                        cleaned.append(float(s))
                    else:
                        cleaned.append(int(s))
                except ValueError:
                    cleaned.append(val)
        cleaned_rows.append(cleaned)

    return headers, cleaned_rows


@app.get("/api/processed/{file_id}/chart")
async def get_processed_chart(file_id: int, chart_type: str = 'bar_chart'):
    try:
        from database import get_processed_file_by_id
        file = get_processed_file_by_id(file_id)
        if not file:
            return {"success": False, "message": "File not found in database", "chart_image": None}

        # Extract data safely
        try:
            headers, data_rows = _safe_extract_chart_data(file['file_path'])
        except ValueError as ve:
            logger.warning(f"Chart extraction failed for file_id={file_id}: {ve}")
            return {"success": False, "message": str(ve), "chart_image": None}

        # Build DataFrame with aligned dimensions
        chart_df = pd.DataFrame(data_rows, columns=headers)

        if not _MATPLOTLIB_AVAILABLE:
            return {
                "success": False,
                "message": "Chart generation unavailable - matplotlib not installed. Run: venv\\Scripts\\pip install matplotlib",
                "chart_image": None
            }

        import matplotlib.pyplot as plt
        import io
        import base64

        fig, ax = plt.subplots(figsize=(10, 6))

        x_col = headers[0]

        # Pick y_col: must be a different column from x_col, and preferably numeric
        y_col = None
        if len(headers) > 1:
            # Try to find a numeric column after the first one
            for candidate in headers[1:]:
                if candidate == x_col:
                    continue
                try:
                    numeric_test = pd.to_numeric(chart_df[candidate], errors='coerce')
                    if numeric_test.notna().any():
                        y_col = candidate
                        break
                except Exception:
                    continue
            # Fallback to second column if no numeric found
            if y_col is None:
                y_col = headers[1]

        if y_col is None or y_col == x_col:
            plt.close(fig)
            return {"success": False, "message": "Need at least 2 distinct columns to generate chart", "chart_image": None}

        # Convert y values safely
        y_values = pd.to_numeric(chart_df[y_col], errors='coerce').fillna(0)
        x_values = chart_df[x_col].astype(str)

        # Limit to top 20 items for readability
        if len(x_values) > 20:
            x_values = x_values.head(20)
            y_values = y_values.head(20)

        if chart_type == 'pie_chart' and y_values.sum() == 0:
            plt.close(fig)
            return {"success": False, "message": "All values are zero - cannot create pie chart", "chart_image": None}

        # Use professional chart rendering
        title = f"{file['filename']} - {chart_type.replace('_', ' ').title()}"
        _render_professional_chart(ax, chart_type, x_values.tolist(), y_values.tolist(), x_col, y_col, title)

        plt.tight_layout()

        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=120, bbox_inches='tight', facecolor='#ffffff', edgecolor='none')
        buf.seek(0)
        img_base64 = base64.b64encode(buf.read()).decode('utf-8')
        plt.close(fig)

        return {
            "success": True,
            "chart_image": f"data:image/png;base64,{img_base64}",
            "chart_type": chart_type,
            "message": f"Chart generated with {len(x_values)} data points"
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        logger.error(f"Unexpected chart error for file_id={file_id}: {e}\n{error_detail}")
        return {
            "success": False,
            "message": f"Chart generation failed: {str(e)}",
            "chart_image": None
        }

@app.get("/api/processed/{file_id}/data")
async def get_processed_raw_data(file_id: int, page: int = 1, limit: int = 100):
    try:
        from database import get_processed_file_by_id
        file = get_processed_file_by_id(file_id)
        if not file or not os.path.exists(file['file_path']):
            raise HTTPException(status_code=404, detail="File not found")
        
        try:
            df = pd.read_excel(file['file_path'], sheet_name='Results', header=0, dtype=str)
        except Exception:
            df = pd.read_excel(file['file_path'], sheet_name=0, header=0, dtype=str)
        
        # Identify numeric columns for comma formatting in dashboard view
        numeric_cols = set()
        for col in df.columns:
            if col in ('Order ID', 'Source_File_Name', 'Unique_ID'):
                continue
            sample_vals = df[col].dropna().head(20)
            numeric_count = sample_vals.str.match(r'^-?[\d,]+\.?\d*$').sum()
            if numeric_count > len(sample_vals) * 0.5:
                numeric_cols.add(col)
        
        total_rows = len(df)
        start_idx = (page - 1) * limit
        end_idx = start_idx + limit
        
        page_data = clean_nan_values(df.iloc[start_idx:end_idx].to_dict(orient='records'))
        
        # Format numeric values with comma and 2 decimals for dashboard display
        for row in page_data:
            for col in numeric_cols:
                if col in row and row[col] is not None:
                    try:
                        val = float(str(row[col]).replace(',', ''))
                        row[col] = f"{val:,.2f}"
                    except (ValueError, TypeError):
                        pass
        
        return {
            "success": True,
            "file": {
                "id": file['id'],
                "filename": file['filename'],
                "total_rows": total_rows
            },
            "page": page,
            "limit": limit,
            "total_pages": (total_rows + limit - 1) // limit,
            "data": page_data,
            "columns": df.columns.tolist()
        }
    except HTTPException as he:
        raise he
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/processed/{file_id}/download")
async def download_processed_file(file_id: int):
    try:
        from database import get_processed_file_by_id
        file = get_processed_file_by_id(file_id)
        if not file:
            raise HTTPException(status_code=404, detail="File not found")
        
        file_path = file['file_path']
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found on disk")
        
        from starlette.responses import FileResponse as StarletteFileResponse
        return StarletteFileResponse(
            file_path,
            filename=file['filename'],
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{file["filename"]}"',
                "Cache-Control": "no-cache"
            }
        )
    except HTTPException as he:
        raise he
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============ DASHBOARD APIs ============

@app.get("/api/dashboard/stats")
async def dashboard_stats(current_user: Optional[dict] = Depends(get_optional_user)):
    from database import get_db_connection, get_processed_stats
    conn = get_db_connection()
    
    cid, mid = _get_context(current_user)
    
    # Build conditional queries for multi-tenant filtering
    where_clause = ""
    params = []
    if cid is not None:
        where_clause += " WHERE company_id = ?"
        params.append(cid)
        if mid is not None:
            where_clause += " AND module_id = ?"
            params.append(mid)
    
    # Files count
    total_files = conn.execute(f"SELECT COUNT(*) FROM files{where_clause}", params).fetchone()[0]
    
    # Folders count
    total_folders = conn.execute(f"SELECT COUNT(*) FROM folders{where_clause}", params).fetchone()[0]
    
    # Masters count
    total_masters = conn.execute(f"SELECT COUNT(*) FROM master_files{where_clause}", params).fetchone()[0]
    
    # Rules count
    total_rules = conn.execute(f"SELECT COUNT(*) FROM rules{where_clause}", params).fetchone()[0]
    
    # Recent files
    recent_files = conn.execute(
        f"SELECT original_name, created_at FROM files{where_clause} ORDER BY created_at DESC LIMIT 5", params
    ).fetchall()
    
    conn.close()
    
    processed_stats = get_processed_stats(company_id=cid, module_id=mid)
    
    return {
        "success": True,
        "stats": {
            "total_files": total_files,
            "total_folders": total_folders,
            "total_masters": total_masters,
            "total_rules": total_rules,
            "recent_files": [dict(f) for f in recent_files],
            "processed": processed_stats,
            "processed_files": processed_stats.get("total_files", 0),
            "financial_years": processed_stats.get("financial_years", 0),
            "report_types": processed_stats.get("report_types", 0),
            "months": processed_stats.get("months", 0)
        }
    }

# ============ RECYCLE BIN APIs ============

@app.get("/api/recycle-bin")
async def get_recycle_bin(current_user: Optional[dict] = Depends(get_optional_user)):
    """Get all items in the recycle bin"""
    try:
        cid, mid = _get_context(current_user)
        items = get_recycle_bin_items(company_id=cid, module_id=mid)
        return {"success": True, "items": items, "count": len(items)}
    except Exception as e:
        logger.error(f"Get recycle bin error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/recycle-bin/{recycle_id}/restore")
async def restore_recycle_bin_item(recycle_id: int, current_user: Optional[dict] = Depends(get_optional_user)):
    """Restore an item from the recycle bin back to its original location"""
    try:
        restored = restore_from_recycle_bin(recycle_id)
        if restored is None:
            raise HTTPException(status_code=404, detail="Recycle bin item not found")
        
        cid, mid = _get_context(current_user)
        try:
            from database import save_audit_log
            save_audit_log(
                user_id=current_user.get('user_id') if current_user else None,
                action='RESTORE',
                entity_type=restored.get('type', 'unknown'),
                entity_id=restored.get('id'),
                details=f"Restored {restored.get('type')} '{restored.get('name')}' from recycle bin",
                company_id=cid,
                user_role=current_user.get('role') if current_user else None
            )
        except Exception as e:
            logger.warning(f"Audit log failed: {e}")
        
        clear_file_cache()
        return {
            "success": True,
            "message": f"Item restored successfully",
            "restored": restored
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Restore recycle bin error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/recycle-bin/{recycle_id}/permanent-delete")
async def permanent_delete_recycle_bin_item(recycle_id: int, current_user: Optional[dict] = Depends(get_optional_user)):
    """Permanently delete an item from the recycle bin (removes physical files too)"""
    try:
        item = get_recycle_bin_item(recycle_id)
        if not item:
            raise HTTPException(status_code=404, detail="Recycle bin item not found")
        
        cid, mid = _get_context(current_user)
        try:
            from database import save_audit_log
            save_audit_log(
                user_id=current_user.get('user_id') if current_user else None,
                action='PERMANENT_DELETE',
                entity_type=item.get('entity_type', 'unknown'),
                entity_id=item.get('entity_id'),
                details=f"Permanently deleted {item.get('entity_type')} '{item.get('entity_name')}' from recycle bin",
                company_id=cid,
                user_role=current_user.get('role') if current_user else None
            )
        except Exception as e:
            logger.warning(f"Audit log failed: {e}")
        
        permanent_delete_from_recycle_bin(recycle_id)
        clear_file_cache()
        return {"success": True, "message": "Item permanently deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Permanent delete recycle bin error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ GRACEFUL SHUTDOWN ============
def signal_handler(sig, frame):
    logger.info(f"Received signal {sig}, shutting down gracefully...")
    sys.exit(0)

signal.signal(signal.SIGINT, signal_handler)
signal.signal(signal.SIGTERM, signal_handler)

if __name__ == "__main__":
    import uvicorn
    logger.info("Starting Reconciliation Tool Server v2.0.0")
    uvicorn.run(app, host="0.0.0.0", port=5000)

# --- AUTO-SYNC ENDPOINTS ---

from fastapi import BackgroundTasks
from backend.auto_sync import trigger_folder_sync
from backend.database import get_files_with_sync_status

@app.post("/api/folders/{folder_id}/sync")
async def trigger_manual_sync(folder_id: int, background_tasks: BackgroundTasks, current_user: Optional[dict] = Depends(get_optional_user)):
    # This endpoint is called when the user clicks 'Sync Now' in the UI
    background_tasks.add_task(trigger_folder_sync, folder_id, False, current_user.get('user_id') if current_user else None)
    return {"success": True, "message": "Sync started in the background."}

@app.get("/api/folders/{folder_id}/sync-status")
async def get_folder_sync_status(folder_id: int):
    # Returns a list of files with their sync status
    files = get_files_with_sync_status(folder_id)
    return {"success": True, "files": files}


# ============ DEDUP / REJECTED-ARTEFACT APIs ============

@app.get("/api/master/{folder_id}/config-dedup")
async def api_get_dedup_config(folder_id: int, current_user: Optional[dict] = Depends(get_optional_user)):
    """Return the persisted dedup config for a folder so the frontend can re-hydrate the modal."""
    try:
        cfg = get_dedup_config(folder_id)
        if not cfg:
            return {"success": True, "exists": False, "config": {"dedup_enabled": False, "dedup_columns": [], "dedup_separator": " | "}}
        return {
            "success": True,
            "exists": True,
            "config": {
                "dedup_enabled":   bool(cfg.get('dedup_enabled')),
                "dedup_columns":   cfg.get('dedup_columns_list') or [],
                "dedup_separator": cfg.get('dedup_separator') or " | ",
            },
        }
    except Exception as e:
        logger.error(f"Get dedup config error for folder {folder_id}: {e}")
        return {"success": False, "message": str(e)}


@app.get("/api/master/{folder_id}/rejected-artefacts")
async def api_list_rejected_artefacts(folder_id: int, current_user: Optional[dict] = Depends(get_optional_user)):
    """List all reject reports for a folder, plus which file each came from."""
    try:
        items = list_rejected_artefacts(folder_id=folder_id)
        return {"success": True, "items": items, "count": len(items)}
    except Exception as e:
        logger.error(f"List rejected artefacts error: {e}")
        return {"success": False, "message": str(e)}


@app.get("/api/files/{file_id}/rejected-download")
async def api_download_rejected_artefact(file_id: int, current_user: Optional[dict] = Depends(get_optional_user)):
    """Stream the most-recent rejected artefact for a file. Falls back to 404 if no artefact exists."""
    try:
        artefact = get_latest_rejected_artefact_for_file(file_id)
        if not artefact or not artefact.get('artefact_path'):
            raise HTTPException(status_code=404, detail="No rejected artefact available for this file")
        artefact_path = artefact['artefact_path']
        if not os.path.exists(artefact_path):
            raise HTTPException(status_code=404, detail="Artefact file not found on disk")
        # Pick a sane Content-Type and download filename
        is_xlsx = artefact_path.lower().endswith('.xlsx')
        media_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            if is_xlsx else "text/csv"
        )
        download_name = os.path.basename(artefact_path)
        return FileResponse(
            artefact_path,
            media_type=media_type,
            filename=download_name,
            headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Rejected download error for file {file_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


def _enrich_files_with_rejected_artefact(files: list) -> list:
    """Helper: decorate each file row with the latest rejected artefact summary (if any).

    Behaviour:
      - Decorate ANY file that has a rejected_artefacts row in the DB (regardless
        of its current sync_status). This is the source of truth — the artefact
        table is what we trust, not the volatile sync_status field.
      - This fixes Bug 6: a file that was correctly rejected by dedup but is
        still in 'in_processing' (because the previous auto-sync run crashed
        before the post-dedup set_file_sync_status call) would otherwise be
        invisible to the pill render in the frontend.
    """
    for f in files:
        if not isinstance(f, dict):
            continue
        try:
            latest = get_latest_rejected_artefact_for_file(f.get('id'))
            if latest:
                f['rejected_artefact_id']    = latest.get('id')
                f['rejected_artefact_rows']  = latest.get('rejected_rows')
                f['rejected_artefact_total'] = latest.get('total_rows')
                f['rejected_download_url']   = f"/api/files/{f.get('id')}/rejected-download"
                # If the file has a rejected artefact on disk, ALWAYS show the
                # Rejected & Download pill — even if sync_status is stale
                # (in_processing/pending). The pill is a UI hint that the
                # rejected file is available for download.
                if f.get('sync_status') not in ('rejected',):
                    f['rejected_status_override'] = True
        except Exception:
            pass
    return files
